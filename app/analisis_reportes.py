import os
import sys
import locale
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl import Workbook
import re

COLUMNAS_REQUERIDAS = {
    'CLICODFAC': 2,
    'NOMBRE': 3,
    'URBANIZAC': 4,
    'CALLE': 5,
    'CLIMUNRO': 6,
    'MEDCODYGO': 7,
    'LECTURA': 8,
    'FECLEC': 9,
    'OBS1': 11,
    'OBS2': 12,
    'REFUBIME': 15,
    'NEWMED': 16,
    'CICLO': 18,
    'CARGA': 19,
    'ORDENRUTA': 20,
    'TIPOLECTURA': 21,
    'NOMBREOPERADOR': 26,
    'PROMEDIOSEDALIB': 34
}

# Intentar usar locale español para fechas
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')  # Linux/macOS
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')  # Windows
    except:
        pass

MESES_ES = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
    5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
    9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
}

def combinatoria_xlsx(dfs):
    # Datos fijos
    cols_fijas = ['CLICODFAC', 'NOMBRE', 'URBANIZAC', 'CALLE', 'CICLO', 'CLIMUNRO', 'MEDCODYGO']
    
    campos_variables_por_mes = [
        'LECTURA', 'OBS1', 'OBS2', 'REFUBIME', 'NEWMED',
        'CARGA', 'ORDENRUTA', 'TIPOLECTURA', 'NOMBREOPERADOR'
    ]

    meses_a_actualizar = set()

    # Extraer todos los meses y años de cada DataFrame nuevo que vas a subir
    for df in dfs:
        df['FECLEC'] = pd.to_datetime(df['FECLEC'], errors='coerce')
        df['MES'] = df['FECLEC'].dt.month
        df['ANIO'] = df['FECLEC'].dt.year
        meses_a_actualizar.update(set(zip(df['MES'], df['ANIO'])))

    # Diagnóstico antes de concatenar
    for i, df in enumerate(dfs):
        print(f"DataFrame {i}: total filas = {len(df)}, filas con CLICODFAC = {df['CLICODFAC'].notna().sum()}")
        print(f"DataFrame {i} valores únicos CLICODFAC (primeros 10):")
        print(df['CLICODFAC'].unique()[:10])

    df_all = pd.concat(dfs, ignore_index=True)
    cols_fijas = ['NOMBRE', 'URBANIZAC', 'CALLE', 'CICLO', 'CLIMUNRO', 'MEDCODYGO']
    base = df_all.dropna(subset=['CLICODFAC']).drop_duplicates(subset=['CLICODFAC']).set_index('CLICODFAC')[cols_fijas]

    # Aquí pones el bloque para debug
    print("Clientes únicos en df_all:", df_all['CLICODFAC'].nunique())
    print("Filas df_all:", df_all.shape[0])

    # Mostrar filas con valores nulos en cols_fijas
    print("Filas con valores nulos en columnas fijas:")
    print(df_all[df_all[cols_fijas].isnull().any(axis=1)][cols_fijas].head(10))

    # Procesar fechas
    df_all['FECLEC'] = pd.to_datetime(df_all['FECLEC'], errors='coerce')
    df_all['MES'] = df_all['FECLEC'].dt.month
    df_all['ANIO'] = df_all['FECLEC'].dt.year
    df_all['MES_NOMBRE'] = df_all['MES'].apply(lambda x: MESES_ES.get(x, ''))

    # Base con datos fijos
    base = df_all.dropna(subset=['CLICODFAC']).drop_duplicates(subset=['CLICODFAC']).set_index('CLICODFAC')[cols_fijas]
    data_final = base.copy()

    # Obtener lista ordenada de meses
    meses_ordenados = df_all[['MES', 'ANIO', 'MES_NOMBRE']].drop_duplicates()
    meses_ordenados = meses_ordenados.sort_values(by=['ANIO', 'MES'])

    # Agregar todos los campos variables por mes
    campos_importantes = ['LECTURA', 'OBS1']
    campos_restantes = ['OBS2', 'REFUBIME', 'NEWMED', 'CARGA', 'ORDENRUTA', 'TIPOLECTURA', 'NOMBREOPERADOR']

    # Agregar columnas variables por mes, primero campos importantes, luego restantes
    for _, row in meses_ordenados.iterrows():
        mes = row['MES']
        anio = row['ANIO']
        mes_nombre = row['MES_NOMBRE']
        df_mes = df_all[(df_all['MES'] == mes) & (df_all['ANIO'] == anio)]

        # Primero campos importantes
        for campo in campos_importantes:
            pivot = df_mes.pivot_table(index='CLICODFAC', values=campo, aggfunc='first')
            if not pivot.empty:
                pivot.columns = [f"{campo} {mes_nombre}"]
                col = pivot.columns[0]
                if col in data_final.columns:
                    data_final[col] = data_final[col].combine_first(pivot[col])
                else:
                    data_final = data_final.join(pivot, how='left')

        # Luego campos restantes
        for campo in campos_restantes:
            pivot = df_mes.pivot_table(index='CLICODFAC', values=campo, aggfunc='first')
            if not pivot.empty:
                pivot.columns = [f"{campo} {mes_nombre}"]
                col = pivot.columns[0]
                if col in data_final.columns:
                    data_final[col] = data_final[col].combine_first(pivot[col])
                else:
                    data_final = data_final.join(pivot, how='left')


    # --- PROMEDIOSEDALIB: Calcular promedio últimos 5 meses por cliente ---
    # Crear columna auxiliar año-mes para filtrar últimos 5 meses
    df_all['YM'] = df_all['FECLEC'].dt.to_period('M')
    meses_disponibles = df_all['YM'].dropna().unique()
    if len(meses_disponibles) > 0:
        mes_max = df_all['YM'].max()
        ultimos_5_meses = sorted([mes_max - i for i in range(5)])

        df_prom = df_all[df_all['YM'].isin(ultimos_5_meses)][['CLICODFAC', 'PROMEDIOSEDALIB', 'YM']]

        # Convertir PROMEDIOSEDALIB a numérico, ignorar errores
        df_prom['PROMEDIOSEDALIB'] = pd.to_numeric(df_prom['PROMEDIOSEDALIB'], errors='coerce')

        # Calcular promedio por CLICODFAC (ignorando NaN)
        prom_df = df_prom.groupby('CLICODFAC')['PROMEDIOSEDALIB'].mean().round(2).to_frame()

        # Unir al resultado final
        data_final = data_final.join(prom_df, how='left')
        data_final["DIFERENCIA DESVIACIÓN"] = data_final["PROMEDIOSEDALIB"] * 1.3
        print(">>> COLUMNA CREADA:", data_final[["PROMEDIOSEDALIB", "DIFERENCIA DESVIACIÓN"]].head())
    else:
        # Si no hay meses, solo añadir columna vacía
        data_final['PROMEDIOSEDALIB'] = None
        data_final["DIFERENCIA DESVIACIÓN"] = None

    



    ## --- Calcular CONSUMO/DF correctamente ---
    meses_ordenados_list = meses_ordenados.sort_values(by=['ANIO', 'MES']).reset_index(drop=True)

    for i in range(1, len(meses_ordenados_list)):
        mes_actual = meses_ordenados_list.loc[i, 'MES_NOMBRE']
        mes_anterior = meses_ordenados_list.loc[i - 1, 'MES_NOMBRE']

        lectura_actual_col = f"LECTURA {mes_actual}"
        lectura_anterior_col = f"LECTURA {mes_anterior}"

        if lectura_actual_col in data_final.columns and lectura_anterior_col in data_final.columns:
            print(f"Usando columnas: {lectura_anterior_col} → {lectura_actual_col}")

            # Asegurar valores numéricos
            lectura_actual = pd.to_numeric(data_final[lectura_actual_col], errors='coerce')
            lectura_anterior = pd.to_numeric(data_final[lectura_anterior_col], errors='coerce')

            consumo_df = lectura_actual - lectura_anterior

            print("Ejemplo de cálculo:")
            print(pd.DataFrame({
                lectura_anterior_col: lectura_anterior,
                lectura_actual_col: lectura_actual,
                'CONSUMO/DF': consumo_df
            }).head(10))

            # Borrar columna previa si existe
            if 'CONSUMO/DF' in data_final.columns:
                data_final.drop(columns=['CONSUMO/DF'], inplace=True)

            # Insertar CONSUMO/DF después de PROMEDIOSEDALIB (si existe)
            if 'PROMEDIOSEDALIB' in data_final.columns:
                idx = data_final.columns.get_loc('PROMEDIOSEDALIB')
                data_final.insert(idx + 1, 'CONSUMO/DF', consumo_df)
            else:
                print("La columna 'PROMEDIOSEDALIB' no se encontró. No se insertó 'CONSUMO/DF'.")

    data_final.reset_index(inplace=True)

    print("data_final.shape:", data_final.shape)
    print(data_final.head())
    #print(data_final.columns)

        # --- REORDENAR columnas para que queden primero pares LECTURA y OBS1 por mes ---
    # columnas fijas al inicio
    cols_fijas_orden = ['NOMBRE', 'URBANIZAC', 'CALLE', 'CICLO', 'CLIMUNRO', 'MEDCODYGO']
    
    # obtenemos los meses ordenados con su nombre (ya tienes meses_ordenados_list)
    meses_ordenados_list = meses_ordenados.sort_values(by=['ANIO', 'MES']).reset_index(drop=True)
    
    # Primero columnas pares LECTURA y OBS1 por mes
    cols_lectura_obs1 = []
    for i in range(len(meses_ordenados_list)):
        mes_nombre = meses_ordenados_list.loc[i, 'MES_NOMBRE']
        cols_lectura_obs1.append(f"LECTURA {mes_nombre}")
        cols_lectura_obs1.append(f"OBS1 {mes_nombre}")

    # Luego las otras columnas variables que mencionaste, para todos los meses
    otras_vars = ['OBS2', 'REFUBIME', 'NEWMED', 'CARGA', 'ORDENRUTA', 'TIPOLECTURA', 'NOMBREOPERADOR']
    cols_otras_vars = []
    for i in range(len(meses_ordenados_list)):
        mes_nombre = meses_ordenados_list.loc[i, 'MES_NOMBRE']
        for var in otras_vars:
            col_name = f"{var} {mes_nombre}"
            if col_name in data_final.columns:
                cols_otras_vars.append(col_name)
    
    # Finalmente, las columnas calculadas (PROMEDIOSEDALIB, CONSUMO/DF) si existen
    cols_calculadas = []
    for calc_col in ['PROMEDIOSEDALIB', 'DIFERENCIA DESVIACIÓN', 'CONSUMO/DF']:
        if calc_col in data_final.columns:
            cols_calculadas.append(calc_col)
    
    # Construimos la lista final de columnas
    cols_continuidad = ['CONTINUIDAD TOMA']
    columnas_finales = ['CLICODFAC'] + cols_fijas_orden + cols_lectura_obs1 + cols_otras_vars + cols_continuidad +  cols_calculadas + ['VARI%']
    columnas_finales = [col for col in columnas_finales if col in data_final.columns]


    # Reordenar columnas
    data_final = data_final[columnas_finales]

###################


    # === AGREGAR COLUMNA: FRECUENCIA LLEGADA ===
    carga_cols = [col for col in data_final.columns if col.startswith('CARGA ')]

    def obtener_frecuencia_llegada(row):
        meses_presentes = [col.replace('CARGA ', '') for col in carga_cols if pd.notna(row[col]) and str(row[col]).strip() != '']
        if not meses_presentes:
            return ''
        elif len(meses_presentes) == len(carga_cols):
            return 'COMPLETO'
        elif len(meses_presentes) == 1:
            return f"SOLO {meses_presentes[0]}"
        else:
            return f"SOLO {'-'.join(meses_presentes)}"

    data_final['FRECUENCIA LLEGADA'] = data_final.apply(obtener_frecuencia_llegada, axis=1)

    def evaluar_continuidad(row):
        # Detectar todas las columnas de tipo "LECTURA_"
        lecturas_cols = sorted([col for col in row.index if col.startswith("LECTURA ")])
        #print(row.index.tolist())

        
        # Convertir a booleanos (True si hay dato no nulo y no vacío)
        estados = []
        for col in lecturas_cols:
            val = row[col]
            tiene_valor = pd.notna(val) and str(val).strip() != ''
            estados.append(tiene_valor)
        
        cantidad_con_dato = sum(estados)

        if cantidad_con_dato == 0:
            return "SIN LECTURA"
        elif cantidad_con_dato == 1:
            return "PRIMERA LECTURA"
        elif all(estados):
            return "LECTURA CONTINUA"
        else:
            return "DISCONTINUO"

    # Aplicar la función al dataframe
    data_final["CONTINUIDAD TOMA"] = data_final.apply(evaluar_continuidad, axis=1)

    # === AGREGAR COLUMNA: CONTINUIDAD ANOMALIA ===
    obs1_cols = [col for col in data_final.columns if re.match(r'^OBS1\s+\w+', col)]

    def evaluar_cont_anomalia(row):
        valores = [str(row[col]).strip() for col in obs1_cols]

        # Si alguna columna está vacía o es "nan", entonces no es UNICA
        if any(v == "" or v.lower() == "nan" for v in valores):
            return "DISTINTA"

        # Si todos los valores son iguales
        if all(v == valores[0] for v in valores):
            return "UNICA"

        return "DISTINTA"


    data_final["CONTINUIDAD ANOMALIA"] = data_final.apply(evaluar_cont_anomalia, axis=1)

    def calcular_diferencia(row):
        prom = row.get("PROMEDIOSEDALIB")
        consumo = row.get("CONSUMO/DF")

        # Verifica si alguno es nulo o NaN
        if pd.isna(prom) or pd.isna(consumo):
            return ""
        
        # Si son iguales, retorna 0
        if prom == consumo:
            return 0
        
        # Retorna la diferencia
        return consumo - prom

    data_final["DIFERENCIA CONSUMO/PROM SEDA"] = data_final.apply(calcular_diferencia, axis=1)

    def calcular_diferencia(row):
        prom = row.get("PROMEDIOSEDALIB")
        consumo = row.get("CONSUMO/DF")

        # Verifica si alguno es nulo o NaN
        if pd.isna(prom) or pd.isna(consumo):
            return None  # <-- Cambiado de "" a None

        # Si son iguales, retorna 0
        if prom == consumo:
            return 0

        # Retorna la diferencia
        return consumo - prom


    data_final["DIFERENCIA CONSUMO/PROM SEDA"] = data_final.apply(calcular_diferencia, axis=1)


    def clasificar_diferencia(valor):
        if pd.isna(valor):
            return None
        if valor < 0:
            if valor >= -99:
                return "0-99 NEGATIVO"
            elif valor >= -499:
                return "100-500 NEGATIVO"
            elif valor >= -999:
                return "501-1000 NEGATIVO"
            else:
                return "1000 A MAS NEGATIVO"
        else:
            if valor <= 99:
                return "0-99 POSITIVO"
            elif valor <= 499:
                return "100-500 POSITIVO"
            elif valor <= 999:
                return "501-1000 POSITIVO"
            else:
                return "1000 A MAS POSITIVO"

    data_final["GRUPO COLUMNA DIF. CONS. PROM.SEDALIB"] = data_final["DIFERENCIA CONSUMO/PROM SEDA"].apply(clasificar_diferencia)


    def calcular_vari(row):
        prom = row.get("PROMEDIOSEDALIB")
        diferencia = row.get("DIFERENCIA CONSUMO/PROM SEDA")

        if pd.isna(prom) or pd.isna(diferencia) or prom == 0:
            return None

        vari = diferencia / prom
        vari_porcentaje = vari * 100
        
        # Si es entero, mostrar sin decimales
        if vari_porcentaje.is_integer():
            return f"{int(vari_porcentaje)}%"
        else:
            return f"{round(vari_porcentaje, 2)}%"

   
    data_final["DIFERENCIA CONSUMO/PROM SEDA"] = data_final.apply(calcular_diferencia, axis=1)
    data_final["VARI%"] = data_final.apply(calcular_vari, axis=1)


    def asignar_grupo(vari):
        if pd.isna(vari):
            return None
        if isinstance(vari, str) and vari.upper() == "CERO":
            return "CERO"
        
        # En caso vari venga con %, quitarlo y convertir a float
        if isinstance(vari, str) and vari.endswith('%'):
            try:
                vari = float(vari.rstrip('%'))
            except:
                return None

        # Asumimos que vari está en porcentaje numérico (ejemplo 30.5 significa 30.5%)
        # Ajustar si tus valores están en decimal (0.305)
        
        if vari < 0:
            if vari >= -29.99:
                return "0-29 NEGATIVO"
            elif vari >= -49.99:
                return "30-49 NEGATIVO"
            elif vari >= -100:
                return "50-100 NEGATIVO"
            elif vari >= -150:
                return "101-150 NEGATIVO"
            elif vari >= -300:
                return "151-300 NEGATIVO"
            elif vari >= -1000:
                return "301-1000 NEGATIVO"
            else:
                return "1001%-MAS NEGATIVO"
        else:
            if vari <= 29.99:
                return "0-29 POSITIVO"
            elif vari <= 49.99:
                return "30-49 POSITIVO"
            elif vari <= 100:
                return "50-100 POSITIVO"
            elif vari <= 150:
                return "101-150 POSITIVO"
            elif vari <= 300:
                return "151-300 POSITIVO"
            elif vari <= 1000:
                return "301-1000 POSITIVO"
            else:
                return "1001%-MAS POSITIVO"

    # Aplicar función para crear la columna 'GRUPO'
    data_final['GRUPO'] = data_final['VARI%'].apply(asignar_grupo)


    def condicion_prom(valor):
        if pd.isna(valor):
            return None
        if valor > 0:
            return "POSITIVO"
        elif valor < 0:
            return "NEGATIVO"
        else:
            return "CERO"

    data_final['CONDICION PROM'] = data_final['PROMEDIOSEDALIB'].apply(condicion_prom)

            
    return data_final


def exportar_excel_ajustado(df, ruta_archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Quitar líneas de cuadrícula
    ws.sheet_view.showGridLines = False
    
    # Inmovilizar primera fila
    ws.freeze_panes = "A2"

    # Estilos
    fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    borde_visible = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    sin_borde = Border(
        left=Side(border_style=None),
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None)
    )
    fuente_estandar = Font(name="Calibri", size=9)
    fuente_negrita = Font(name="Calibri", size=9, bold=True)
    alinear_izquierda = Alignment(horizontal='left')
    alinear_centro = Alignment(horizontal='center', vertical='center')
    
    # Escribir encabezados con alineación centrada en todas las columnas
    ws.row_dimensions[1].height = 35
    for col_idx, col_name in enumerate(df.columns, 1):
        letra_columna = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill_blanco
        cell.border = borde_visible
        cell.font = fuente_negrita
        # Colores personalizados de cabeceras
        cabecera = str(col_name).strip().upper()

        if "FRECUENCIA LLEGADA" in cabecera or "CONTINUIDAD ANOMALIA" in cabecera:
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")

        elif "CONTINUIDAD TOMA" in cabecera or "PROMEDIOSEDALIB" in cabecera:
            cell.fill = PatternFill(start_color="2eafe7", end_color="2eafe7", fill_type="solid")
            cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")

        elif "CONSUMO/DF" in cabecera or "DIFERENCIA CONSUMO/PROM SEDA" in cabecera:
            cell.fill = PatternFill(start_color="f3c242", end_color="f3c242", fill_type="solid")

        elif "DIFERENCIA DESVIACIÓN" in cabecera or "CONDICION PROM" in cabecera:
            cell.fill = PatternFill(start_color="f5ed64", end_color="f5ed64", fill_type="solid")

        elif "GRUPO COLUMNA DIF. CONS. PROM.SEDALIB" in cabecera or cabecera == "GRUPO":
            cell.fill = PatternFill(start_color="a7a7a7", end_color="a7a7a7", fill_type="solid")

        elif "VARI%" in cabecera:
            cell.fill = PatternFill(start_color="ff5a5a", end_color="ff5a5a", fill_type="solid")
            cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")

        cell.alignment = alinear_centro


    # Colores especiales para cabeceras A-G, LECTURA y OBS1
    fill_amarillo_AG = PatternFill(start_color="FFC333", end_color="FFC333", fill_type="solid")
    fill_lectura_header = PatternFill(start_color="c4deeb", end_color="c4deeb", fill_type="solid")
    fill_obs1_header = PatternFill(start_color="D6E6A0", end_color="D6E6A0", fill_type="solid")

    columnas_lectura = []
    columnas_obs1 = []

    # Reaplicar colores a cabeceras según reglas
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        # Cabeceras A-G amarillo ffc333
        if col_idx <= 7:
            cell.fill = fill_amarillo_AG
        # Cabeceras con "LECTURA"
        if "LECTURA" in str(col_name).upper():
            cell.fill = fill_lectura_header
            columnas_lectura.append(col_idx)
        # Cabeceras con "OBS1"
        elif "OBS1" in str(col_name).upper():
            cell.fill = fill_obs1_header
            columnas_obs1.append(col_idx)

    # Escribir datos
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            letra_columna = get_column_letter(col_idx)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill_blanco
            cell.border = sin_borde
            cell.font = fuente_estandar
            # Alineación: columnas A-G izquierda, H+ centro
            if col_idx <= 7:
                cell.alignment = alinear_izquierda
            else:
                cell.alignment = alinear_centro
                
    # Ahora sí, aplicar color y fuente a columnas LECTURA y OBS1 desde fila 2 en adelante
    fill_lectura_column = PatternFill(start_color="c4deeb", end_color="c4deeb", fill_type="solid")
    font_lectura_column = Font(color="000000", name="Calibri", size=9)
    fill_obs1_column = PatternFill(start_color="D6E6A0", end_color="D6E6A0", fill_type="solid")

    for fila in range(2, ws.max_row + 1):
        for col_idx in columnas_lectura:
            cell = ws.cell(row=fila, column=col_idx)
            cell.fill = fill_lectura_column
            cell.font = font_lectura_column
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in columnas_obs1:
            cell = ws.cell(row=fila, column=col_idx)
            cell.fill = fill_obs1_column
            cell.alignment = Alignment(horizontal='center', vertical='center')
            

    # Ajustar ancho de columnas
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(df[col].astype(str).map(len).max(), len(str(col)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Ocultar columnas
    columnas_a_ocultar = ['OBS2', 'REFUBIME', 'NEWMED', 'CARGA', 'ORDENRUTA', 'TIPOLECTURA', 'NOMBREOPERADOR']
    for idx, col in enumerate(df.columns, 1):
        for base in columnas_a_ocultar:
            if str(col).startswith(base):
                ws.column_dimensions[get_column_letter(idx)].hidden = True

    wb.save(ruta_archivo)

def main(folder_path):
    if not os.path.isdir(folder_path):
        print(f"Error: La carpeta '{folder_path}' no existe.")
        return

    archivos = [f for f in os.listdir(folder_path) if f.lower().endswith(('.xlsx', '.xls'))]
    if not archivos:
        print("No se encontraron archivos Excel en la carpeta.")
        return

    dfs = []
    for archivo in archivos:
        ruta_completa = os.path.join(folder_path, archivo)
        print(f"Leyendo archivo: {archivo}")
        try:
            df = pd.read_excel(ruta_completa, dtype=str, header=0)
        except Exception as e:
            print(f"Error al leer {archivo}: {e}")
            continue

        max_col = max(COLUMNAS_REQUERIDAS.values())
        
        if df.shape[1] < max_col:
            print(f"Archivo {archivo} no tiene suficientes columnas ({df.shape[1]} < {max_col}). Se omite.")
            continue

        cols_indices = [i - 1 for i in COLUMNAS_REQUERIDAS.values()]
        df_subset = df.iloc[:, cols_indices].copy()
        df_subset.columns = COLUMNAS_REQUERIDAS.keys()
        df_subset['FECLEC'] = pd.to_datetime(df_subset['FECLEC'], dayfirst=True, errors='coerce')

        dfs.append(df_subset)

    if not dfs:
        print("No se cargaron datos válidos de los archivos.")
        return

    print(f"Total archivos cargados: {len(dfs)}")

    # Ejecutar análisis
    resultado = combinatoria_xlsx(dfs)

    # Guardar resultado en Excel con estilo pizarra y columnas ocultas
    output_file = os.path.join(folder_path, f"resultado_combinado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    # Reordenar columnas según requerimiento
    columnas = list(resultado.columns)

    # Mapeo de posiciones
    reemplazos = {
        "PROMEDIOSEDALIB": "FRECUENCIA LLEGADA",
        "DIFERENCIA DESVIACIÓN": "CONTINUIDAD TOMA",
        "CONSUMO/DF": "CONTINUIDAD ANOMALIA",
        "FRECUENCIA LLEGADA": "PROMEDIOSEDALIB",
        "CONTINUIDAD TOMA": "CONSUMO/DF",
        "CONTINUIDAD ANOMALIA": "DIFERENCIA CONSUMO/PROM SEDA",
        "DIFERENCIA CONSUMO/PROM SEDA": "DIFERENCIA DESVIACIÓN",
    }

    # Obtener copia de columnas y aplicar reemplazos
    nuevas_columnas = columnas.copy()
    for original, nuevo in reemplazos.items():
        if original in nuevas_columnas and nuevo in columnas:
            idx = nuevas_columnas.index(original)
            nuevas_columnas[idx] = f"<<{nuevo}>>"  # marcamos temporalmente para evitar colisión

    # Ahora sustituimos los marcadores temporales por los nombres definitivos
    for i, col in enumerate(nuevas_columnas):
        if col.startswith("<<") and col.endswith(">>"):
            nuevas_columnas[i] = col.strip("<>")

    # Insertar columnas adicionales en posiciones específicas
    def insertar_despues(de_col, nueva_col):
        if nueva_col in nuevas_columnas:
            nuevas_columnas.remove(nueva_col)
        try:
            idx = nuevas_columnas.index(de_col)
            nuevas_columnas.insert(idx + 1, nueva_col)
        except ValueError:
            pass  # si no existe la columna base, no hacemos nada

    insertar_despues("GRUPO COLUMNA DIF. CONS. PROM.SEDALIB", "VARI%")
    insertar_despues("VARI%", "GRUPO")
    insertar_despues("GRUPO", "CONDICION PROM")

    # Reordenar el DataFrame
    resultado = resultado.reindex(columns=nuevas_columnas)


    exportar_excel_ajustado(resultado, output_file)
    print(f"Archivo generado: {output_file}")

        # Preguntar si desea cerrar el programa
    while True:
        respuesta = input("\n¿Deseas cerrar el programa? (si/no): ").strip().lower()
        if respuesta in ['si', 's']:
            print("Cerrando el programa...")
            break
        elif respuesta in ['no', 'n']:
            print("El programa seguirá abierto. Puedes realizar otra operación.")
            break
        else:
            print("Respuesta no válida. Por favor escribe 'si' o 'no'.")



if __name__ == '__main__':
    carpeta = input("Por favor, ingresa la ruta de la carpeta con los archivos Excel: ").strip()
    main(carpeta)
