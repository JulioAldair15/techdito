import os
from flask import Flask, after_this_request, render_template, request, redirect, url_for, flash, session, jsonify, send_file, current_app, send_from_directory, render_template, make_response
from .models import Usuario, Empleado, DataCatastroV2, RegistroTrabajo, EmpleadoLectura, EmpleadoDistribucion, EmpleadoInspecciones, EmpleadoCatastro, EmpleadoPersuasivas, EmpleadoMedidores, EmpleadoRecaudacion, EmpleadoAdministrativo, EmpleadoNorte, ReporteLectura, AuditoriaAcceso ,CargaDia, MaterialAsignado, CargaEjecutada, MaterialDevuelto,Remuneracion, DatosBancarios, BeneficioSocial, DocumentoEmpleado, Imagen, Categoria, Producto, Proveedor, Entrada, Salida, MovimientoDetalle, InventarioAuditoria, UnidadMedida, MatrizValidacion, Carta
from flask_bcrypt import check_password_hash 
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func, case
from app import app, db
from datetime import datetime, timedelta, date
from sqlalchemy import extract, or_, text
from calendar import monthrange
import traceback  # Importamos para imprimir detalles de errores
# from barcode import Code128
from openpyxl import Workbook 
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, PatternFill 
from collections import defaultdict
from openpyxl.utils import get_column_letter, column_index_from_string 
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.dataframe import dataframe_to_rows
from PyPDF2 import PdfMerger, PdfWriter, PdfReader
from collections import defaultdict
from openpyxl import load_workbook 
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter 
import locale
import calendar
import tempfile
import subprocess
import time
import re
import pandas as pd
import io
import re
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse
import json
from werkzeug.utils import secure_filename
from pdf2image import convert_from_path
from reportlab.lib.utils import ImageReader
import barcode
from barcode.writer import ImageWriter
from PIL import ImageFont
import zipfile
from dateutil import parser
from PIL import Image

# IVARGAS 11/07/2026
# ====================================
try:
    import pymupdf as fitz
except ImportError:
    import fitz
# ====================================

from fpdf import FPDF
import numpy as np
import shutil
import atexit
from zipfile import ZipFile
import requests
from requests.auth import HTTPBasicAuth
from functools import wraps
from zoneinfo import ZoneInfo
from requests_ntlm import HttpNtlmAuth
import os
import cv2
import zxingcpp
import decimal
import dbf
from xhtml2pdf import pisa
from openpyxl.utils import get_column_letter
import zxingcpp
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import aliased
import csv
from thefuzz import fuzz
import sys

# IVARGAS  11/07/2026
# ========================================
from google.cloud import storage
from google.oauth2.service_account import Credentials
# ========================================

from pdf2image import convert_from_path

app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

LIMA_TZ = ZoneInfo("America/Lima")
def get_timestamp():
    """Devuelve fecha/hora actual en Lima (UTC-5)."""
    return datetime.now(LIMA_TZ)

app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

# Lista global de archivos a eliminar al finalizar
archivos_a_borrar = []

def marcar_para_borrar(path):
    archivos_a_borrar.append(path)

@atexit.register
def limpiar_archivos():
    for path in archivos_a_borrar:
        try:
            if os.path.exists(path):
                os.remove(path)
                print(f"🧹 Eliminado temporal: {path}")
        except Exception as e:
            print(f"⚠️ Error al borrar archivo temporal: {e}")


# Caché global
INDEX_CACHE = {
    "timestamp_1": 0,
    "timestamp_2": 0,
    "data": []
}

# Rutas a tus archivos
INDEX_PATH_1 = "index_archivos_1.json"
INDEX_PATH_2 = "index_archivos_2.json"

def obtener_index_actualizado():
    global INDEX_CACHE

    # Obtener timestamps actuales
    ts_1 = os.path.getmtime(INDEX_PATH_1)
    ts_2 = os.path.getmtime(INDEX_PATH_2)

    # Si alguno cambió, recargar
    if ts_1 != INDEX_CACHE["timestamp_1"] or ts_2 != INDEX_CACHE["timestamp_2"]:
        print("[LOG] Cambios detectados en los archivos de índice. Recargando...")
        with open(INDEX_PATH_1, "r", encoding="utf-8") as f1, \
             open(INDEX_PATH_2, "r", encoding="utf-8") as f2:
            data = json.load(f1) + json.load(f2)
            INDEX_CACHE = {
                "timestamp_1": ts_1,
                "timestamp_2": ts_2,
                "data": data
            }
    else:
        print("[LOG] Índice en caché utilizado (sin cambios en archivos).")

    return INDEX_CACHE["data"]

### AUDITORIA ###
def registrar_evento(user_id, usuario, evento, modulo=None, duracion_sesion=None):
    try:
        evento_obj = AuditoriaAcceso(
            id_usuario_a=user_id,
            usuario=usuario,
            evento=evento,
            modulo=modulo,
            duracion_sesion=duracion_sesion,
            timestamp=get_timestamp()
        )
        db.session.add(evento_obj)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error registrando evento: {e}")


def log_evento(nombre_modulo):
    def decorador(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if 'user_id' in session:
                registrar_evento(
                    user_id=session['user_id'],
                    usuario=session['user_name'],
                    evento='acceso_modulo',
                    modulo=nombre_modulo
                )
            return f(*args, **kwargs)
        return wrapper
    return decorador

# Ruta raíz que redirige a la pantalla de login
@app.route('/')
def home():
    return redirect(url_for('login'))


@app.route('/add-or-update-imagenes', methods=['POST'])
def insertar_actualizar_imagenes():
    
    if not request.is_json:
        return jsonify({ "success": False, "message": "El contenido debe ser JSON" }), 400
    
    data = request.get_json()
    imagenes = data.get("imagenes")

    if not imagenes or not isinstance(imagenes, list):
        return jsonify({ "success": False, "message": "El campo 'imagenes' debe ser una lista" }), 400
    
    sql = text("""
        INSERT INTO imagenes (carpeta, filename, path, leyenda, origen)
        VALUES (:carpeta, :filename, :path, :leyenda, :origen)
        ON DUPLICATE KEY UPDATE
            path = VALUES(path),
            leyenda = VALUES(leyenda),
            origen = VALUES(origen),
            updated_at = CURRENT_TIMESTAMP
    """)

    try:
        for img in imagenes:
            if not img.get("carpeta") or not img.get("filename"):
                continue

            db.session.execute(sql, {
                "carpeta": img["carpeta"],
                "filename": img["filename"],
                "path": img.get("path"),
                "leyenda": img.get("leyenda"),
                "origen": img.get("origen")
            })

        db.session.commit()

        return jsonify({ "success": True, "total": len(imagenes) }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({ "success": False, "message": str(e) }), 500


"""""""""""
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = request.form['user']
        password = request.form['password']
        usuario = Usuario.query.filter_by(user=user).first()

        if usuario and check_password_hash(usuario.contraseña, password):
            # Hacer la sesión permanente
            session.permanent = True

            # Guardar datos en sesión
            session['user_id'] = usuario.id_usuario
            session['user_name'] = usuario.user
            session['login_time'] = get_timestamp().isoformat()

            if usuario.empleado:
                nombre_real = usuario.empleado.nombres or ''
                session['nombre_completo'] = nombre_real.strip()
            else:
                session['nombre_completo'] = usuario.user

            # Auditoría login
            registrar_evento(usuario.id_usuario, usuario.user, 'login')

            return redirect(url_for('inicio'))

        flash('Usuario o contraseña incorrectos', 'error')
        return redirect(url_for('login'))

    return render_template('login.html')
"""""

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # 1. Detectar si la petición es JSON (App Móvil) o Formulario (Web)
        es_app_movil = request.is_json

        if es_app_movil:
            data = request.get_json()
            username = data.get('user')
            password = data.get('password')
        else:
            username = request.form.get('user')
            password = request.form.get('password')

        # 2. Consulta a la tabla 'usuario'
        usuario = Usuario.query.filter_by(user=username).first()

        # 3. Verificación de credenciales
        if usuario and check_password_hash(usuario.contraseña, password):
            
            # --- RESTRICCIÓN PARA OPERARIOS EN LA WEB ---
            # Si el tipo es operario y NO viene de la app, bloqueamos.
            if usuario.tipousu == 'operario' and not es_app_movil:
                flash('Acceso denegado: Los operarios solo pueden ingresar mediante la App Móvil.', 'error')
                return redirect(url_for('login'))

            # 4. Configuración de sesión (Común para ambos)
            session.permanent = True
            session['user_id'] = usuario.id_usuario
            session['user_name'] = usuario.user
            session['login_time'] = datetime.now().isoformat()
            session['rol'] = usuario.rol

            # Obtener nombre del empleado mediante la relación backref 'empleado'
            if usuario.empleado:
                # Si es None, asignamos una cadena vacía "" para evitar que se imprima la palabra "None"
                nombres = usuario.empleado.nombres or ""
                apellidos = usuario.empleado.apellidos or ""
                
                # Juntamos ambos y eliminamos espacios extra al inicio/final o en medio
                nombre_display = f"{nombres} {apellidos}".strip()
                
                # Por seguridad: si por alguna razón ambos estaban vacíos, guardamos el nombre de usuario
                session['nombre_completo'] = nombre_display if nombre_display else usuario.user
            else:
                session['nombre_completo'] = usuario.user

            # Auditoría (si tienes la función)
            registrar_evento(usuario.id_usuario, usuario.user, 'login')

            # 5. Respuesta diferenciada
            if es_app_movil:
                return jsonify({
                    "success": True,
                    "message": "Login correcto",
                    "user_id": usuario.id_usuario,
                    "user_name": usuario.user,
                    "nombres": session.get('nombre_completo')
                }), 200
            
            return redirect(url_for('inicio'))

        # 6. Manejo de errores de credenciales
        if es_app_movil:
            return jsonify({"success": False, "message": "Usuario o clave incorrectos"}), 401
            
        flash('Usuario o contraseña incorrectos', 'error')
        return redirect(url_for('login'))

    return render_template('login.html')



# Ruta para la pantalla inicial después de iniciar sesión
@app.route('/inicio')
@log_evento("pantalla_inicio")
def inicio():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    return render_template(
        'inicio.html',
        user_name=session.get('user_name'),
        nombre_completo=session.get('nombre_completo')
    )


# Ruta para cerrar sesión
@app.route('/logout')
def logout():
    user_id = session.get('user_id')
    usuario = session.get('user_name')
    login_time = session.get('login_time')

    if user_id and usuario and login_time:
        try:
            inicio = datetime.fromisoformat(login_time)
            fin = get_timestamp()
            duracion = int((fin - inicio).total_seconds())
            registrar_evento(user_id, usuario, 'logout', duracion_sesion=duracion)
        except Exception as e:
            print(f'Error al registrar logout: {e}')

    session.clear()
    return redirect(url_for('login'))

@app.before_request
def actualizar_timeout():
    session.modified = True

@app.route('/registrar-modulo', methods=['POST'])
def registrar_modulo():
    if 'user_id' in session:
        data = request.get_json()
        modulo = data.get('modulo', 'desconocido')
        detalle = data.get('detalle', None)

        registrar_evento(
            user_id=session['user_id'],
            usuario=session['user_name'],
            evento='acceso_modulo',
            modulo=modulo + (f" ({detalle})" if detalle else "")
        )
        return jsonify({'status': 'ok'})
    return jsonify({'status': 'no_session'}), 401


@app.route('/auditar-cambio-pasajes', methods=['POST'])
def auditar_cambio_pasajes():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        tipo = data.get('tipo')  # 'monto', 'select', 'viaticos', 'ruta'
        valor_anterior = data.get('valor_anterior')
        nuevo_valor = data.get('nuevo_valor')

        if not id_empleado:
            return jsonify({'success': False, 'message': 'ID de empleado faltante'}), 400

        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        if 'user_id' in session:
            detalle = (
                f"Empleado: {nombre_empleado} | "
                f"Tipo: {tipo} | "
                f"Anterior: {valor_anterior if valor_anterior else 'VACÍO'} | "
                f"Nuevo: {nuevo_valor if nuevo_valor else 'VACÍO'}"
            )

            evento = {
                'monto': 'modificar_pasajes_monto',
                'select': 'modificar_pasajes_select',
                'viaticos': 'modificar_viaticos',
                'ruta': 'modificar_ruta',
                'estado': 'modificar_asistencia'
            }.get(tipo, 'modificacion_desconocida')

            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento=evento,
                modulo=detalle
            )

        return jsonify({'success': True})

    except Exception as e:
        print("❌ Error en auditar_cambio_pasajes:", str(e))
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/auditar-agregar-empleado', methods=['POST'])
def auditar_agregar_empleado():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')

        if not id_empleado:
            return jsonify({'success': False, 'message': 'ID de empleado faltante'}), 400

        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='agregar_empleado_asistencia',
                modulo=f"Empleado añadido: {nombre_empleado}"
            )

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/auditar-guardar-asistencia', methods=['POST'])
def auditar_guardar_asistencia():
    try:
        data = request.get_json()
        fecha = data.get('fecha')

        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='guardar_asistencia',
                modulo=f"Asistencia guardada para la fecha: {fecha}"
            )

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/auditar-busqueda-fecha', methods=['POST'])
def auditar_busqueda_fecha():
    data = request.get_json()
    fecha_inicio = data.get('fecha_inicio')
    fecha_fin = data.get('fecha_fin')
    area = data.get('area')
    accion = data.get('accion', 'buscar')

    usuario_id = session.get('user_id')
    usuario_nombre = session.get('user_name')

    if not usuario_id or not usuario_nombre:
        return jsonify({"status": "error", "message": "Usuario no identificado"}), 400

    if accion == "buscar":
        evento = f"busqueda_reporte_{area}"
        modulo = f"Fechas: {fecha_inicio} a {fecha_fin}"
    elif accion == "descargar_actual":
        evento = f"descargar_reporte_actual_{area}"
        modulo = f"Fechas: {fecha_inicio} a {fecha_fin}"
    elif accion == "descargar_completo":
        evento = f"descargar_planilla_completa_{area}"
        modulo = f"Fechas: {fecha_inicio} a {fecha_fin}"
    else:
        evento = f"accion_desconocida_{area}"
        modulo = f"Fechas: {fecha_inicio} a {fecha_fin}"

    registrar_evento(
        user_id=usuario_id,
        usuario=usuario_nombre,
        evento=evento,
        modulo=modulo
    )

    return jsonify({"status": "ok"})




# MÓDULO RECAUDACIÓN
@app.route('/filtrar-empleados', methods=['GET'])
@log_evento("acceso_modulo")
def filtrar_empleado():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'RECAUDACION', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en RECAUDACION"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/añadir-empleados', methods=['GET'])
def filtrar_empleados_añadir():
    # Filtrar todos los empleados sin especificar el área
    empleados = Empleado.query.filter(Empleado.estado != 'CESADO').all()
    
    # Convertir los datos a JSON
    empleados_data = [
        {
            "id_empleado": empleado.id_empleado,
            "dni": empleado.dni,
            "nombres": empleado.nombres,
            "cargo": empleado.cargo,
            "cod_ope": empleado.cod_ope
        }
        for empleado in empleados
    ]
    
    return jsonify(empleados_data)

@app.route('/cargar-asistencia', methods=['GET'])
def cargar_asistencia():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: Si ya hay registros en la fecha seleccionada, 
        # hacemos JOIN con Empleado para traer solo a los que NO están cesados.
        # ====================================================================
        registros_asistencia = EmpleadoRecaudacion.query.join(Empleado).filter(
            EmpleadoRecaudacion.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_recaudacion": registro.id_recaudacion,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": registro.pasajes if registro.pasajes is not None else "",
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Si no hay registros (hoja nueva), filtramos a los 
        # empleados de RECAUDACION agregando la condición estado != 'CESADO'
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoRecaudacion.id_empleado).filter(
                    EmpleadoRecaudacion.mes == mes_consulta
                )
            ),
            Empleado.area == 'RECAUDACION',
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500


@app.route('/guardar-asistencia-detalle', methods=['POST'])
def guardar_asistencia_detalle():
    try:
        data = request.get_json()
        print("Datos recibidos:", data)  # Depuración

        if "asistencias" not in data:
            return jsonify({'success': False, 'message': "Clave 'asistencias' no encontrada en JSON"}), 400

        asistencias = data['asistencias']

        for asistencia in asistencias:
            print("Asistencia procesada:", asistencia)

            if "fecha" not in asistencia:
                return jsonify({'success': False, 'message': "Clave 'fecha' no encontrada en asistencia"}), 400

            fecha = asistencia['fecha']
            id_empleado = asistencia['id_empleado']

            # Buscar si ya existe un registro de asistencia para ese empleado en esa fecha
            asistencia_existente = EmpleadoRecaudacion.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado)
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes) if asistencia.get('pasajes') else None
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()
                if empleado_original:
                    nuevo_registro = EmpleadoRecaudacion(
                        id_empleado=empleado_original.id_empleado,
                        dni=empleado_original.dni,
                        nombres=empleado_original.nombres,
                        cargo=empleado_original.cargo,
                        area=empleado_original.area,
                        cod_ope=empleado_original.cod_ope,
                        mes=asistencia['mes'],
                        fec_asist=fecha,
                        estado=asistencia.get('estado', ''),
                        pasajes=asistencia.get('pasajes', None),  # Ahora permite guardar "PR"
                        ruta=asistencia.get('ruta', ''),
                        viaticos=asistencia.get('viaticos', 0)
                    )
                    db.session.add(nuevo_registro)

        db.session.commit()
        return jsonify({'success': True, 'message': 'Registros actualizados correctamente en EmpleadoRecaudacion.'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al guardar la asistencia: {str(e)}'})

@app.route('/eliminar-asistencia', methods=['POST'])
def eliminar_asistencia():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro en la base de datos
        asistencia = EmpleadoRecaudacion.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Obtener nombre del empleado desde la tabla principal (Empleado)
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # ✅ Registrar en la auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Recaudación | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})



# MÓDULO TOMA DE ESTADO
@app.route('/filtrar-empleados-lectura', methods=['GET'])
def filtrar_empleado_lectura():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'TOMA DE ESTADO', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en TOMA DE ESTADO"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/cargar-asistencia-lectura', methods=['GET'])
def cargar_asistencia_lectura():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: JOIN con Empleado para excluir a los cesados en 
        # registros ya existentes.
        # ====================================================================
        registros_asistencia = EmpleadoLectura.query.join(Empleado).filter(
            EmpleadoLectura.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_lectura": registro.id_lectura,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": registro.pasajes if registro.pasajes is not None else "",
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Filtramos empleados de TOMA DE ESTADO y DISTRIBUCION
        # agregando la condición Empleado.estado != 'CESADO' para listas nuevas.
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoLectura.id_empleado).filter(
                    EmpleadoLectura.mes == mes_consulta
                )
            ),
            Empleado.area.in_(['TOMA DE ESTADO', 'DISTRIBUCION']),
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500


@app.route('/guardar-asistencia-detalle-lectura', methods=['POST'])
def guardar_asistencia_detalle_lectura():
    try:
        data = request.get_json()
        print("\n🔍 JSON recibido:", data)  # ✅ Depuración

        if not data or "asistencias" not in data:
            mensaje_error = "Clave 'asistencias' no encontrada en JSON"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        asistencias = data['asistencias']
        if not asistencias:
            mensaje_error = "Lista de asistencias vacía"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        tablas = [
            EmpleadoDistribucion, EmpleadoInspecciones, EmpleadoCatastro,
            EmpleadoPersuasivas, EmpleadoMedidores, EmpleadoRecaudacion, EmpleadoAdministrativo, EmpleadoNorte
        ]

        for asistencia in asistencias:
            print("\n📌 Procesando asistencia:", asistencia)  # ✅ Depuración

            if "fecha" not in asistencia or "id_empleado" not in asistencia or "mes" not in asistencia:
                mensaje_error = "Faltan datos obligatorios en la asistencia"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': mensaje_error}), 400

            try:
                fecha = datetime.strptime(asistencia['fecha'], '%Y-%m-%d').date()
            except ValueError:
                mensaje_error = f"Formato de fecha inválido -> {asistencia['fecha']}"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': "Formato de fecha inválido. Use 'YYYY-MM-DD'"}), 400

            id_empleado = asistencia['id_empleado']
            print(f"👤 Verificando asistencia de empleado {id_empleado} para {fecha}")

            for tabla in tablas: 
                asistencia_existente_otras = tabla.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()
                if asistencia_existente_otras:
                    # Obtener el nombre del empleado desde la tabla principal (Empleado)
                    empleado = Empleado.query.filter_by(id_empleado=id_empleado).first()
                    nombre_empleado = f"{empleado.nombres}" if empleado else f"ID {id_empleado}"

                    mensaje_error = f"El empleado {nombre_empleado} ya cuenta con asistencia en la fecha {fecha} en el área {tabla.__name__}."
                    print(f"🚨 ERROR: {mensaje_error}")

                    return jsonify({'success': False, 'message': mensaje_error}), 400

            asistencia_existente = EmpleadoLectura.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                print(f"✏️ Actualizando asistencia existente para {id_empleado}")
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado).strip() or None
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes) if asistencia.get('pasajes') else None
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                print(f"➕ Creando nueva asistencia para {id_empleado}")
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()

                if not empleado_original:
                    mensaje_error = f"Empleado {id_empleado} no encontrado"
                    print(f"🚨 ERROR: {mensaje_error}")
                    return jsonify({'success': False, 'message': mensaje_error}), 400

                nuevo_registro = EmpleadoLectura(
                    id_empleado=empleado_original.id_empleado,
                    nombres=empleado_original.nombres,
                    dni=empleado_original.dni,
                    cargo=empleado_original.cargo,
                    area=empleado_original.area,
                    cod_ope=empleado_original.cod_ope,
                    mes=asistencia['mes'],
                    fec_asist=fecha,
                    estado=asistencia.get('estado', '').strip() or None,
                    pasajes=asistencia.get('pasajes', None),
                    ruta=asistencia.get('ruta', '').strip() or None,
                    viaticos=asistencia.get('viaticos', 0) or 0.00
                )
                db.session.add(nuevo_registro)

        db.session.commit()
        mensaje_exito = "Registros guardados correctamente en EmpleadoLectura."
        print(f"✅ {mensaje_exito}")
        return jsonify({'success': True, 'message': mensaje_exito})

    except SQLAlchemyError as e:
        db.session.rollback()
        mensaje_error = f"Error en la base de datos: {str(e)}"
        print(f"❌ ERROR SQL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500

    except Exception as e:
        db.session.rollback()
        mensaje_error = f"Error inesperado: {str(e)}"
        print(f"❌ ERROR GENERAL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500


@app.route('/eliminar-asistencia-lectura', methods=['POST'])
def eliminar_asistencia_lectura():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro de asistencia
        asistencia = EmpleadoLectura.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Obtener nombre del empleado desde la tabla principal Empleado
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # ✅ Registrar en auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Lecturas | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})



# MÓDULO DISTRIBUCION DE RECIBOS
@app.route('/filtrar-empleados-distribucion', methods=['GET'])
def filtrar_empleado_distribucion():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'DISTRIBUCION', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en DISTRIBUCION"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/cargar-asistencia-distribucion', methods=['GET'])
def cargar_asistencia_distribucion():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: JOIN con Empleado para excluir a los cesados en 
        # los registros de asistencia ya existentes.
        # ====================================================================
        registros_asistencia = EmpleadoDistribucion.query.join(Empleado).filter(
            EmpleadoDistribucion.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_distribucion": registro.id_distribucion,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": registro.pasajes if registro.pasajes is not None else "",
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Filtramos empleados de DISTRIBUCION y TOMA DE ESTADO
        # que aún no tienen asistencia, asegurando que no estén 'CESADO'.
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoDistribucion.id_empleado).filter(
                    EmpleadoDistribucion.mes == mes_consulta
                )
            ),
            Empleado.area.in_(['DISTRIBUCION', 'TOMA DE ESTADO']),
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500


@app.route('/guardar-asistencia-detalle-distribucion', methods=['POST'])
def guardar_asistencia_detalle_distribucion():
    try:
        data = request.get_json()
        print("\n🔍 JSON recibido:", data)  # ✅ Depuración

        if not data or "asistencias" not in data:
            mensaje_error = "Clave 'asistencias' no encontrada en JSON"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        asistencias = data['asistencias']
        if not asistencias:
            mensaje_error = "Lista de asistencias vacía"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        tablas = [
            EmpleadoLectura, EmpleadoInspecciones, EmpleadoCatastro,
            EmpleadoPersuasivas, EmpleadoMedidores, EmpleadoRecaudacion, EmpleadoAdministrativo, EmpleadoNorte
        ]

        for asistencia in asistencias:
            print("\n📌 Procesando asistencia:", asistencia)  # ✅ Depuración

            if "fecha" not in asistencia or "id_empleado" not in asistencia or "mes" not in asistencia:
                mensaje_error = "Faltan datos obligatorios en la asistencia"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': mensaje_error}), 400

            try:
                fecha = datetime.strptime(asistencia['fecha'], '%Y-%m-%d').date()
            except ValueError:
                mensaje_error = f"Formato de fecha inválido -> {asistencia['fecha']}"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': "Formato de fecha inválido. Use 'YYYY-MM-DD'"}), 400

            id_empleado = asistencia['id_empleado']
            print(f"👤 Verificando asistencia de empleado {id_empleado} para {fecha}")

            for tabla in tablas: 
                asistencia_existente_otras = tabla.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()
                if asistencia_existente_otras:
                    # Obtener el nombre del empleado desde la tabla principal (Empleado)
                    empleado = Empleado.query.filter_by(id_empleado=id_empleado).first()
                    nombre_empleado = f"{empleado.nombres}" if empleado else f"ID {id_empleado}"

                    mensaje_error = f"El empleado {nombre_empleado} ya cuenta con asistencia en la fecha {fecha} en el área {tabla.__name__}."
                    print(f"🚨 ERROR: {mensaje_error}")

                    return jsonify({'success': False, 'message': mensaje_error}), 400

            asistencia_existente = EmpleadoDistribucion.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                print(f"✏️ Actualizando asistencia existente para {id_empleado}")
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado).strip() or None
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes) if asistencia.get('pasajes') else None
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                print(f"➕ Creando nueva asistencia para {id_empleado}")
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()

                if not empleado_original:
                    mensaje_error = f"Empleado {id_empleado} no encontrado"
                    print(f"🚨 ERROR: {mensaje_error}")
                    return jsonify({'success': False, 'message': mensaje_error}), 400

                nuevo_registro = EmpleadoDistribucion(
                    id_empleado=empleado_original.id_empleado,
                    nombres=empleado_original.nombres,
                    dni=empleado_original.dni,
                    cargo=empleado_original.cargo,
                    area=empleado_original.area,
                    cod_ope=empleado_original.cod_ope,
                    mes=asistencia['mes'],
                    fec_asist=fecha,
                    estado=asistencia.get('estado', '').strip() or None,
                    pasajes=asistencia.get('pasajes', None),
                    ruta=asistencia.get('ruta', '').strip() or None,
                    viaticos=asistencia.get('viaticos', 0) or 0.00
                )
                db.session.add(nuevo_registro)

        db.session.commit()
        mensaje_exito = "Registros guardados correctamente en EmpleadoDistribucion."
        print(f"✅ {mensaje_exito}")
        return jsonify({'success': True, 'message': mensaje_exito})

    except SQLAlchemyError as e:
        db.session.rollback()
        mensaje_error = f"Error en la base de datos: {str(e)}"
        print(f"❌ ERROR SQL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500

    except Exception as e:
        db.session.rollback()
        mensaje_error = f"Error inesperado: {str(e)}"
        print(f"❌ ERROR GENERAL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500


@app.route('/eliminar-asistencia-distribucion', methods=['POST'])
def eliminar_asistencia_distribucion():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro en la base de datos
        asistencia = EmpleadoDistribucion.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Buscar el nombre del empleado
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # ✅ Registrar en la tabla de auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Distribución | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})


# MÓDULO INSPECCIONES
@app.route('/filtrar-empleados-inspecciones', methods=['GET'])
def filtrar_empleado_inspecciones():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'INSPECCIONES', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en INSPECCIONES"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/cargar-asistencia-inspecciones', methods=['GET'])
def cargar_asistencia_inspecciones():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: JOIN con Empleado para excluir a los cesados en 
        # registros de asistencia ya existentes.
        # ====================================================================
        registros_asistencia = EmpleadoInspecciones.query.join(Empleado).filter(
            EmpleadoInspecciones.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_inspecciones": registro.id_inspecciones,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": registro.pasajes if registro.pasajes is not None else "",
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Filtramos empleados de INSPECCIONES que aún no 
        # tienen asistencia, asegurando que su estado laboral no sea 'CESADO'.
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoInspecciones.id_empleado).filter(
                    EmpleadoInspecciones.mes == mes_consulta
                )
            ),
            Empleado.area.in_(['INSPECCIONES']),
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500

@app.route('/guardar-asistencia-detalle-inspecciones', methods=['POST'])
def guardar_asistencia_detalle_inspecciones():
    try:
        data = request.get_json()
        print("\n🔍 JSON recibido:", data)  # ✅ Depuración

        if not data or "asistencias" not in data:
            mensaje_error = "Clave 'asistencias' no encontrada en JSON"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        asistencias = data['asistencias']
        if not asistencias:
            mensaje_error = "Lista de asistencias vacía"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        tablas = [
            EmpleadoLectura, EmpleadoDistribucion, EmpleadoCatastro,
            EmpleadoPersuasivas, EmpleadoMedidores, EmpleadoRecaudacion, EmpleadoAdministrativo, EmpleadoNorte
        ]

        for asistencia in asistencias:
            print("\n📌 Procesando asistencia:", asistencia)  # ✅ Depuración

            if "fecha" not in asistencia or "id_empleado" not in asistencia or "mes" not in asistencia:
                mensaje_error = "Faltan datos obligatorios en la asistencia"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': mensaje_error}), 400

            try:
                fecha = datetime.strptime(asistencia['fecha'], '%Y-%m-%d').date()
            except ValueError:
                mensaje_error = f"Formato de fecha inválido -> {asistencia['fecha']}"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': "Formato de fecha inválido. Use 'YYYY-MM-DD'"}), 400

            id_empleado = asistencia['id_empleado']
            print(f"👤 Verificando asistencia de empleado {id_empleado} para {fecha}")

            for tabla in tablas: 
                asistencia_existente_otras = tabla.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()
                if asistencia_existente_otras:
                    # Obtener el nombre del empleado desde la tabla principal (Empleado)
                    empleado = Empleado.query.filter_by(id_empleado=id_empleado).first()
                    nombre_empleado = f"{empleado.nombres}" if empleado else f"ID {id_empleado}"

                    mensaje_error = f"El empleado {nombre_empleado} ya cuenta con asistencia en la fecha {fecha} en el área {tabla.__name__}."
                    print(f"🚨 ERROR: {mensaje_error}")

                    return jsonify({'success': False, 'message': mensaje_error}), 400

            asistencia_existente = EmpleadoInspecciones.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                print(f"✏️ Actualizando asistencia existente para {id_empleado}")
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado).strip() or None
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes) if asistencia.get('pasajes') else None
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                print(f"➕ Creando nueva asistencia para {id_empleado}")
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()

                if not empleado_original:
                    mensaje_error = f"Empleado {id_empleado} no encontrado"
                    print(f"🚨 ERROR: {mensaje_error}")
                    return jsonify({'success': False, 'message': mensaje_error}), 400

                nuevo_registro = EmpleadoInspecciones(
                    id_empleado=empleado_original.id_empleado,
                    nombres=empleado_original.nombres,
                    dni=empleado_original.dni,
                    cargo=empleado_original.cargo,
                    area=empleado_original.area,
                    cod_ope=empleado_original.cod_ope,
                    mes=asistencia['mes'],
                    fec_asist=fecha,
                    estado=asistencia.get('estado', '').strip() or None,
                    pasajes=asistencia.get('pasajes', None),
                    ruta=asistencia.get('ruta', '').strip() or None,
                    viaticos=asistencia.get('viaticos', 0) or 0.00
                )
                db.session.add(nuevo_registro)

        db.session.commit()
        mensaje_exito = "Registros guardados correctamente en EmpleadoInspecciones."
        print(f"✅ {mensaje_exito}")
        return jsonify({'success': True, 'message': mensaje_exito})

    except SQLAlchemyError as e:
        db.session.rollback()
        mensaje_error = f"Error en la base de datos: {str(e)}"
        print(f"❌ ERROR SQL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500

    except Exception as e:
        db.session.rollback()
        mensaje_error = f"Error inesperado: {str(e)}"
        print(f"❌ ERROR GENERAL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500


@app.route('/eliminar-asistencia-inspecciones', methods=['POST'])
def eliminar_asistencia_inspecciones():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro de asistencia
        asistencia = EmpleadoInspecciones.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Buscar el nombre del empleado
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # ✅ Registrar en la tabla de auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Inspecciones Comerciales | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})


# MÓDULO CATASTRO
@app.route('/filtrar-empleados-catastro', methods=['GET'])
def filtrar_empleado_catastro():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'CATASTRO', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en CATASTRO"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/cargar-asistencia-catastro', methods=['GET'])
def cargar_asistencia_catastro():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: JOIN con Empleado para excluir a los cesados en 
        # registros de asistencia ya guardados.
        # ====================================================================
        registros_asistencia = EmpleadoCatastro.query.join(Empleado).filter(
            EmpleadoCatastro.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_catastro": registro.id_catastro,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": registro.pasajes if registro.pasajes is not None else "",
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Filtramos empleados de CATASTRO que aún no 
        # tienen asistencia, asegurando que su estado no sea 'CESADO'.
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoCatastro.id_empleado).filter(
                    EmpleadoCatastro.mes == mes_consulta
                )
            ),
            Empleado.area.in_(['CATASTRO']),
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500


@app.route('/guardar-asistencia-detalle-catastro', methods=['POST'])
def guardar_asistencia_detalle_catastro():
    try:
        data = request.get_json()
        print("\n🔍 JSON recibido:", data)  # ✅ Depuración

        if not data or "asistencias" not in data:
            mensaje_error = "Clave 'asistencias' no encontrada en JSON"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        asistencias = data['asistencias']
        if not asistencias:
            mensaje_error = "Lista de asistencias vacía"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        tablas = [
            EmpleadoLectura, EmpleadoDistribucion, EmpleadoInspecciones,
            EmpleadoPersuasivas, EmpleadoMedidores, EmpleadoRecaudacion, EmpleadoAdministrativo, EmpleadoNorte
        ]

        for asistencia in asistencias:
            print("\n📌 Procesando asistencia:", asistencia)  # ✅ Depuración

            if "fecha" not in asistencia or "id_empleado" not in asistencia or "mes" not in asistencia:
                mensaje_error = "Faltan datos obligatorios en la asistencia"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': mensaje_error}), 400

            try:
                fecha = datetime.strptime(asistencia['fecha'], '%Y-%m-%d').date()
            except ValueError:
                mensaje_error = f"Formato de fecha inválido -> {asistencia['fecha']}"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': "Formato de fecha inválido. Use 'YYYY-MM-DD'"}), 400

            id_empleado = asistencia['id_empleado']
            print(f"👤 Verificando asistencia de empleado {id_empleado} para {fecha}")

            for tabla in tablas: 
                asistencia_existente_otras = tabla.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()
                if asistencia_existente_otras:
                    # Obtener el nombre del empleado desde la tabla principal (Empleado)
                    empleado = Empleado.query.filter_by(id_empleado=id_empleado).first()
                    nombre_empleado = f"{empleado.nombres}" if empleado else f"ID {id_empleado}"

                    mensaje_error = f"El empleado {nombre_empleado} ya cuenta con asistencia en la fecha {fecha} en el área {tabla.__name__}."
                    print(f"🚨 ERROR: {mensaje_error}")

                    return jsonify({'success': False, 'message': mensaje_error}), 400

            asistencia_existente = EmpleadoCatastro.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                print(f"✏️ Actualizando asistencia existente para {id_empleado}")
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado).strip() or None
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes) if asistencia.get('pasajes') else None
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                print(f"➕ Creando nueva asistencia para {id_empleado}")
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()

                if not empleado_original:
                    mensaje_error = f"Empleado {id_empleado} no encontrado"
                    print(f"🚨 ERROR: {mensaje_error}")
                    return jsonify({'success': False, 'message': mensaje_error}), 400

                nuevo_registro = EmpleadoCatastro(
                    id_empleado=empleado_original.id_empleado,
                    nombres=empleado_original.nombres,
                    dni=empleado_original.dni,
                    cargo=empleado_original.cargo,
                    area=empleado_original.area,
                    cod_ope=empleado_original.cod_ope,
                    mes=asistencia['mes'],
                    fec_asist=fecha,
                    estado=asistencia.get('estado', '').strip() or None,
                    pasajes=asistencia.get('pasajes', None),
                    ruta=asistencia.get('ruta', '').strip() or None,
                    viaticos=asistencia.get('viaticos', 0) or 0.00
                )
                db.session.add(nuevo_registro)

        db.session.commit()
        mensaje_exito = "Registros guardados correctamente en EmpleadoCatastro."
        print(f"✅ {mensaje_exito}")
        return jsonify({'success': True, 'message': mensaje_exito})

    except SQLAlchemyError as e:
        db.session.rollback()
        mensaje_error = f"Error en la base de datos: {str(e)}"
        print(f"❌ ERROR SQL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500

    except Exception as e:
        db.session.rollback()
        mensaje_error = f"Error inesperado: {str(e)}"
        print(f"❌ ERROR GENERAL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500


@app.route('/eliminar-asistencia-catastro', methods=['POST']) 
def eliminar_asistencia_catastro():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro en la base de datos
        asistencia = EmpleadoCatastro.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Obtener el nombre del empleado (opcional para el log)
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # Registrar en auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Gestión Catastral | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})


# MÓDULO MEDICIÓN
@app.route('/filtrar-empleados-medidores', methods=['GET'])
def filtrar_empleado_medidores():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'MEDICION', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en MEDICION"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/cargar-asistencia-medidores', methods=['GET'])
def cargar_asistencia_medidores():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: JOIN con Empleado para excluir a los cesados en 
        # registros de asistencia ya guardados.
        # ====================================================================
        registros_asistencia = EmpleadoMedidores.query.join(Empleado).filter(
            EmpleadoMedidores.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_medidores": registro.id_medidores,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": registro.pasajes if registro.pasajes is not None else "",
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Filtramos empleados de MEDICION que aún no 
        # tienen asistencia, asegurando que su estado no sea 'CESADO'.
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoMedidores.id_empleado).filter(
                    EmpleadoMedidores.mes == mes_consulta
                )
            ),
            Empleado.area.in_(['MEDICION']),
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500


@app.route('/guardar-asistencia-detalle-medidores', methods=['POST'])
def guardar_asistencia_detalle_medidores():
    try:
        data = request.get_json()
        print("\n🔍 JSON recibido:", data)  # ✅ Depuración

        if not data or "asistencias" not in data:
            mensaje_error = "Clave 'asistencias' no encontrada en JSON"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        asistencias = data['asistencias']
        if not asistencias:
            mensaje_error = "Lista de asistencias vacía"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        tablas = [
            EmpleadoLectura, EmpleadoDistribucion, EmpleadoInspecciones,
            EmpleadoPersuasivas, EmpleadoCatastro, EmpleadoRecaudacion, EmpleadoAdministrativo, EmpleadoNorte
        ]

        for asistencia in asistencias:
            print("\n📌 Procesando asistencia:", asistencia)  # ✅ Depuración

            if "fecha" not in asistencia or "id_empleado" not in asistencia or "mes" not in asistencia:
                mensaje_error = "Faltan datos obligatorios en la asistencia"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': mensaje_error}), 400

            try:
                fecha = datetime.strptime(asistencia['fecha'], '%Y-%m-%d').date()
            except ValueError:
                mensaje_error = f"Formato de fecha inválido -> {asistencia['fecha']}"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': "Formato de fecha inválido. Use 'YYYY-MM-DD'"}), 400

            id_empleado = asistencia['id_empleado']
            print(f"👤 Verificando asistencia de empleado {id_empleado} para {fecha}")

            for tabla in tablas: 
                asistencia_existente_otras = tabla.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()
                if asistencia_existente_otras:
                    # Obtener el nombre del empleado desde la tabla principal (Empleado)
                    empleado = Empleado.query.filter_by(id_empleado=id_empleado).first()
                    nombre_empleado = f"{empleado.nombres}" if empleado else f"ID {id_empleado}"

                    mensaje_error = f"El empleado {nombre_empleado} ya cuenta con asistencia en la fecha {fecha} en el área {tabla.__name__}."
                    print(f"🚨 ERROR: {mensaje_error}")

                    return jsonify({'success': False, 'message': mensaje_error}), 400

            asistencia_existente = EmpleadoMedidores.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                print(f"✏️ Actualizando asistencia existente para {id_empleado}")
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado).strip() or None
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes) if asistencia.get('pasajes') else None
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                print(f"➕ Creando nueva asistencia para {id_empleado}")
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()

                if not empleado_original:
                    mensaje_error = f"Empleado {id_empleado} no encontrado"
                    print(f"🚨 ERROR: {mensaje_error}")
                    return jsonify({'success': False, 'message': mensaje_error}), 400

                nuevo_registro = EmpleadoMedidores(
                    id_empleado=empleado_original.id_empleado,
                    nombres=empleado_original.nombres,
                    dni=empleado_original.dni,
                    cargo=empleado_original.cargo,
                    area=empleado_original.area,
                    cod_ope=empleado_original.cod_ope,
                    mes=asistencia['mes'],
                    fec_asist=fecha,
                    estado=asistencia.get('estado', '').strip() or None,
                    pasajes=asistencia.get('pasajes', None),
                    ruta=asistencia.get('ruta', '').strip() or None,
                    viaticos=asistencia.get('viaticos', 0) or 0.00
                )
                db.session.add(nuevo_registro)

        db.session.commit()
        mensaje_exito = "Registros guardados correctamente en EmpleadoMedidores."
        print(f"✅ {mensaje_exito}")
        return jsonify({'success': True, 'message': mensaje_exito})

    except SQLAlchemyError as e:
        db.session.rollback()
        mensaje_error = f"Error en la base de datos: {str(e)}"
        print(f"❌ ERROR SQL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500

    except Exception as e:
        db.session.rollback()
        mensaje_error = f"Error inesperado: {str(e)}"
        print(f"❌ ERROR GENERAL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500


@app.route('/eliminar-asistencia-medidores', methods=['POST'])
def eliminar_asistencia_medidores():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro en la base de datos
        asistencia = EmpleadoMedidores.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Obtener el nombre del empleado (para registrar en auditoría)
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # Registrar en auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Gestión de Medidores | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})


# MÓDULO PERSUASIVAS
@app.route('/filtrar-empleados-persuasivas', methods=['GET'])
def filtrar_empleado_persuasivas():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'PERSUASIVAS', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en PERSUASIVAS"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/cargar-asistencia-persuasivas', methods=['GET'])
def cargar_asistencia_persuasivas():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: JOIN con Empleado para excluir a los cesados en 
        # registros de asistencia ya guardados.
        # ====================================================================
        registros_asistencia = EmpleadoPersuasivas.query.join(Empleado).filter(
            EmpleadoPersuasivas.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_persuasivas": registro.id_persuasivas,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": registro.pasajes if registro.pasajes is not None else "",
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Filtramos empleados de PERSUASIVAS que aún no 
        # tienen asistencia, asegurando que su estado no sea 'CESADO'.
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoPersuasivas.id_empleado).filter(
                    EmpleadoPersuasivas.mes == mes_consulta
                )
            ),
            Empleado.area.in_(['PERSUASIVAS']),
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500


@app.route('/guardar-asistencia-detalle-persuasivas', methods=['POST'])
def guardar_asistencia_detalle_persuasivas():
    try:
        data = request.get_json()
        print("\n🔍 JSON recibido:", data)  # ✅ Depuración

        if not data or "asistencias" not in data:
            mensaje_error = "Clave 'asistencias' no encontrada en JSON"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        asistencias = data['asistencias']
        if not asistencias:
            mensaje_error = "Lista de asistencias vacía"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        tablas = [
            EmpleadoLectura, EmpleadoDistribucion, EmpleadoInspecciones,
            EmpleadoMedidores, EmpleadoCatastro, EmpleadoRecaudacion, EmpleadoAdministrativo, EmpleadoNorte
        ]

        for asistencia in asistencias:
            print("\n📌 Procesando asistencia:", asistencia)  # ✅ Depuración

            if "fecha" not in asistencia or "id_empleado" not in asistencia or "mes" not in asistencia:
                mensaje_error = "Faltan datos obligatorios en la asistencia"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': mensaje_error}), 400

            try:
                fecha = datetime.strptime(asistencia['fecha'], '%Y-%m-%d').date()
            except ValueError:
                mensaje_error = f"Formato de fecha inválido -> {asistencia['fecha']}"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': "Formato de fecha inválido. Use 'YYYY-MM-DD'"}), 400

            id_empleado = asistencia['id_empleado']
            print(f"👤 Verificando asistencia de empleado {id_empleado} para {fecha}")

            for tabla in tablas: 
                asistencia_existente_otras = tabla.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()
                if asistencia_existente_otras:
                    # Obtener el nombre del empleado desde la tabla principal (Empleado)
                    empleado = Empleado.query.filter_by(id_empleado=id_empleado).first()
                    nombre_empleado = f"{empleado.nombres}" if empleado else f"ID {id_empleado}"

                    mensaje_error = f"El empleado {nombre_empleado} ya cuenta con asistencia en la fecha {fecha} en el área {tabla.__name__}."
                    print(f"🚨 ERROR: {mensaje_error}")

                    return jsonify({'success': False, 'message': mensaje_error}), 400

            asistencia_existente = EmpleadoPersuasivas.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                print(f"✏️ Actualizando asistencia existente para {id_empleado}")
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado).strip() or None
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes) if asistencia.get('pasajes') else None
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                print(f"➕ Creando nueva asistencia para {id_empleado}")
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()

                if not empleado_original:
                    mensaje_error = f"Empleado {id_empleado} no encontrado"
                    print(f"🚨 ERROR: {mensaje_error}")
                    return jsonify({'success': False, 'message': mensaje_error}), 400

                nuevo_registro = EmpleadoPersuasivas(
                    id_empleado=empleado_original.id_empleado,
                    nombres=empleado_original.nombres,
                    dni=empleado_original.dni,
                    cargo=empleado_original.cargo,
                    area=empleado_original.area,
                    cod_ope=empleado_original.cod_ope,
                    mes=asistencia['mes'],
                    fec_asist=fecha,
                    estado=asistencia.get('estado', '').strip() or None,
                    pasajes=asistencia.get('pasajes', None),
                    ruta=asistencia.get('ruta', '').strip() or None,
                    viaticos=asistencia.get('viaticos', 0) or 0.00
                )
                db.session.add(nuevo_registro)

        db.session.commit()
        mensaje_exito = "Registros guardados correctamente en EmpleadoPersuasivas."
        print(f"✅ {mensaje_exito}")
        return jsonify({'success': True, 'message': mensaje_exito})

    except SQLAlchemyError as e:
        db.session.rollback()
        mensaje_error = f"Error en la base de datos: {str(e)}"
        print(f"❌ ERROR SQL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500

    except Exception as e:
        db.session.rollback()
        mensaje_error = f"Error inesperado: {str(e)}"
        print(f"❌ ERROR GENERAL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500


@app.route('/eliminar-asistencia-persuasivas', methods=['POST'])
def eliminar_asistencia_persuasivas():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro de asistencia
        asistencia = EmpleadoPersuasivas.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Obtener el nombre del empleado
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # Registrar en auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Acciones Persuasivas | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})





# MÓDULO NORTE
@app.route('/filtrar-empleados-norte', methods=['GET'])
def filtrar_empleado_norte():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'NORTE', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en NORTE"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/cargar-asistencia-norte', methods=['GET'])
def cargar_asistencia_norte():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: JOIN con Empleado para excluir a los cesados en 
        # registros de asistencia ya guardados para el área Norte.
        # ====================================================================
        registros_asistencia = EmpleadoNorte.query.join(Empleado).filter(
            EmpleadoNorte.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_norte": registro.id_norte,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": registro.pasajes if registro.pasajes is not None else "",
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Filtramos empleados del área NORTE que aún no 
        # tienen asistencia, asegurando que su estado laboral no sea 'CESADO'.
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoNorte.id_empleado).filter(
                    EmpleadoNorte.mes == mes_consulta
                )
            ),
            Empleado.area.in_(['NORTE']),
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500


@app.route('/guardar-asistencia-detalle-norte', methods=['POST'])
def guardar_asistencia_detalle_norte():
    try:
        data = request.get_json()
        print("\n🔍 JSON recibido:", data)  # ✅ Depuración

        if not data or "asistencias" not in data:
            mensaje_error = "Clave 'asistencias' no encontrada en JSON"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        asistencias = data['asistencias']
        if not asistencias:
            mensaje_error = "Lista de asistencias vacía"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        tablas = [
            EmpleadoLectura, EmpleadoDistribucion, EmpleadoInspecciones,
            EmpleadoMedidores, EmpleadoCatastro, EmpleadoRecaudacion, EmpleadoAdministrativo, EmpleadoPersuasivas
        ]

        for asistencia in asistencias:
            print("\n📌 Procesando asistencia:", asistencia)  # ✅ Depuración

            if "fecha" not in asistencia or "id_empleado" not in asistencia or "mes" not in asistencia:
                mensaje_error = "Faltan datos obligatorios en la asistencia"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': mensaje_error}), 400

            try:
                fecha = datetime.strptime(asistencia['fecha'], '%Y-%m-%d').date()
            except ValueError:
                mensaje_error = f"Formato de fecha inválido -> {asistencia['fecha']}"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': "Formato de fecha inválido. Use 'YYYY-MM-DD'"}), 400

            id_empleado = asistencia['id_empleado']
            print(f"👤 Verificando asistencia de empleado {id_empleado} para {fecha}")

            for tabla in tablas: 
                asistencia_existente_otras = tabla.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()
                if asistencia_existente_otras:
                    # Obtener el nombre del empleado desde la tabla principal (Empleado)
                    empleado = Empleado.query.filter_by(id_empleado=id_empleado).first()
                    nombre_empleado = f"{empleado.nombres}" if empleado else f"ID {id_empleado}"

                    mensaje_error = f"El empleado {nombre_empleado} ya cuenta con asistencia en la fecha {fecha} en el área {tabla.__name__}."
                    print(f"🚨 ERROR: {mensaje_error}")

                    return jsonify({'success': False, 'message': mensaje_error}), 400

            asistencia_existente = EmpleadoNorte.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                print(f"✏️ Actualizando asistencia existente para {id_empleado}")
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado).strip() or None
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes) if asistencia.get('pasajes') else None
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                print(f"➕ Creando nueva asistencia para {id_empleado}")
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()

                if not empleado_original:
                    mensaje_error = f"Empleado {id_empleado} no encontrado"
                    print(f"🚨 ERROR: {mensaje_error}")
                    return jsonify({'success': False, 'message': mensaje_error}), 400

                nuevo_registro = EmpleadoNorte(
                    id_empleado=empleado_original.id_empleado,
                    nombres=empleado_original.nombres,
                    dni=empleado_original.dni,
                    cargo=empleado_original.cargo,
                    area=empleado_original.area,
                    cod_ope=empleado_original.cod_ope,
                    mes=asistencia['mes'],
                    fec_asist=fecha,
                    estado=asistencia.get('estado', '').strip() or None,
                    pasajes=asistencia.get('pasajes', None),
                    ruta=asistencia.get('ruta', '').strip() or None,
                    viaticos=asistencia.get('viaticos', 0) or 0.00
                )
                db.session.add(nuevo_registro)

        db.session.commit()
        mensaje_exito = "Registros guardados correctamente en EmpleadoNorte."
        print(f"✅ {mensaje_exito}")
        return jsonify({'success': True, 'message': mensaje_exito})

    except SQLAlchemyError as e:
        db.session.rollback()
        mensaje_error = f"Error en la base de datos: {str(e)}"
        print(f"❌ ERROR SQL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500

    except Exception as e:
        db.session.rollback()
        mensaje_error = f"Error inesperado: {str(e)}"
        print(f"❌ ERROR GENERAL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500


@app.route('/eliminar-asistencia-norte', methods=['POST'])
def eliminar_asistencia_norte():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro en la base de datos
        asistencia = EmpleadoNorte.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Obtener el nombre del empleado desde la tabla general Empleado
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # Registrar evento en auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Zona Norte | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})


# MÓDULO ASISTENCIAS ADMINISTRATIVO
@app.route('/filtrar-empleados-administrativo_1', methods=['GET'])
def filtrar_empleado_administrativo_1():
    try:
        empleados = Empleado.query.filter(
            Empleado.area == 'ADMINSTRATIVO', 
            Empleado.estado != 'CESADO'
        ).all()

        if not empleados:
            return jsonify({"mensaje": "No se encontraron empleados en ADMINSTRATIVO"}), 404

        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados
        ]

        return jsonify(empleados_data)

    except Exception as e:
        return jsonify({"error": "Error al obtener empleados", "detalles": str(e)}), 500


@app.route('/cargar-asistencia-administrativo_1', methods=['GET'])
def cargar_asistencia_administrativo_1():
    try:
        fecha_str = request.args.get('fecha')  # Obtener la fecha de la solicitud
        if not fecha_str:
            return jsonify({"error": "Debe proporcionar una fecha"}), 400

        fecha_consulta = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mes_consulta = fecha_consulta.strftime("%Y-%m")  # Formato "YYYY-MM"

        # ====================================================================
        # MODIFICACIÓN 1: JOIN con Empleado para excluir a los cesados en 
        # registros de asistencia ya guardados para el área Administrativa.
        # ====================================================================
        registros_asistencia = EmpleadoAdministrativo.query.join(Empleado).filter(
            EmpleadoAdministrativo.fec_asist == fecha_consulta,
            Empleado.estado != 'CESADO'
        ).all()

        if registros_asistencia:
            # Si hay registros, se devuelven para su modificación
            asistencia_data = [
                {
                    "id_administrativo": registro.id_administrativo,
                    "id_empleado": registro.id_empleado,
                    "dni": registro.dni,
                    "nombres": registro.nombres,
                    "cargo": registro.cargo,
                    "area": registro.area,
                    "mes": registro.mes,
                    "fec_asist": registro.fec_asist.strftime("%Y-%m-%d"),
                    "estado": registro.estado,
                    "justificacion": registro.justificacion,
                    "pasajes": float(registro.pasajes) if registro.pasajes else 0.0,
                    "viaticos": float(registro.viaticos) if registro.viaticos else 0.0,
                    "ruta": registro.ruta,
                    "cod_ope": registro.cod_ope
                }
                for registro in registros_asistencia
            ]
            return jsonify({"tipo": "modificacion", "datos": asistencia_data})

        # ====================================================================
        # MODIFICACIÓN 2: Filtramos empleados del área ADMINISTRATIVO que aún no 
        # tienen asistencia, asegurando que su estado laboral no sea 'CESADO'.
        # ====================================================================
        empleados_sin_asistencia = db.session.query(Empleado).filter(
            ~Empleado.id_empleado.in_(
                db.session.query(EmpleadoAdministrativo.id_empleado).filter(
                    EmpleadoAdministrativo.mes == mes_consulta
                )
            ),
            Empleado.area.in_(['ADMINSTRATIVO']), # Se mantiene el nombre de área de tu código original
            Empleado.estado != 'CESADO'  # <--- Filtro agregado aquí
        ).all()

        if not empleados_sin_asistencia:
            return jsonify({"mensaje": "Todos los empleados ya tienen asistencia en este mes o no hay personal activo"}), 404

        # Construir respuesta con empleados sin asistencia
        empleados_data = [
            {
                "id_empleado": empleado.id_empleado,
                "dni": empleado.dni,
                "nombres": empleado.nombres,
                "cargo": empleado.cargo,
                "cod_ope": empleado.cod_ope
            }
            for empleado in empleados_sin_asistencia
        ]

        return jsonify({"tipo": "nueva_asistencia", "datos": empleados_data})

    except Exception as e:
        return jsonify({"error": "Error al obtener asistencia", "detalles": str(e)}), 500


@app.route('/guardar-asistencia-detalle-administrativo_1', methods=['POST'])
def guardar_asistencia_detalle_administrativo_1():
    try:
        data = request.get_json()
        print("\n🔍 JSON recibido:", data)  # ✅ Depuración

        if not data or "asistencias" not in data:
            mensaje_error = "Clave 'asistencias' no encontrada en JSON"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        asistencias = data['asistencias']
        if not asistencias:
            mensaje_error = "Lista de asistencias vacía"
            print(f"🚨 ERROR: {mensaje_error}")
            return jsonify({'success': False, 'message': mensaje_error}), 400

        tablas = [
            EmpleadoLectura, EmpleadoDistribucion, EmpleadoInspecciones,
            EmpleadoMedidores, EmpleadoCatastro, EmpleadoRecaudacion, EmpleadoPersuasivas, EmpleadoNorte
        ]

        for asistencia in asistencias:
            print("\n📌 Procesando asistencia:", asistencia)  # ✅ Depuración

            if "fecha" not in asistencia or "id_empleado" not in asistencia or "mes" not in asistencia:
                mensaje_error = "Faltan datos obligatorios en la asistencia"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': mensaje_error}), 400

            try:
                fecha = datetime.strptime(asistencia['fecha'], '%Y-%m-%d').date()
            except ValueError:
                mensaje_error = f"Formato de fecha inválido -> {asistencia['fecha']}"
                print(f"🚨 ERROR: {mensaje_error}")
                return jsonify({'success': False, 'message': "Formato de fecha inválido. Use 'YYYY-MM-DD'"}), 400

            id_empleado = asistencia['id_empleado']
            print(f"👤 Verificando asistencia de empleado {id_empleado} para {fecha}")

            for tabla in tablas: 
                asistencia_existente_otras = tabla.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()
                if asistencia_existente_otras:
                    # Obtener el nombre del empleado desde la tabla principal (Empleado)
                    empleado = Empleado.query.filter_by(id_empleado=id_empleado).first()
                    nombre_empleado = f"{empleado.nombres}" if empleado else f"ID {id_empleado}"

                    mensaje_error = f"El empleado {nombre_empleado} ya cuenta con asistencia en la fecha {fecha} en el área {tabla.__name__}."
                    print(f"🚨 ERROR: {mensaje_error}")

                    return jsonify({'success': False, 'message': mensaje_error}), 400

            asistencia_existente = EmpleadoAdministrativo.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

            if asistencia_existente:
                print(f"✏️ Actualizando asistencia existente para {id_empleado}")
                asistencia_existente.estado = asistencia.get('estado', asistencia_existente.estado).strip() or None
                asistencia_existente.pasajes = asistencia.get('pasajes', asistencia_existente.pasajes)
                asistencia_existente.ruta = asistencia.get('ruta', asistencia_existente.ruta)
                asistencia_existente.viaticos = asistencia.get('viaticos', asistencia_existente.viaticos)
            else:
                print(f"➕ Creando nueva asistencia para {id_empleado}")
                empleado_original = Empleado.query.filter_by(id_empleado=id_empleado).first()

                if not empleado_original:
                    mensaje_error = f"Empleado {id_empleado} no encontrado"
                    print(f"🚨 ERROR: {mensaje_error}")
                    return jsonify({'success': False, 'message': mensaje_error}), 400

                nuevo_registro = EmpleadoAdministrativo(
                    id_empleado=empleado_original.id_empleado,
                    nombres=empleado_original.nombres,
                    dni=empleado_original.dni,
                    cargo=empleado_original.cargo,
                    area=empleado_original.area,
                    mes=asistencia['mes'],
                    fec_asist=fecha,
                    estado=asistencia.get('estado', '').strip() or None,
                    pasajes=asistencia.get('pasajes', 0) or 0.00,
                    ruta=asistencia.get('ruta', '').strip() or None,
                    viaticos=asistencia.get('viaticos', 0) or 0.00
                )
                db.session.add(nuevo_registro)

        db.session.commit()
        mensaje_exito = "Registros guardados correctamente en EmpleadoAdministrativo."
        print(f"✅ {mensaje_exito}")
        return jsonify({'success': True, 'message': mensaje_exito})

    except SQLAlchemyError as e:
        db.session.rollback()
        mensaje_error = f"Error en la base de datos: {str(e)}"
        print(f"❌ ERROR SQL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500

    except Exception as e:
        db.session.rollback()
        mensaje_error = f"Error inesperado: {str(e)}"
        print(f"❌ ERROR GENERAL: {mensaje_error}")
        return jsonify({'success': False, 'message': mensaje_error}), 500


@app.route('/eliminar-asistencia-administrativo_1', methods=['POST'])
def eliminar_asistencia_administrativo_1():
    try:
        data = request.get_json()
        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')

        if not id_empleado or not fecha:
            return jsonify({'success': False, 'message': 'Datos insuficientes'}), 400

        # Buscar el registro en la base de datos
        asistencia = EmpleadoAdministrativo.query.filter_by(fec_asist=fecha, id_empleado=id_empleado).first()

        if not asistencia:
            return jsonify({'success': False, 'message': 'No se encontró el registro de asistencia'}), 404

        # Obtener el nombre del empleado desde la tabla general Empleado
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # Eliminar el registro
        db.session.delete(asistencia)
        db.session.commit()

        # Registrar evento en auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_asistencia',
                modulo=f"Administrativo | Fecha: {fecha} | Empleado: {nombre_empleado}"
            )

        return jsonify({'success': True, 'message': 'Registro eliminado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar: {str(e)}'})
   


##ADMINISTRATIVO
TABLAS_ASISTENCIA = {
    "RECAUDACION": EmpleadoRecaudacion,
    "TOMA DE ESTADO": EmpleadoLectura,
    "DISTRIBUCION": EmpleadoDistribucion,
    "CATASTRO": EmpleadoCatastro,
    "MEDICION": EmpleadoMedidores,
    "INSPECCIONES": EmpleadoInspecciones,
    "PERSUASIVAS": EmpleadoPersuasivas,
    "NORTE": EmpleadoNorte,
    "ADMINSTRATIVO": EmpleadoAdministrativo,
}

@app.route('/api/getAsistencia', methods=['POST'])
def get_asistencia():
    try:
        # Obtener datos de la petición
        data = request.get_json()
        area = data.get('area')
        fecha_inicio = data.get("fechaInicio")
        fecha_fin = data.get("fechaFin")

        # Validar que se envíen área y fechas
        if not area or not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Área y fechas son obligatorias"}), 400

        # Validar que el área sea correcta
        modelo_asistencia = TABLAS_ASISTENCIA.get(area.upper())
        if not modelo_asistencia:
            print(f"ERROR: Área no válida: {area}")  # Mensaje detallado en consola
            return jsonify({"error": f"Área no válida: {area}"}), 400

        # Convertir fechas de string a objeto datetime
        fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

        # Verificar que la fecha inicio no sea mayor a la fecha fin
        if fecha_inicio_dt > fecha_fin_dt:
            return jsonify({"error": "La fecha de inicio no puede ser mayor que la fecha de fin"}), 400

        # Filtrar datos dentro del rango de fechas
        empleados = db.session.query(
            modelo_asistencia.dni,
            modelo_asistencia.nombres,
            modelo_asistencia.cargo,
            modelo_asistencia.fec_asist,
            modelo_asistencia.estado,
            modelo_asistencia.pasajes,
            modelo_asistencia.viaticos,
            modelo_asistencia.ruta
        ).filter(
            modelo_asistencia.fec_asist.between(fecha_inicio_dt, fecha_fin_dt)
        ).all()

        # Estructurar los datos para el JSON de respuesta
        empleados_dict = {}

        for empleado in empleados:
            dni = empleado.dni

            if dni not in empleados_dict:
                empleados_dict[dni] = {
                    "dni": empleado.dni,
                    "nombres": empleado.nombres,
                    "cargo": empleado.cargo,
                    "asistencia": {},
                    "pasajes": {},
                    "viaticos": {},
                    "rutas": {},
                }

            # Si hay fecha de asistencia, asignar los valores en el día correspondiente
            if empleado.fec_asist:
                dia = empleado.fec_asist.strftime("%Y-%m-%d")  # Guardar con formato YYYY-MM-DD
                empleados_dict[dni]["asistencia"][dia] = empleado.estado or ""
                empleados_dict[dni]["pasajes"][dia] = str(empleado.pasajes or "")
                empleados_dict[dni]["viaticos"][dia] = str(empleado.viaticos or "")
                empleados_dict[dni]["rutas"][dia] = empleado.ruta or ""

        # Convertir el diccionario a lista para enviarlo como JSON
        respuesta = list(empleados_dict.values())

        return jsonify(respuesta)

    except Exception as e:
        error_trace = traceback.format_exc()  # Obtener el error detallado
        print(f"ERROR INTERNO DEL SERVIDOR:\n{error_trace}")  # Imprimir en consola
        return jsonify({"error": "Error interno del servidor", "detalle": str(e)}), 500
    


@app.route('/api/getAsistenciaCompleta', methods=['POST'])
def get_asistencia_completa():
    data = request.get_json()

    # Validación de parámetros requeridos
    if not data or 'fechaInicio' not in data or 'fechaFin' not in data:
        return jsonify({"error": "Se requieren los campos 'fechaInicio' y 'fechaFin'"}), 400

    fechaInicio = data['fechaInicio']
    fechaFin = data['fechaFin']

    # Validar que la fecha de inicio no sea mayor a la fecha de fin
    if datetime.strptime(fechaInicio, "%Y-%m-%d") > datetime.strptime(fechaFin, "%Y-%m-%d"):
        return jsonify({"error": "La fecha de inicio no puede ser mayor que la fecha de fin"}), 400

    # Modelos por área
    modelos = {
        "empleado_inspecciones": EmpleadoInspecciones,
        "empleado_lectura": EmpleadoLectura,
        "empleado_distribucion": EmpleadoDistribucion,
        "empleado_catastro": EmpleadoCatastro,
        "empleado_persuasivas": EmpleadoPersuasivas,
        "empleado_medidores": EmpleadoMedidores,
        "empleado_recaudacion": EmpleadoRecaudacion,
        "empleado_norte": EmpleadoNorte,
        "empleado_administrativo": EmpleadoAdministrativo
    }

    asistencia_completa = {}

    for area, modelo in modelos.items():
        empleados = modelo.query.all()
        asistencias = modelo.query.filter(modelo.fec_asist.between(fechaInicio, fechaFin)).all()

        asistencia_por_empleado = {}
        for a in asistencias:
            if a.id_empleado not in asistencia_por_empleado:
                asistencia_por_empleado[a.id_empleado] = []
            asistencia_por_empleado[a.id_empleado].append(a)

        asistencia_completa[area] = []
        for empleado in empleados:
            registros = asistencia_por_empleado.get(empleado.id_empleado, [])

            for registro in registros:
                asistencia_completa[area].append({
                    "id": getattr(empleado, f"id_{area}", None),
                    "nombres": empleado.nombres,
                    "dni": empleado.dni,
                    "cargo": empleado.cargo,
                    "area_global": area,
                    "area": empleado.area,
                    "mes": registro.mes if registro else None,
                    "fec_asist": registro.fec_asist.strftime('%Y-%m-%d') if registro.fec_asist else None,
                    "estado": registro.estado if registro else "-",
                    "pasajes": registro.pasajes if registro.pasajes is not None else "0",
                    "ruta": registro.ruta if registro else "-",
                    "viaticos": float(registro.viaticos) if registro.viaticos is not None else 0.0,
                    "cod_ope": registro.cod_ope if registro else None
                })

    # Generar y devolver el archivo Excel
    return generar_reporte_excel(asistencia_completa, fechaInicio, fechaFin)

def safe_float(value):
    """Convierte un valor a float si es numérico, de lo contrario devuelve 0.0"""
    try:
        if isinstance(value, (int, float)):
            return float(value)
        elif isinstance(value, str):
            value = value.strip().replace(",", ".")  # Si tiene coma, la cambia por punto
            if value.replace('.', '', 1).isdigit():  # Verifica si es un número
                return float(value)
        return 0.0  # Si no es numérico, retorna 0.0
    except (ValueError, TypeError):
        return 0.0



def generar_reporte_excel(datos, fechaInicio, fechaFin):
    wb = Workbook()

    # Paleta de colores para cada área global
    colores_areas = {
        "empleado_catastro": "fff700",
        "empleado_distribucion": "8bff00",
        "empleado_inspecciones": "ffaa00",  
        "empleado_lectura": "00fff3",
        "empleado_medidores": "ffd586",        
        "empleado_persuasivas": "ff7fe6",    
        "empleado_recaudacion": "528cff",
        "empleado_norte": "95b8ee",  
        "empleado_administrativo": "bad931" 
    }

    # Paleta de colores para los estados en la hoja "Asistencias"
    colores_estados = {
        "A": "4fe548",   # Verde limón
        "F": "FF0000",   # Rojo
        "DT": "FFA500",  # Naranja
        "FT": "FFFF00",  # Amarillo
        "LG": "87CEEB",  # Celeste
        "DM": "8A2BE2",  # Violeta
        "V": "008000",   # Verde oscuro
        "LSG": "FFC0CB", # Rosa
        "SU": "A52A2A",  # Marrón
        "CE": "00CED1",  # Turquesa
        "FG": "9400D3",  # Púrpura
        "LD": "4682B4",  # Azul acero
        "DC": "ff00ff",  # Verde claro
        "AP": "D2691E",  # Chocolate
        "LP": "BDB76B",  # Caqui oscuro
        "TC": "DC143C"   # Carmesí
    }

    # Diccionario de colores para empleados con área diferente a su área global
    colores_areas_especificas = {
        "CATASTRO": "ffe683",
        "PERSUASIVAS": "ffb5f1",
        "TOMA DE ESTADO": "a6eae7",
        "INSPECCIONES": "ffaa59",
        "DISTRIBUCION": "cae39c",
        "MEDICION": "cead6f",
        "RECAUDACION": "a2bbeb",
        "NORTE": "95b8ee",
        "ADMINSTRATIVO": "adb781"

    }
    
    orden_areas = [
    "empleado_lectura",
    "empleado_catastro",
    "empleado_persuasivas",
    "empleado_inspecciones",
    "empleado_distribucion",
    "empleado_medidores",
    "empleado_recaudacion",
    "empleado_norte",
    "empleado_administrativo"
]

    # Estilo de bordes negros
    borde_negro = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    color_domingo = PatternFill(start_color="f17575", end_color="f17575", fill_type="solid")

    # Obtener fechas del rango
    fechas = []
    fecha_actual = datetime.strptime(fechaInicio, "%Y-%m-%d")
    fecha_fin = datetime.strptime(fechaFin, "%Y-%m-%d")

    while fecha_actual <= fecha_fin:
        fechas.append(fecha_actual.strftime("%Y-%m-%d"))
        fecha_actual += timedelta(days=1)

    # Crear hojas para cada tipo de datos
    hojas = {
        "Asistencias": "estado",
        "Pasajes": "pasajes",
        "Viáticos": "viaticos",
        "Rutas": "ruta"
    }

    for nombre_hoja, campo in hojas.items():
        ws = wb.create_sheet(title=nombre_hoja)


        # 🔹 Definir estilos para la cabecera
        cabecera_fill = PatternFill(start_color="327bc0", end_color="327bc0", fill_type="solid")
        cabecera_font = Font(bold=True, color="FFFFFF", size=12)
        cabecera_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        locale.setlocale(locale.LC_TIME, "es_ES.utf8")  

        # 🔹 Agrupar las fechas por mes
        meses = defaultdict(list)
        for col_idx, fecha in enumerate(fechas, start=7):  # Desde la columna 7 en adelante
            mes = datetime.strptime(fecha, "%Y-%m-%d").strftime("%B")  # Obtener nombre del mes
            meses[mes].append(col_idx)

        # 🔹 Insertar fila para los nombres de los meses
        ws.insert_rows(1)  # Insertar nueva fila antes de los encabezados

        # 🔹 Aplicar los nombres de los meses fusionando las celdas correspondientes
        mes_font = Font(bold=True, size=12, color="FFFFFF")
        mes_alignment = Alignment(horizontal="center", vertical="center")
        mes_fill = PatternFill(start_color="184875", end_color="184875", fill_type="solid")

        for mes, columnas in meses.items():
            inicio = get_column_letter(columnas[0])  # Primera columna del mes
            fin = get_column_letter(columnas[-1])   # Última columna del mes
            rango = f"{inicio}1:{fin}1"  # Rango de celdas a fusionar

            ws.merge_cells(rango)  # Fusionar celdas
            celda_mes = ws[f"{inicio}1"]  # Celda principal del mes
            celda_mes.value = mes.upper()  # Escribir el nombre del mes en mayúsculas
            celda_mes.font = mes_font
            celda_mes.alignment = mes_alignment
            celda_mes.fill = mes_fill

        # 🔹 Ajustar altura de la fila de los meses
        ws.row_dimensions[1].height = 25

        # 🔹 Asegurar que los encabezados estén en la segunda fila
        encabezado = ["N°", "DNI", "NOMBRES", "CARGO", "ÁREA GLOBAL", "ÁREA"] + [f"{d[-2:]}/{d[5:7]}" for d in fechas]

        for col_idx, valor in enumerate(encabezado, start=1):
            cell = ws.cell(row=2, column=col_idx, value=valor)
            cell.fill = cabecera_fill
            cell.font = cabecera_font
            cell.alignment = cabecera_alignment

        # 🔹 Ajustar altura de la fila de encabezados
        ws.row_dimensions[2].height = 25

        if campo in ["pasajes", "viaticos"]:
            encabezado.append("Total")

        for col_num, cell in enumerate(ws[1], 1):  # Primera fila (cabecera)
            cell.fill = cabecera_fill
            cell.font = cabecera_font
            cell.border = borde_negro  # Aplicar borde
        
        # Pintar columnas de fechas en domingo de rojo
        for col_idx, fecha in enumerate(fechas, start=7):  # Columnas desde la 7 en adelante
            if datetime.strptime(fecha, "%Y-%m-%d").weekday() == 6:  # Domingo
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.fill = color_domingo

        contador = 1
        
        for area in orden_areas:
            if area in datos:  # Solo iteramos si el área está en el diccionario
                empleados = datos[area]
                empleados_map = {}
                color_fondo = PatternFill(start_color=colores_areas.get(area, "FFFFFF"), 
                                        end_color=colores_areas.get(area, "FFFFFF"), 
                                        fill_type="solid")
                
            for empleado in empleados:
                clave = f"{empleado['dni']}-{area}"
                if clave not in empleados_map:
                    empleados_map[clave] = {
                        "dni": empleado["dni"],
                        "nombres": empleado["nombres"],
                        "cargo": empleado["cargo"],
                        "area_global": area,
                        "area": empleado["area"],
                        "fechas": {fecha: "" for fecha in fechas},
                        "totalMonto": 0
                    }

                emp_data = empleados_map[clave]
                fecha_registro = empleado["fec_asist"]

                if fecha_registro in fechas:
                    if campo == "estado":
                        emp_data["fechas"][fecha_registro] = empleado["estado"] or "-"
                    elif campo == "ruta":
                        emp_data["fechas"][fecha_registro] = empleado["ruta"] or ""
                    elif campo in ["pasajes", "viaticos"]:
                            valor = empleado.get(campo, "")
                            
                            # Verificar si el valor es numérico
                            try:
                                monto = float(valor)
                                es_numero = True
                            except (ValueError, TypeError):
                                es_numero = False

                            # Registrar el valor (número o texto) en la celda
                            if emp_data["fechas"][fecha_registro] == "":
                                emp_data["fechas"][fecha_registro] = valor
                                if es_numero:
                                    emp_data["totalMonto"] += monto
                            else:
                                # Si ya hay un valor y el nuevo es numérico mayor, se actualiza
                                valor_actual = emp_data["fechas"][fecha_registro]
                                try:
                                    monto_actual = float(valor_actual)
                                    if es_numero and monto > monto_actual:
                                        emp_data["totalMonto"] += (monto - monto_actual)
                                        emp_data["fechas"][fecha_registro] = monto
                                except (ValueError, TypeError):
                                    # Si el valor actual no es numérico y el nuevo sí, se mantiene el texto y solo se suma el número al total
                                    if es_numero:
                                        emp_data["totalMonto"] += monto


            for emp in empleados_map.values():
                fila = [
                    contador,
                    emp["dni"],
                    emp["nombres"],
                    emp["cargo"],
                    emp["area_global"],
                    emp["area"]
                ] + [emp["fechas"][fecha] for fecha in fechas]

                if campo in ["pasajes", "viaticos"]:
                    fila.append(emp["totalMonto"])

                ws.append(fila)

                fila_num = ws.max_row
                # Verificar si el empleado tiene un área diferente a su área global
                if emp["area_global"] != emp["area"]:
                    color_area = colores_areas_especificas.get(emp["area"], None)
                    if color_area:
                        color_fondo = PatternFill(start_color=color_area, end_color=color_area, fill_type="solid")
                        for col_idx in range(1, 7):  # Aplicar color a las primeras 6 columnas
                            cell = ws.cell(row=fila_num, column=col_idx)
                            cell.fill = color_fondo  
                            cell.border = borde_negro   # Aplicar borde negro

                contador += 1

        # Aplicar colores a los estados en la hoja "Asistencias"
        if nombre_hoja == "Asistencias":
            for row in ws.iter_rows(min_row=2, min_col=7, max_col=ws.max_column):
                for cell in row:
                    estado = str(cell.value).strip()
                    if estado in colores_estados:
                        color_hex = colores_estados[estado]
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        cell.font = Font(bold=True, color="000000")  # Negrita, texto negro
                        cell.alignment = Alignment(horizontal="center", vertical="center")  # Centrado

         # 🔹 Aplicar color rojo a TODA la columna si la fecha cae en domingo
        for col_idx, fecha in enumerate(fechas, start=7):  # Columnas desde la 7 en adelante
            if datetime.strptime(fecha, "%Y-%m-%d").weekday() == 6:  # Si es domingo
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.fill = color_domingo  # Pintar de rojo toda la columna


    # Aplicar bordes a todas las celdas
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        for row in ws.iter_rows():
            for cell in row:
                cell.border = borde_negro

                # Mantener atributos anteriores y solo cambiar el tamaño
                cell.font = Font(
                    name=cell.font.name,  # Mantiene la fuente original
                    size=9,  # Cambia solo el tamaño
                    bold=cell.font.bold,  # Mantiene negrita si la tenía
                    italic=cell.font.italic,  # Mantiene cursiva si la tenía
                    color=cell.font.color  # Mantiene el color original
                )



        # Consolidar asistencias en una nueva hoja llamada "Consolidado"
        if "Consolidado" in wb.sheetnames:
            del wb["Consolidado"]

        ws_consolidado = wb.create_sheet(title="Consolidado")

        # Definir encabezados para la hoja Consolidado
        encabezado_consolidado = ["N°", "DNI", "NOMBRES", "CARGO", "ÁREA GLOBAL", "ÁREA"] + [f"{d[-2:]}/{d[5:7]}" for d in fechas]

        # Estilo de encabezado para la hoja Consolidado
        for col_idx, valor in enumerate(encabezado_consolidado, start=1):
            cell = ws_consolidado.cell(row=1, column=col_idx, value=valor)
            cell.fill = cabecera_fill
            cell.font = cabecera_font
            cell.alignment = cabecera_alignment

        # Mapeo de áreas
        mapa_areas = {
            "empleado_lectura": "TOMA DE ESTADO",
            "empleado_catastro": "CATASTRO",
            "empleado_persuasivas": "PERSUASIVAS",
            "empleado_inspecciones": "INSPECCIONES",
            "empleado_distribucion": "DISTRIBUCION",
            "empleado_medidores": "MEDICION",
            "empleado_recaudacion": "RECAUDACION",
            "empleado_norte": "NORTE",
            "empleado_administrativo": "ADMINISTRATIVO"
        }

        consolidado_map = {}

        # Consolidar datos de todas las áreas
        for area, empleados in datos.items():
            area_nombre = mapa_areas.get(area, area)
            
            for empleado in empleados:
                dni = empleado["dni"]
                clave = f"{dni}-{empleado['nombres']}"
                
                if clave not in consolidado_map:
                    consolidado_map[clave] = {
                        "dni": dni,
                        "nombres": empleado["nombres"],
                        "cargo": empleado["cargo"],
                        "area_global": area_nombre,
                        "area": empleado["area"],
                        "fechas": {fecha: "" for fecha in fechas}
                    }

                fecha_registro = empleado["fec_asist"]
                if fecha_registro in fechas:
                    estado = empleado["estado"] or "-"
                    consolidado_map[clave]["fechas"][fecha_registro] = estado

                # Priorizar área global y área coincidente
                if empleado["area_global"] == empleado["area"]:
                    consolidado_map[clave]["area"] = empleado["area"]
        
        # Ordenar alfabéticamente por el área
        consolidado_ordenado = sorted(consolidado_map.values(), key=lambda x: x["area"])

        # Paleta de colores para los estados
        colores_estados_1 = { 
            "A": "4fe548",   # Verde limón
            "F": "FF0000",   # Rojo
            "DT": "FFA500",  # Naranja
            "FT": "FFFF00",  # Amarillo
            "LG": "87CEEB",  # Celeste
            "DM": "8A2BE2",  # Violeta
            "V": "008000",   # Verde oscuro
            "LSG": "FFC0CB", # Rosa
            "SU": "A52A2A",  # Marrón
            "CE": "00CED1",  # Turquesa
            "FG": "9400D3",  # Púrpura
            "LD": "4682B4",  # Azul acero
            "DC": "ff00ff",  # Verde claro
            "AP": "D2691E",  # Chocolate
            "LP": "BDB76B",  # Caqui oscuro
            "TC": "DC143C"   # Carmesí
        }

        # Escribir datos consolidados en la hoja
        contador = 1
        for empleado in consolidado_ordenado:
            fila = [
                contador,
                empleado["dni"],
                empleado["nombres"],
                empleado["cargo"],
                empleado["area_global"],
                empleado["area"]
            ] + [empleado["fechas"][fecha] for fecha in fechas]

            ws_consolidado.append(fila)
            contador += 1
        
        # Aplicar colores a los estados en la hoja Consolidado
        for fila in ws_consolidado.iter_rows(min_row=2, max_row=ws_consolidado.max_row, min_col=7, max_col=6+len(fechas)):
            for celda in fila:
                estado = celda.value
                if estado in colores_estados_1:
                    color_hex = colores_estados_1[estado]
                    celda.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        
        # Colores para las áreas específicas
        colores_areas_especificas_1 = { 
            "CATASTRO": "ffe683",
            "PERSUASIVAS": "ffb5f1",
            "TOMA DE ESTADO": "a6eae7",
            "INSPECCIONES": "ffaa59",
            "DISTRIBUCION": "cae39c",
            "MEDICION": "cead6f",
            "RECAUDACION": "a2bbeb",
            "NORTE": "95b8ee",
            "ADMINSTRATIVO": "adb781"
        }

        # Aplicar colores a las áreas en las columnas de la 1 a la 6 (excepto las cabeceras)
        for fila in ws_consolidado.iter_rows(min_row=2, max_row=ws_consolidado.max_row, min_col=1, max_col=6):
            # Obtener el valor del Área en la columna 6
            area = fila[5].value  # fila[5] corresponde a la columna 6 (Área)
            
            # Si el área tiene un color especificado, aplicar el color en las columnas 1 a 6
            if area in colores_areas_especificas_1:
                color_hex = colores_areas_especificas_1[area]
                
                for celda in fila:
                    celda.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")


        # Pintar de rojo las columnas de domingos
        rojo_domingo = PatternFill(start_color="e15858", end_color="e15858", fill_type="solid")

        for idx, fecha in enumerate(fechas):
            dia_semana = datetime.strptime(fecha, "%Y-%m-%d").weekday()
            
            # Si es domingo (6 = domingo en Python), colorear la columna
            if dia_semana == 6:
                columna_domingo = 7 + idx
                for fila in ws_consolidado.iter_rows(min_row=1, max_row=ws_consolidado.max_row, min_col=columna_domingo, max_col=columna_domingo):
                    for celda in fila:
                        celda.fill = rojo_domingo

        # Aplicar bordes y colores a la hoja Consolidado
        for row in ws_consolidado.iter_rows():
            for cell in row:
                cell.border = borde_negro
                cell.font = Font(
                    name=cell.font.name,  
                    size=9,  
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color  
                )
        
        # Ajustar anchura de columnas
        for column_cells in ws_consolidado.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws_consolidado.column_dimensions[column_cells[0].column_letter].width = length + 2
    

    # Crear una nueva hoja llamada "Consolidado_Pasajes"
    if "consolidado_pasajes" in wb.sheetnames:
        del wb["consolidado_pasajes"]
    ws_consolidado_pasajes = wb.create_sheet(title="consolidado_pasajes")

    # Definir encabezados para "consolidado_pasajes"
    encabezado_pasajes = ["N°", "DNI", "NOMBRES", "CARGO", "ÁREA GLOBAL", "ÁREA"] + [f"{d[-2:]}/{d[5:7]}" for d in fechas]
    encabezado_pasajes.append("Total Pasajes")

    # Estilo de encabezado
    cabecera_fill_pasajes = PatternFill(start_color="327BC0", end_color="327BC0", fill_type="solid")  # Color azul #327bc0
    cabecera_font_pasajes = Font(bold=True, color="FFFFFF")  # Texto blanco y en negrita
    cabecera_alignment_pasajes = Alignment(horizontal="center", vertical="center")

    # Aplicar encabezado con estilo
    for col_idx, valor in enumerate(encabezado_pasajes, start=1):
        cell = ws_consolidado_pasajes.cell(row=1, column=col_idx, value=valor)
        cell.fill = cabecera_fill_pasajes
        cell.font = Font(bold=True, color="FFFFFF", size=9)  # Negrita, blanco y tamaño 9
        cell.alignment = cabecera_alignment_pasajes
        cell.border = borde_negro

    # Ordenar las áreas alfabéticamente para el consolidado
    empleados_por_area_pasajes = defaultdict(list)
    for area, empleados in datos.items():
        for empleado in empleados:
            empleados_por_area_pasajes[empleado["area"]].append(empleado)

    areas_ordenadas_pasajes = sorted(empleados_por_area_pasajes.keys())

    contador = 1
    for area in areas_ordenadas_pasajes:
        empleados = empleados_por_area_pasajes[area]
        empleados_map = {}

        for empleado in empleados:
            clave = f"{empleado['dni']}-{empleado['area']}"
            if clave not in empleados_map:
                empleados_map[clave] = {
                    "dni": empleado["dni"],
                    "nombres": empleado["nombres"],
                    "cargo": empleado["cargo"],
                    "area_global": empleado["area_global"],
                    "area": empleado["area"],
                    "fechas": {fecha: 0.0 for fecha in fechas},
                    "totalMonto": 0.0
                }

            emp_data = empleados_map[clave]
            fecha_registro = empleado["fec_asist"]

            if fecha_registro in fechas:
                valor_pasaje = empleado.get("pasajes", "")
                fecha_registro = empleado["fec_asist"]

                # Intentar convertir a número, si no se puede, dejar como texto
                try:
                    monto = float(valor_pasaje)
                    es_numero = True
                except (ValueError, TypeError):
                    es_numero = False

                # Procesar el valor para la fecha correspondiente
                if emp_data["fechas"][fecha_registro] in [0.0, ""]:
                    emp_data["fechas"][fecha_registro] = valor_pasaje
                    if es_numero:
                        emp_data["totalMonto"] += monto
                else:
                    # Si ya hay un valor y el nuevo es numérico mayor, actualizar
                    valor_actual = emp_data["fechas"][fecha_registro]
                    try:
                        monto_actual = float(valor_actual)
                        if es_numero and monto > monto_actual:
                            emp_data["totalMonto"] += (monto - monto_actual)
                            emp_data["fechas"][fecha_registro] = monto
                    except (ValueError, TypeError):
                        # Si el valor actual es texto y el nuevo es numérico, sumar al total sin cambiar el texto
                        if es_numero:
                            emp_data["totalMonto"] += monto


        for emp in empleados_map.values():
            fila = [
                contador,
                emp["dni"],
                emp["nombres"],
                emp["cargo"],
                emp["area_global"],
                emp["area"]
            ] + [emp["fechas"][fecha] if emp["fechas"][fecha] != 0 else "" for fecha in fechas]

            fila.append(emp["totalMonto"] if emp["totalMonto"] != 0 else "")
            ws_consolidado_pasajes.append(fila)

            fila_num = ws_consolidado_pasajes.max_row
            # Aplicar color a las áreas específicas
            if emp["area_global"] != emp["area"]:
                color_area = colores_areas_especificas_1.get(emp["area"], None)
                if color_area:
                    color_fondo = PatternFill(start_color=color_area, end_color=color_area, fill_type="solid")
                    for col_idx in range(1, 7):
                        cell = ws_consolidado_pasajes.cell(row=fila_num, column=col_idx)
                        cell.fill = color_fondo
                        cell.border = borde_negro

            contador += 1

    # Pintar columnas de domingos en rojo
    for col_idx, fecha in enumerate(fechas, start=7):
        if datetime.strptime(fecha, "%Y-%m-%d").weekday() == 6:
            for row in ws_consolidado_pasajes.iter_rows(min_row=1, max_row=ws_consolidado_pasajes.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = color_domingo
    
    # Ajustar altura de fila para encabezado
    for row in ws_consolidado_pasajes.iter_rows(min_row=2, max_row=ws_consolidado_pasajes.max_row, min_col=1, max_col=len(encabezado_pasajes)):
        for cell in row:
            cell.border = borde_negro  # Bordes negros en toda la tabla
            cell.font = Font(size=9)    # Tamaño de fuente 9



    # Eliminar la hoja vacía por defecto
    wb.remove(wb["Sheet"])

    # Guardar en memoria y devolver el archivo
    output = BytesIO()
    nombre_archivo = f"Reporte_Completo_{fechaInicio.replace('-', '')}_{fechaFin.replace('-', '')}.xlsx"
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=nombre_archivo, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/api/getConsolidado', methods=['POST'])
def get_consolidado():
    db.session.expire_all()

    data = request.get_json()
    if not data or 'fechaInicio' not in data or 'fechaFin' not in data or 'tipo' not in data:
        return jsonify({"error": "Se requieren los campos 'fechaInicio', 'fechaFin' y 'tipo'"}), 400

    fechaInicio = data['fechaInicio']
    fechaFin = data['fechaFin']
    tipo = data['tipo']  # "asistencias" o "pasajes"

    if tipo not in ["asistencias", "pasajes"]:
        return jsonify({"error": "Tipo no válido"}), 400

    if datetime.strptime(fechaInicio, "%Y-%m-%d") > datetime.strptime(fechaFin, "%Y-%m-%d"):
        return jsonify({"error": "La fecha de inicio no puede ser mayor que la fecha de fin"}), 400

    area_nombre_map = {
        "lectura": "TOMA DE ESTADO",
        "catastro": "CATASTRO",
        "persuasivas": "PERSUASIVAS",
        "inspecciones": "INSPECCIONES",
        "distribucion": "DISTRIBUCION",
        "medidores": "MEDICION",
        "norte": "NORTE"
    }

    modelos = [
        (EmpleadoLectura, "lectura"),
        (EmpleadoCatastro, "catastro"),
        (EmpleadoPersuasivas, "persuasivas"),
        (EmpleadoInspecciones, "inspecciones"),
        (EmpleadoDistribucion, "distribucion"),
        (EmpleadoMedidores, "medidores"),
        (EmpleadoNorte, "norte"),
    ]

    fechas = [
        (datetime.strptime(fechaInicio, "%Y-%m-%d") + timedelta(days=i)).strftime("%Y-%m-%d")
        for i in range((datetime.strptime(fechaFin, "%Y-%m-%d") - datetime.strptime(fechaInicio, "%Y-%m-%d")).days + 1)
    ]

    consolidado_map = {}

    for modelo, nombre_area in modelos:
        registros = modelo.query.filter(modelo.fec_asist.between(fechaInicio, fechaFin)).all()

        for registro in registros:
            empleado_actual = Empleado.query.filter_by(id_empleado=registro.id_empleado).first()
            if not empleado_actual:
                continue

            clave = f"{empleado_actual.dni}-{empleado_actual.nombres}"
            if clave not in consolidado_map:
                consolidado_map[clave] = {
                    "dni": empleado_actual.dni,
                    "nombres": empleado_actual.nombres,
                    "cargo": empleado_actual.cargo,
                    "area_global": empleado_actual.area or "-",
                    "area": empleado_actual.area or "-",
                    "fechas": {fecha: "" for fecha in fechas},
                    "areas_dia": {fecha: empleado_actual.area or "-" for fecha in fechas}
                }

            fecha_registro = registro.fec_asist.strftime('%Y-%m-%d') if registro.fec_asist else None
            if fecha_registro in fechas:
                if tipo == "asistencias":
                    consolidado_map[clave]["fechas"][fecha_registro] = registro.estado or ""
                elif tipo == "pasajes":
                    consolidado_map[clave]["fechas"][fecha_registro] = registro.pasajes or "0"
                consolidado_map[clave]["areas_dia"][fecha_registro] = area_nombre_map.get(nombre_area, nombre_area)

    consolidado_lista = []
    for idx, empleado in enumerate(sorted(consolidado_map.values(), key=lambda x: x["area"]), start=1):
        fila = {
            "n": idx,
            "dni": empleado["dni"],
            "nombres": empleado["nombres"],
            "cargo": empleado["cargo"],
            "area_global": empleado["area_global"],
            "area": empleado["area"],
        }
        for fecha in fechas:
            if tipo == "asistencias":
                fila[fecha] = {
                    "estado": empleado["fechas"][fecha],
                    "area_dia": empleado["areas_dia"][fecha]
                }
            elif tipo == "pasajes":
                fila[fecha] = {
                    "pasajes": empleado["fechas"][fecha],
                    "area_dia": empleado["areas_dia"][fecha]
                }
        consolidado_lista.append(fila)

    return jsonify(consolidado_lista)

### CONSOLIDADO ###

@app.route('/api/getAsistenciaConsolidada', methods=['POST'])
def get_asistencia_consolidada():
    db.session.expire_all()  # refrescar datos

    data = request.get_json()
    if not data or 'fechaInicio' not in data or 'fechaFin' not in data:
        return jsonify({"error": "Se requieren los campos 'fechaInicio' y 'fechaFin'"}), 400

    fechaInicio = data['fechaInicio']
    fechaFin = data['fechaFin']

    if datetime.strptime(fechaInicio, "%Y-%m-%d") > datetime.strptime(fechaFin, "%Y-%m-%d"):
        return jsonify({"error": "La fecha de inicio no puede ser mayor que la fecha de fin"}), 400

    # 🔹 Mapeo para normalizar nombres de área según el frontend
    area_nombre_map = {
        "lectura": "TOMA DE ESTADO",
        "catastro": "CATASTRO",
        "persuasivas": "PERSUASIVAS",
        "inspecciones": "INSPECCIONES",
        "distribucion": "DISTRIBUCION",
        "medidores": "MEDICION",
        "norte": "NORTE"
    }

    modelos = [
        (EmpleadoLectura, "lectura"),
        (EmpleadoCatastro, "catastro"),
        (EmpleadoPersuasivas, "persuasivas"),
        (EmpleadoInspecciones, "inspecciones"),
        (EmpleadoDistribucion, "distribucion"),
        (EmpleadoMedidores, "medidores"),
        (EmpleadoNorte, "norte"),
    ]

    fechas = [
        (datetime.strptime(fechaInicio, "%Y-%m-%d") + timedelta(days=i)).strftime("%Y-%m-%d")
        for i in range((datetime.strptime(fechaFin, "%Y-%m-%d") - datetime.strptime(fechaInicio, "%Y-%m-%d")).days + 1)
    ]

    consolidado_map = {}

    for modelo, nombre_area in modelos:
        registros = modelo.query.filter(modelo.fec_asist.between(fechaInicio, fechaFin)).all()

        for registro in registros:
            empleado_actual = Empleado.query.filter_by(id_empleado=registro.id_empleado).first()
            if not empleado_actual:
                continue  # ignorar empleados eliminados

            clave = f"{empleado_actual.dni}-{empleado_actual.nombres}"
            if clave not in consolidado_map:
                consolidado_map[clave] = {
                    "dni": empleado_actual.dni,
                    "nombres": empleado_actual.nombres,
                    "cargo": empleado_actual.cargo,
                    "area_global": empleado_actual.area or "-",
                    "area": empleado_actual.area or "-",
                    "fechas": {fecha: "" for fecha in fechas},
                    "areas_dia": {fecha: empleado_actual.area or "-" for fecha in fechas}
                }

            fecha_registro = registro.fec_asist.strftime('%Y-%m-%d') if registro.fec_asist else None
            if fecha_registro in fechas:
                consolidado_map[clave]["fechas"][fecha_registro] = registro.estado or "-"
                # 🔹 Normalizar el área para que coincida con coloresAreas
                consolidado_map[clave]["areas_dia"][fecha_registro] = area_nombre_map.get(nombre_area, nombre_area)

    # Convertir a lista para el frontend
    consolidado_lista = []
    for idx, empleado in enumerate(sorted(consolidado_map.values(), key=lambda x: x["area"]), start=1):
        fila = {
            "n": idx,
            "dni": empleado["dni"],
            "nombres": empleado["nombres"],
            "cargo": empleado["cargo"],
            "area_global": empleado["area_global"],
            "area": empleado["area"],
        }
        # Agregamos fechas y áreas por día
        for fecha in fechas:
            fila[fecha] = {
                "estado": empleado["fechas"][fecha],
                "area_dia": empleado["areas_dia"][fecha]
            }
        consolidado_lista.append(fila)

    return jsonify(consolidado_lista)


#############BOLETAS################
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, 'templates_excel', 'BOLETA.xlsx')
TEMP_PDF_DIR = os.path.join(BASE_DIR, 'static', 'temp_pdfs')
FINAL_PDF_DIR = os.path.join(BASE_DIR, 'static', 'pdfs')

os.makedirs(TEMP_PDF_DIR, exist_ok=True)
os.makedirs(FINAL_PDF_DIR, exist_ok=True)

def get_latest_pdf(directory):
    pdf_files = [f for f in os.listdir(directory) if f.endswith(".pdf")]
    if not pdf_files:
        return None
    pdf_files.sort(key=lambda f: os.path.getctime(os.path.join(directory, f)), reverse=True)
    return os.path.join(directory, pdf_files[0])

@app.route('/generar_boletas', methods=['POST'])
def generar_boletas():
    mes = request.form.get('mes')
    fecha_pago = request.form.get('fecha')
    file = request.files.get('file')

    if not mes or not fecha_pago or not file:
        return "Faltan datos requeridos.", 400

    if not os.path.exists(TEMPLATE_PATH):
        return "Plantilla no encontrada.", 404

    try:
        data_wb = load_workbook(BytesIO(file.read()))
        data_ws = data_wb.active

        temp_pdf_files = []

        for index, row in enumerate(data_ws.iter_rows(min_row=2, max_row=data_ws.max_row), start=1):
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb.active

            periodo = f'PERIODO 2025 - {mes.upper()}'
            ws.merge_cells('B15:G15')
            ws.cell(row=15, column=2).value = periodo
            ws['H5'].value = fecha_pago

            data_map = {
                'C': 'C7', 'E': 'C8', 'F': 'C9', 'B': 'C10', 'D': 'C11',
                'G': 'G7','H': 'G8', 'N': 'G10', 'Y': 'G9', 'O': 'G9', 'Z': 'D17',
                'AM': 'D27', 'AW': 'D28', 'AX': 'D29', 'AU': ['G25', 'G29'], 'AY': ['H25', 'H29'], 'AW': ['H23', 'D28']
            }

            # Obtener los valores de las columnas necesarias
            col_aa_index = column_index_from_string('AA') - 1  # Índice de columna AA (Asignación Familiar)
            col_w_index = column_index_from_string('W') - 1    # Índice de columna W (Domingos Trabajados)
            col_al_index = column_index_from_string('AL') - 1  # Índice de columna AH (Monto Domingos)
            col_v_index = column_index_from_string('V') - 1    # Índice de columna V (Feriados Laborados)
            col_ak_index = column_index_from_string('AK') - 1  # Índice de columna AG (Monto Feriados)
            col_an_index = column_index_from_string('AN') - 1  # Índice de columna AJ (Bonificación Afecta)
            col_u_index = column_index_from_string('U') - 1  # Índice de columna AJ (Monto Horas extras)
            col_aj_index = column_index_from_string('AJ') - 1  # Índice de columna AJ (Horas extras)
            col_af_index = column_index_from_string('AF') - 1  # Índice de columna AF (GRATI TRUNCAS)
            col_ag_index = column_index_from_string('AG') - 1  # Índice de columna AG (VACAS TRUNCAS)
            col_ah_index = column_index_from_string('AH') - 1  # Índice de columna AH (CTS TRUNCAS)
            col_q_index = column_index_from_string('Q') - 1  # Índice de columna Q (DESCANSO MEDICO)
            col_r_index = column_index_from_string('R') - 1  # Índice de columna R (LICENCIA PATERNIDAD)
            col_s_index = column_index_from_string('S') - 1  # Índice de columna S (LICENCIA PATERNIDAD)

            col_ao_index = column_index_from_string('AO') - 1  # Índice de columna AO (Aport. Oblig. Prima)
            col_ap_index = column_index_from_string('AP') - 1  # Índice de columna AP (Seguro Oblig. Prima)
            col_ar_index = column_index_from_string('AR') - 1  # Índice de columna AR (Sist. Nac. Pens. DL 19990)
            col_as_index = column_index_from_string('AS') - 1  # Índice de columna AS (Rent. 5ta. Cat. Reten.)
            col_at_index = column_index_from_string('AT') - 1  # Índice de columna AT (Adelantos)

            asig_familiar = row[col_aa_index].value if col_aa_index < len(row) else None
            domingos_trabajados = row[col_w_index].value if col_w_index < len(row) else None
            monto_domingos = row[col_al_index].value if col_al_index < len(row) else None
            feriados_laborados = row[col_v_index].value if col_v_index < len(row) else None
            monto_feriados = row[col_ak_index].value if col_ak_index < len(row) else None
            bonificacion_afecta = row[col_an_index].value if col_an_index < len(row) else None
            monto_extras = row[col_u_index].value if col_u_index < len(row) else None
            horas_extras = row[col_aj_index].value if col_aj_index < len(row) else None
            grati_truncas = row[col_af_index].value if col_af_index < len(row) else None
            vacas_truncas = row[col_ag_index].value if col_ag_index < len(row) else None
            cts_truncas = row[col_ah_index].value if col_ah_index < len(row) else None
            descanso_medico = row[col_q_index].value if col_q_index < len(row) else None
            licencia_paternidad = row[col_r_index].value if col_r_index < len(row) else None
            fallecimiento = row[col_s_index].value if col_s_index < len(row) else None

            aport_oblig_prima = row[col_ao_index].value if col_ao_index < len(row) else 0
            seguro_oblig_prima = row[col_ap_index].value if col_ap_index < len(row) else 0
            sist_nac_pension = row[col_ar_index].value if col_ar_index < len(row) else 0
            renta_5ta = row[col_as_index].value if col_as_index < len(row) else None
            adelantos = row[col_at_index].value if col_at_index < len(row) else None

            # Validaciones para B, C y D (beneficios laborales)
            row_offset = 18  # Comenzamos en la fila 18

            if asig_familiar:  
                ws[f'B{row_offset}'].value = 'ASIG. FAMILIAR LEY NRO. 25129'
                ws[f'D{row_offset}'].value = asig_familiar
                row_offset += 1  

            if domingos_trabajados:  
                ws[f'B{row_offset}'].value = 'DOMINGOS TRABAJADOS'
                ws[f'C{row_offset}'].value = domingos_trabajados
                ws[f'D{row_offset}'].value = monto_domingos if monto_domingos else ""
                row_offset += 1  

            if feriados_laborados:  
                ws[f'B{row_offset}'].value = 'FERIADOS LABORADOS'
                ws[f'C{row_offset}'].value = feriados_laborados
                ws[f'D{row_offset}'].value = monto_feriados if monto_feriados else ""
                row_offset += 1  
            
            if horas_extras:  
                ws[f'B{row_offset}'].value = 'HORAS EXTRAS'
                ws[f'C{row_offset}'].value = horas_extras
                ws[f'D{row_offset}'].value = monto_extras if monto_extras else ""
                row_offset += 1 

            if bonificacion_afecta:  
                ws[f'B{row_offset}'].value = 'BONIFICACIÓN AFECTA'
                ws[f'D{row_offset}'].value = bonificacion_afecta
                row_offset += 1 

            if grati_truncas:  
                ws[f'B{row_offset}'].value = 'GRATIFICACIONES TRUNCAS'
                ws[f'D{row_offset}'].value = grati_truncas
                row_offset += 1

            if vacas_truncas:  
                ws[f'B{row_offset}'].value = 'VACACIONES TRUNCAS'
                ws[f'D{row_offset}'].value = vacas_truncas
                row_offset += 1

            if cts_truncas:  
                ws[f'B{row_offset}'].value = 'CTS TRUNCAS'
                ws[f'D{row_offset}'].value = cts_truncas
                row_offset += 1

            if descanso_medico:  
                ws[f'B{row_offset}'].value = 'DESCANSO MEDICO'
                ws[f'C{row_offset}'].value = descanso_medico
                row_offset += 1 
            
            if licencia_paternidad:  
                ws[f'B{row_offset}'].value = 'LIC. PATERNIDAD'
                ws[f'C{row_offset}'].value = licencia_paternidad
                row_offset += 1

            if fallecimiento:  
                ws[f'B{row_offset}'].value = 'DESC. FALLECIMIENTO'
                ws[f'C{row_offset}'].value = fallecimiento
                row_offset += 1


            # Validaciones para E y H (aportes y descuentos)
            row_e_offset = 17  # Empezamos en la fila 17 para la columna E y H

            if aport_oblig_prima and aport_oblig_prima != 0:  
                ws[f'E{row_e_offset}'].value = 'APORT. OBLIGT. PRIMA'
                ws[f'H{row_e_offset}'].value = aport_oblig_prima
                row_e_offset += 1  

            if seguro_oblig_prima and seguro_oblig_prima != 0:  
                ws[f'E{row_e_offset}'].value = 'SEGURO OBLIGT. PRIMA'
                ws[f'H{row_e_offset}'].value = seguro_oblig_prima
                row_e_offset += 1  

            if (not aport_oblig_prima or aport_oblig_prima == 0) and (not seguro_oblig_prima or seguro_oblig_prima == 0):  
                if sist_nac_pension and sist_nac_pension != 0:  
                    ws[f'E{row_e_offset}'].value = 'SIST. NAC. DE PENSI. DL 19990'
                    ws[f'H{row_e_offset}'].value = sist_nac_pension
                    row_e_offset += 1  

            if renta_5ta:  
                ws[f'E{row_e_offset}'].value = 'RENT. 5TA. CAT. RETEN.'
                ws[f'H{row_e_offset}'].value = renta_5ta
                row_e_offset += 1  

            if adelantos:  
                ws[f'E{row_e_offset}'].value = 'ADELANTOS'
                ws[f'H{row_e_offset}'].value = adelantos




            for col, target in data_map.items():
                col_index = column_index_from_string(col) - 1
                if col_index < len(row):  # Verificar que la columna existe en la fila
                    value = row[col_index].value
                    print(f"Columna: {col} (Índice {col_index}) -> Valor obtenido: {value}")
                    if value:
                        targets = target if isinstance(target, list) else [target]
                        for tgt in targets:
                            cell = ws[tgt]
                            if isinstance(cell, MergedCell):
                                for merged_range in ws.merged_cells.ranges:
                                    if tgt in merged_range.coord:
                                        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                        top_left_cell.value = value
                                        break
                            else:
                                cell.value = value

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_excel:
                wb.save(temp_excel.name)
                temp_excel_path = temp_excel.name

            before_conversion = set(os.listdir(TEMP_PDF_DIR))
            subprocess.run([
                r'C:\Program Files\LibreOffice\program\soffice.exe',
                '--headless', '--convert-to', 'pdf:calc_pdf_Export', '--outdir',
                TEMP_PDF_DIR, temp_excel_path
            ], capture_output=True, text=True)

            time.sleep(2)
            after_conversion = set(os.listdir(TEMP_PDF_DIR))

            new_files = after_conversion - before_conversion
            if new_files:
                temp_pdf_path = os.path.join(TEMP_PDF_DIR, list(new_files)[0])
                temp_pdf_files.append(temp_pdf_path)
                print(f"✅ Archivo PDF generado: {temp_pdf_path}")
            else:
                print(f"❌ Error: No se generó el PDF para {temp_excel_path}")

            os.remove(temp_excel_path)

        final_pdf_path = os.path.join(FINAL_PDF_DIR, f'boletas_{mes}.pdf')
        merger = PdfMerger()

        for pdf_file in temp_pdf_files:
            if os.path.exists(pdf_file):
                merger.append(pdf_file)

        merger.write(final_pdf_path)
        merger.close()

        @after_this_request
        def remove_temp_files(response):
            try:
                for pdf_file in temp_pdf_files:
                    if os.path.exists(pdf_file):
                        os.remove(pdf_file)
            except Exception as e:
                print(f"❌ Error al eliminar archivos temporales: {e}")
            return response

        return send_file(
            final_pdf_path,
            download_name=f'boletas_{mes}.pdf',
            mimetype='application/pdf',
            as_attachment=False
        )

    except Exception as e:
        print(f"❌ Error en el procesamiento: {traceback.format_exc()}")
        return "Error en el procesamiento. Ver logs.", 500


####### REPORTE LECTURAS #######
COLUMNAS_REQUERIDAS = {
    'CLICODFAC': 2,
    'NOMBRE': 3,
    'URBANIZAC': 4,
    'CALLE': 5,
    'CLIMUNRO': 6,
    'MEDCODYGO': 7,
    'LECTURA': 8,
    'FECLEC': 9,
    'OBS1': 11,
    'OBS2': 12,
    'REFUBIME': 15,
    'NEWMED': 16,
    'CICLO': 18,
    'CARGA': 19,
    'ORDENRUTA': 20,
    'TIPOLECTURA': 21,
    'NOMBREOPERADOR': 26,
    'PROMEDIOSEDALIB': 34
}

# Intentamos usar locale en español
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')  # Linux/macOS
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')  # Windows
    except:
        pass  # Si falla, usamos traducción manual

# Traducción manual de nombres de meses
MESES_ES = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
    5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
    9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
}

def combinatoria_xlsx(dfs):
    # Datos fijos
    cols_fijas = ['CLICODFAC', 'NOMBRE', 'URBANIZAC', 'CALLE', 'CICLO', 'CLIMUNRO', 'MEDCODYGO']
    
    # Variables divididas por prioridad de orden
    campos_primero = ['LECTURA', 'OBS1']
    campos_restantes = ['OBS2', 'REFUBIME', 'NEWMED', 'CARGA', 'ORDENRUTA', 'TIPOLECTURA', 'NOMBREOPERADOR']

    meses_a_actualizar = set()

    # Extraer todos los meses y años de cada DataFrame nuevo que vas a subir
    for df in dfs:
        df['FECLEC'] = pd.to_datetime(df['FECLEC'], errors='coerce')
        df['MES'] = df['FECLEC'].dt.month
        df['ANIO'] = df['FECLEC'].dt.year
        meses_a_actualizar.update(set(zip(df['MES'], df['ANIO'])))

    # Diagnóstico antes de concatenar
    for i, df in enumerate(dfs):
        print(f"DataFrame {i}: total filas = {len(df)}, filas con CLICODFAC = {df['CLICODFAC'].notna().sum()}")
        print(f"DataFrame {i} valores únicos CLICODFAC (primeros 10):")
        print(df['CLICODFAC'].unique()[:10])

    df_all = pd.concat(dfs, ignore_index=True)
    base = df_all.dropna(subset=['CLICODFAC'])[cols_fijas].drop_duplicates(subset=['CLICODFAC']).set_index('CLICODFAC')

    # Aquí pones el bloque para debug
    print("Clientes únicos en df_all:", df_all['CLICODFAC'].nunique())
    print("Filas df_all:", df_all.shape[0])

    # Mostrar filas con valores nulos en cols_fijas
    print("Filas con valores nulos en columnas fijas:")
    print(df_all[df_all[cols_fijas].isnull().any(axis=1)][cols_fijas].head(10))

    # Procesar fechas
    df_all['FECLEC'] = pd.to_datetime(df_all['FECLEC'], errors='coerce')
    df_all['MES'] = df_all['FECLEC'].dt.month
    df_all['ANIO'] = df_all['FECLEC'].dt.year
    df_all['MES_NOMBRE'] = df_all['MES'].apply(lambda x: MESES_ES.get(x, ''))

    # Base con datos fijos
    base = df_all[cols_fijas].drop_duplicates(subset=['CLICODFAC'], keep='last')

    data_final = base.copy()

    # Obtener lista ordenada de meses
    meses_ordenados = df_all[['MES', 'ANIO', 'MES_NOMBRE']].drop_duplicates()
    meses_ordenados = meses_ordenados.sort_values(by=['ANIO', 'MES'])

    # Primero agregar campos LECTURA y OBS1 por mes
    for _, row in meses_ordenados.iterrows():
        mes = row['MES']
        anio = row['ANIO']
        mes_nombre = row['MES_NOMBRE']
        df_mes = df_all[(df_all['MES'] == mes) & (df_all['ANIO'] == anio)]

        for campo in campos_primero:
            pivot = df_mes.pivot_table(index='CLICODFAC', values=campo, aggfunc='first')
            if not pivot.empty and pivot.shape[1] == 1:
                pivot.columns = [f"{campo} {mes_nombre}"]
                
                # Si la columna ya existe en data_final, combínalas
                col = pivot.columns[0]
                if col in data_final.columns:
                    # Combina: tomar valores no nulos de pivot, si no usar data_final
                    data_final[col] = data_final[col].combine_first(pivot[col])
                else:
                    data_final = data_final.join(pivot, how='left')

            else:
                print(f"⚠️ Se omitió el campo {campo} para {mes_nombre} porque pivot está vacío o mal formado.")



    # Luego agregar el resto de campos por mes
    for _, row in meses_ordenados.iterrows():
        mes = row['MES']
        anio = row['ANIO']
        mes_nombre = row['MES_NOMBRE']
        df_mes = df_all[(df_all['MES'] == mes) & (df_all['ANIO'] == anio)]

        for campo in campos_restantes:
            pivot = df_mes.pivot_table(index='CLICODFAC', values=campo, aggfunc='first')
            if not pivot.empty and pivot.shape[1] == 1:
                pivot.columns = [f"{campo} {mes_nombre}"]
                
                col = pivot.columns[0]
                if col in data_final.columns:
                    data_final[col] = data_final[col].combine_first(pivot[col])
                else:
                    data_final = data_final.join(pivot, how='left')

            else:
                print(f"⚠️ Se omitió el campo {campo} para {mes_nombre} porque pivot está vacío o mal formado.")


    # --- PROMEDIOSEDALIB: Calcular promedio últimos 5 meses por cliente ---
    # Crear columna auxiliar año-mes para filtrar últimos 5 meses
    df_all['YM'] = df_all['FECLEC'].dt.to_period('M')
    meses_disponibles = df_all['YM'].dropna().unique()
    if len(meses_disponibles) > 0:
        mes_max = df_all['YM'].max()
        ultimos_5_meses = sorted([mes_max - i for i in range(5)])

        df_prom = df_all[df_all['YM'].isin(ultimos_5_meses)][['CLICODFAC', 'PROMEDIOSEDALIB', 'YM']]

        # Convertir PROMEDIOSEDALIB a numérico, ignorar errores
        df_prom['PROMEDIOSEDALIB'] = pd.to_numeric(df_prom['PROMEDIOSEDALIB'], errors='coerce')

        # Calcular promedio por CLICODFAC (ignorando NaN)
        prom_df = df_prom.groupby('CLICODFAC')['PROMEDIOSEDALIB'].mean().round(2).to_frame()

        # Unir al resultado final
        data_final = data_final.join(prom_df, how='left')
    else:
        # Si no hay meses, solo añadir columna vacía
        data_final['PROMEDIOSEDALIB'] = None


    ## --- Calcular CONSUMO/DF correctamente ---
    meses_ordenados_list = meses_ordenados.sort_values(by=['ANIO', 'MES']).reset_index(drop=True)

    for i in range(1, len(meses_ordenados_list)):
        mes_actual = meses_ordenados_list.loc[i, 'MES_NOMBRE']
        mes_anterior = meses_ordenados_list.loc[i - 1, 'MES_NOMBRE']

        lectura_actual_col = f"LECTURA {mes_actual}"
        lectura_anterior_col = f"LECTURA {mes_anterior}"

        if lectura_actual_col in data_final.columns and lectura_anterior_col in data_final.columns:
            print(f"Usando columnas: {lectura_anterior_col} → {lectura_actual_col}")

            # Asegurar valores numéricos
            lectura_actual = pd.to_numeric(data_final[lectura_actual_col], errors='coerce')
            lectura_anterior = pd.to_numeric(data_final[lectura_anterior_col], errors='coerce')

            consumo_df = lectura_actual - lectura_anterior

            print("Ejemplo de cálculo:")
            print(pd.DataFrame({
                lectura_anterior_col: lectura_anterior,
                lectura_actual_col: lectura_actual,
                'CONSUMO/DF': consumo_df
            }).head(10))

            # Borrar columna previa si existe
            if 'CONSUMO/DF' in data_final.columns:
                data_final.drop(columns=['CONSUMO/DF'], inplace=True)

            # Asegurar mínimo 56 columnas
            while len(data_final.columns) < 56:
                data_final[f'FILL_{len(data_final.columns)}'] = None

            # Insertar en posición exacta (índice 56 = columna 57 en Excel = BE)
            data_final.insert(56, 'CONSUMO/DF', consumo_df)

            break  # Solo una vez

    # Paso final: Eliminar columnas FILL_ si se crearon
    data_final = data_final.loc[:, ~data_final.columns.str.startswith('FILL_')]

    data_final.reset_index(inplace=True)

    print("data_final.shape:", data_final.shape)
    print(data_final.head())
    print(data_final.columns)

    
    return data_final

@app.route('/upload', methods=['POST'])
def upload_files():
    area = request.form.get('areas')
    if area != 'TOMA DE ESTADO':
        return "Área no soportada", 400

    files = request.files.getlist('file')
    if not files:
        return "No se subieron archivos", 400

    dfs = []
    total_registros = 0
    nuevos_o_cambios = 0

    for file in files:
        try:
            df = pd.read_excel(file, dtype=str, header=0)
        except Exception as e:
            return f"Error al leer archivo {file.filename}: {str(e)}", 400

        max_col = max(COLUMNAS_REQUERIDAS.values())
        if df.shape[1] < max_col:
            return f"Archivo {file.filename} no tiene suficientes columnas.", 400

        cols_indices = [i-1 for i in COLUMNAS_REQUERIDAS.values()]
        df_subset = df.iloc[:, cols_indices].copy()
        df_subset.columns = COLUMNAS_REQUERIDAS.keys()
        df_subset['FECLEC'] = pd.to_datetime(df_subset['FECLEC'], dayfirst=True, errors='coerce')

        dfs.append(df_subset)

        # -- NUEVO: Borrar registros antiguos del mismo mes/año antes de insertar --
        # Extraer todos los meses y años del df_subset
        meses_anios = df_subset['FECLEC'].dropna().dt.to_period('M').unique()
        for periodo in meses_anios:
            mes = periodo.month
            anio = periodo.year
            print(f"Eliminando registros previos de {mes}/{anio} para archivo {file.filename}")
            db.session.query(ReporteLectura).filter(
                db.extract('month', ReporteLectura.FECLEC) == mes,
                db.extract('year', ReporteLectura.FECLEC) == anio
            ).delete()
        db.session.commit()
        # -- FIN NUEVO --

        for _, row in df_subset.iterrows():
            total_registros += 1
            clean_row = {k: (None if pd.isna(v) else v) for k, v in row.items()}

            feclec = clean_row['FECLEC']
            clicodfac = str(clean_row['CLICODFAC']).strip() if clean_row['CLICODFAC'] else None

            if not feclec or not clicodfac:
                continue

            mes = feclec.month
            anio = feclec.year

            # Buscar si ya existe un registro con mismo CLICODFAC y MES/AÑO
            existente = db.session.query(ReporteLectura).filter(
                ReporteLectura.CLICODFAC == clicodfac,
                db.extract('month', ReporteLectura.FECLEC) == mes,
                db.extract('year', ReporteLectura.FECLEC) == anio
            ).first()

            print(f"Procesando: {clicodfac} - {feclec.strftime('%Y-%m-%d')}")

            if existente:
                hay_cambio = False
                for campo, valor_nuevo in clean_row.items():
                    valor_existente = getattr(existente, campo)
                    if str(valor_existente) != str(valor_nuevo):
                        setattr(existente, campo, valor_nuevo)
                        hay_cambio = True

                if hay_cambio:
                    nuevos_o_cambios += 1
                print("Ya existe registro, actualizando")

            else:
                nuevo = ReporteLectura(**clean_row)
                db.session.add(nuevo)
                nuevos_o_cambios += 1
                print("Nuevo registro, agregando")

        db.session.commit()

        hoy = datetime.now()
        fecha_limite = hoy - relativedelta(months=5)

        # Leer todos los datos desde la base de datos y convertirlos en un DataFrame
        registros = ReporteLectura.query.filter(ReporteLectura.FECLEC >= fecha_limite).all()
        df_db = pd.DataFrame([{
            'CLICODFAC': r.CLICODFAC,
            'NOMBRE': r.NOMBRE,
            'URBANIZAC': r.URBANIZAC,
            'CALLE': r.CALLE,
            'CLIMUNRO': r.CLIMUNRO,
            'MEDCODYGO': r.MEDCODYGO,
            'LECTURA': r.LECTURA,
            'FECLEC': r.FECLEC,
            'OBS1': r.OBS1,
            'OBS2': r.OBS2,
            'REFUBIME': r.REFUBIME,
            'NEWMED': r.NEWMED,
            'CICLO': r.CICLO,
            'CARGA': r.CARGA,
            'ORDENRUTA': r.ORDENRUTA,
            'TIPOLECTURA': r.TIPOLECTURA,
            'NOMBREOPERADOR': r.NOMBREOPERADOR,
            'PROMEDIOSEDALIB': r.PROMEDIOSEDALIB
        } for r in registros])

        # Asegurar que FECLEC sea datetime
        df_db['FECLEC'] = pd.to_datetime(df_db['FECLEC'], errors='coerce')

        # Crear columna auxiliar año-mes
        df_db['YM'] = df_db['FECLEC'].dt.to_period('M')

        # Identificar el mes más reciente
        mes_max = df_db['YM'].max()

        # Obtener los últimos 5 meses incluyendo el más reciente
        ultimos_5_meses = [(mes_max - i) for i in range(5)]
        ultimos_5_meses = sorted(ultimos_5_meses)

        # Filtrar el DataFrame
        df_filtrado = df_db[df_db['YM'].isin(ultimos_5_meses)].drop(columns='YM')

        # Generar análisis
        combined_df = combinatoria_xlsx([df_filtrado])
    
    # Agregar columna "FRECUENCIA LLEGADA"
    carga_cols = [col for col in combined_df.columns if col.startswith('CARGA ')]

    def obtener_frecuencia_llegada(row):
        meses_presentes = [col.replace('CARGA ', '') for col in carga_cols if pd.notna(row[col]) and str(row[col]).strip() != '']
        if not meses_presentes:
            return ''
        elif len(meses_presentes) == len(carga_cols):
            return 'COMPLETO'
        elif len(meses_presentes) == 1:
            return f"SOLO {meses_presentes[0]}"
        else:
            return f"SOLO {'-'.join(meses_presentes)}"

    combined_df['FRECUENCIA LLEGADA'] = combined_df.apply(obtener_frecuencia_llegada, axis=1)

    # 1. Detectar columnas tipo "LECTURA MES"
    lectura_cols = [col for col in combined_df.columns if re.match(r'^LECTURA\s+\w+', col)]

    # 2. Evaluar continuidad para cada fila
    def evaluar_continuidad(row):
        valores = [str(row[col]).strip() for col in lectura_cols]
        estados = [bool(v and v != 'nan') for v in valores]

        if all(estados):
            return "LECTURA CONTINUA"
        if not any(estados):
            return "SIN LECTURA"
        if estados[0] and not estados[1] and estados[2:4] == [True, True]:
            return "LECTURA CONTINUA 1"
        if not estados[0] and estados[1] and estados[2] == False:
            return "LECTURA DISCONTINUA 1"
        if estados[0] and estados[1] and not estados[2]:
            return "LECTURA DISCONTINUA 2"
        if not estados[0] and not estados[1] and estados[2]:
            return "PRIMERA LECTURA"
        if estados[0] and not estados[1] and not estados[2]:
            return "LECTURA SOLO FEBRERO"
        return "DISCONTINUO"

    # 3. Agregar la columna nueva
    combined_df["CONTINUIDAD TOMA"] = combined_df.apply(evaluar_continuidad, axis=1)

    # 1. Detectar columnas tipo "OBS1 MES"
    obs1_cols = [col for col in combined_df.columns if re.match(r'^OBS1\s+\w+', col)]

    # 2. Evaluar continuidad de anomalías para cada fila
    def evaluar_cont_anomalia(row):
        valores = [str(row[col]).strip() for col in obs1_cols if str(row[col]).strip() != "" and str(row[col]).strip().lower() != "nan"]
        if len(valores) == 0:
            return "UNICA"
        if all(v == valores[0] for v in valores):
            return "UNICA"
        return "DISTINTA"

    # 3. Agregar la columna nueva
    combined_df["CONTINUIDAD ANOMALIA"] = combined_df.apply(evaluar_cont_anomalia, axis=1)
    

    # --- Reubicar PROMEDIOSEDALIB y CONSUMO/DF correctamente ---
    if 'PROMEDIOSEDALIB' in combined_df.columns and 'CONSUMO/DF' in combined_df.columns:
        # Extraer las columnas que quieres mover
        promedio_col = combined_df['PROMEDIOSEDALIB']
        consumo_col = combined_df['CONSUMO/DF']

        # Eliminar las columnas del DataFrame
        combined_df = combined_df.drop(columns=['PROMEDIOSEDALIB', 'CONSUMO/DF'])

        # Insertar en la posición deseada
        insert_idx = min(55, len(combined_df.columns))
        combined_df.insert(insert_idx, 'PROMEDIOSEDALIB', promedio_col.reset_index(drop=True))

        combined_df.insert(len(combined_df.columns), 'CONSUMO/DF', consumo_col.reset_index(drop=True))
        # BE



    

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False, sheet_name='REPORTE')
        workbook = writer.book
        worksheet = writer.sheets['REPORTE']

        # Ajustar ancho de columnas automáticamente según el contenido
        for i, column in enumerate(combined_df.columns, 1):
            max_length = max(
                combined_df[column].astype(str).map(len).max(),
                len(str(column))
            )
            worksheet.column_dimensions[get_column_letter(i)].width = max_length + 2  # Margen extra

        # Ocultar columnas de la R (col 18) a la AZ (col 52)
        for col_idx in range(18, 53):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].hidden = True

            # Mover columna una sola vez
            col_index = combined_df.columns.get_loc("CONTINUIDAD TOMA") + 1
            if col_index != 54:
                worksheet.move_range(
                    f"{get_column_letter(col_index)}1:{get_column_letter(col_index)}{combined_df.shape[0]+1}",
                    cols=54 - col_index
                )

    # Guardar archivo al disco
    output.seek(0)
    output_dir = r"C:\RADIAN\ASISTENCIAS\archivos_generados"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, 'combinado_toma_de_estado.xlsx')
    with open(output_path, 'wb') as f:
        f.write(output.getvalue())

    return jsonify({
        'message': 'Análisis generado correctamente',
        'download_url': f'/descargar-analisis?archivo=combinado_toma_de_estado.xlsx'
    })


@app.route('/descargar-analisis')
def descargar_analisis():
    archivo = request.args.get('archivo')
    ruta = os.path.join(r"C:\RADIAN\ASISTENCIAS\archivos_generados", archivo)
    print("Buscando archivo en:", ruta)  # Para depuración en consola
    if os.path.exists(ruta):
        return send_file(ruta, as_attachment=True)
    return "Archivo no encontrado", 404


##FOTO CONSULTA
BASE_FOLDER_ORDENES = r"\\192.168.1.201\images\ordenes"
BASE_FOLDER_LECTURAS = r"\\192.168.1.201\images\lecturas"


# Leyenda completa
LEYENDA = {
    "29": "CIERRE ALCANTARILLADO",
    "63": "VERIFICACION ORDEN DRASTICO",
    "27": "CIERRE SIMPLE",
    "28": "CIERRE DRASTICO",
    "73": "VERIFICACION DE ACCION COERCITIVA",
    "64": "VERIFICACION ORDEN OBSTURACION",
    "30": "REAPERTURA SIMPLE",
    "37": "SELLADO DE ALCANTARILLADO",
    "39": "VERIFICACION DE ORDEN",
    "33": "LEVANTAMIENTO ALCANTARILLADO",
    "32": "REAPERTURA ALCANTARILLADO",
    "31": "REAPERTURA DRASTICA",
    "41": "LEVANTAMIENTO ACUEDUCTO",
    "10": "CATASTRO",
    "72": "CATASTRO-FICHA LEVANTADA",
    "22": "SUPERVISION TOMA DE ESTADO",
    "49": "INSTALACION MEDIDORES",
    "56": "NUEVOS SUMINISTROS",
    "45": "MANTENIMIENTO MEDIDORES",
    "47": "VERIFICACION POSTERIOR MEDIDORES",
    "12": "DISTRIBUCION DISPERSO",
    "46": "COMUNICACIÓN VERIFICACION POSTERIOR",
    "48": "COMUNICACIÓN INSTALACION MEDIDORES",
    "50": "COMUNICACION NUEVOS SUMINISTROS",
    "51": "COMUNICACION CARTA VP",
    "52": "COMUNICACION RESULTADO VP",
    "11": "DISTRIBUCION CONTINUO",
    "13": "COMUNICACIÓN TARIFA",
    "14": "COMUNICACIÓN ATENCION",
    "15": "COMUNICACIÓN CONSUMO",
    "16": "COMUNICACIÓN RECUPERO",
    "44": "COMUNICACION MANTENIMIENTO MEDIDORES",
    "17": "COMUNICACIÓN RESOLUCION",
    "23": "COMUNICACION ATIPICAS",
    "3": "INSPECCION",
    "7": "INSPECCION INTERNA ATIPICA USO MULTIPLE",
    "4": "INSPECCION INTERNA RECLAMO USO UNICO",
    "6": "INSPECCION INTERNA ATIPICA USO UNICO",
    "5": "INSPECCION INTERNA RECLAMO USO MULTIPLE",
    "66": "INSPECCION INTERNA TARIFA USO UNICO",
    "67": "INSPECCION INTERNA TARIFA USO MULTIPLE",
    "68": "INSPECCION INTERNA VERIFICACION USO UNICO",
    "69": "INSPECCION INTERNA VERIFICACION USO MULTIPLE",
    "70": "INSPECCION ESPECIAL GEOFONO USO UNICO",
    "71": "INSPECCION ESPECIAL GEOFONO USO MULTIPLE",
    "53": "INSPECCION INTERNA VERIFICACION",
    "59": "INSPECCION EXTERNA MEDICION",
    "55": "INSPECCION TARIFA POR SOLICITUD",
    "21": "INSPECCION EXTERNA RECLAMO",
    "24": "INSPECCION INTERNA TARIFA",
    "54": "INSPECCION EXTERNA VERIFICACION",
    "65": "INSPECCION EXTERNA LEVANTAMIENTO",
    "43": "INSPECCION EXTERNA",
    "19": "INSPECCION INTERNA RECLAMO",
    "18": "INSPECCION INTERNA ATIPICA",
    "9": "INSPECCION EXTERNA ATIPICA",
    "8": "INSPECCION ESPECIAL GEOFONO",
    "36": "REAPERTURA DRASTICA CON PAVIMIENTO",
    "35": "CIERRE DRASTICO SIN PAVIMIENTO",
    "40": "LEVANTAMIENTO ALCANTARILLADO CON PAVIMENTO",
    "34": "CIERRE DRASTICO CON PAVIMIENTO",
    "42": "LEVANTAMIENTO ALCANTARILLADO SIN PAVIMENTO",
    "38": "REAPERTURA DRASTICA SIN PAVIMENTO",
    "26": "LEVANTAMIENTO ACUEDUCTO SIN PAVIMENTO",
    "25": "LEVANTAMIENTO ACUEDUCTO CON PAVIMENTO",
    "928": "SUP CIERRE DRASTICO",
    "973": "SUP VERIFICACION CLIENTE MOROSOS",
    "964": "SUP VERIFICACION ORDEN OBSTURACION",
    "963": "SUP VERIFICACION ORDEN DRASTICO",
    "938": "SUP REAPERTURA DRASTICA SIN PAVIMENTO",
    "936": "SUP REAPERTURA DRASTICA CON PAVIMIENTO",
    "935": "SUP CIERRE DRASTICO SIN PAVIMIENTO",
    "934": "SUP CIERRE DRASTICO CON PAVIMIENTO",
    "941": "SUP LEVANTAMIENTO ACUEDUCTO",
    "939": "SUP VERIFICACION DE ORDEN",
    "937": "SUP SELLADO DE ALCANTARILLADO",
    "933": "SUP LEVANTAMIENTO ALCANTARILLADO",
    "932": "SUP REAPERTURA ALCANTARILLADO",
    "931": "SUP REAPERTURA DRASTICA",
    "930": "SUP REAPERTURA SIMPLE",
    "929": "SUP CIERRE ALCANTARILLADO",
    "927": "SUP CIERRE SIMPLE",
    "910": "SUP CATASTRO",
    "945": "SUP MANTENIMIENTO MEDIDORES",
    "949": "SUP INSTALACION MEDIDORES",
    "947": "SUP VERIFICACION MEDIDORES",
    "950": "SUP COMUNICACION NUEVOS SUMINISTROS",
    "948": "SUP COMUNICACIÓN INSTALACION MEDIDORES",
    "923": "SUP COMUNICACION ATIPICAS",
    "946": "SUP COMUNICACIÓN VERIFICACION POSTERIOR",
    "917": "SUP COMUNICACIÓN RESOLUCION",
    "916": "SUP COMUNICACIÓN RECUPERO",
    "915": "SUP COMUNICACIÓN CONSUMO",
    "914": "SUP COMUNICACIÓN ATENCION",
    "913": "SUP COMUNICACIÓN TARIFA",
    "951": "SUP CARTA VERIFICACION POSTERIOR",
    "952": "SUP RESULTADO VERIFICACION POSTERIOR",
    "911": "SUP DISTRIBUCION CONTINUO",
    "912": "SUP DISTRIBUCION DISPERSO",
    "944": "SUP COMUNICACION MANTENIMIENTO MEDIDORES",
    "943": "SUP INSPECCION EXTERNA",
    "953": "SUP INSPECCION INTERNA RECLAMO VERIFICACION",
    "954": "SUP INSPECCION EXTERNA RECLAMO VERIFICACION",
    "955": "SUP INSPECCION TARIFA POR SOLICITUD",
    "908": "SUP INSPECCION ESPECIAL GEOFONO",
    "909": "SUP INSPECCION EXTERNA ATIPICA",
    "918": "SUP INSPECCION INTERNA ATIPICA",
    "919": "SUP INSPECCION INTERNA RECLAMO",
    "921": "SUP INSPECCION EXTERNA RECLAMO",
    "959": "SUP INSPECCION EXTERNA MEDICION",
    "965": "SUP INSPECCION EXTERNA LEVANTAMIENTO",
    "904": "SUP INSPECCION INTERNA RECLAMO USO UNICO",
    "905": "SUP INSPECCION INTERNA RECLAMO USO MULTIPLE",
    "906": "SUP INSPECCION INTERNA USO UNICO",
    "907": "SUP INSPECCION INTERNA USO MULTIPLE",
    "924": "SUP INSPECCION INTERNA TARIFA",
    "57": "TOMA DE ESTADOS",
    "58": "RELECTURA",
    "60": "GENERICA",
    "956": "SUP NUEVOS SUMINISTROS",
    "957": "SUP TOMA DE ESTADO",
    "962": "SUP GENERICA",
    "960": "SUP GENERICA",
    "61": "GENERICA",
    "62": "GENERICA",
    "961": "SUP GENERICA",
    "958": "SUP RELECTURA"
}

def formato_fecha_lectura(carpeta):
    if carpeta.isdigit() and len(carpeta) == 6:
        anio = carpeta[:4]
        mes = carpeta[4:]
        mes_nombre = calendar.month_name[int(mes)]
        # Capitalizar solo primera letra y dejar resto en minúsculas
        mes_nombre = mes_nombre.capitalize()
        return f"{mes_nombre} - {anio}"
    return carpeta

def buscar_imagenes_por_codigo_v1(codigo):
    print(f"[LOG] Iniciando búsqueda en índice para código: {codigo}")
    resultados_por_carpeta = {}

    index_actual = obtener_index_actualizado()

    codigo = codigo.strip()
    codigo_sin_ceros = codigo.lstrip("0")

    for entry in index_actual:
        if codigo in entry["filename"] or codigo_sin_ceros in entry["filename"]:

            carpeta = entry["carpeta"]  # puede ser "202505" o "72"
            imagen = entry["filename"]

            if carpeta.isdigit() and len(carpeta) == 6:
                import re
                match = re.search(r'_C(\d+)_', imagen)
                if match:
                    subcarpeta = match.group(1)
                    carpeta_completa = os.path.normpath(os.path.join(carpeta, subcarpeta))
                else:
                    carpeta_completa = carpeta
                leyenda = formato_fecha_lectura(carpeta)  # Fecha con mes capitalizado
                clave_agrupacion = carpeta  # Para ordenamiento
            else:
                carpeta_completa = carpeta
                leyenda = entry.get("leyenda") or LEYENDA.get(carpeta) or carpeta
                clave_agrupacion = carpeta_completa

            if clave_agrupacion not in resultados_por_carpeta:
                resultados_por_carpeta[clave_agrupacion] = {
                    "carpeta": carpeta_completa,
                    "leyenda": leyenda,
                    "imagenes": []
                }

            resultados_por_carpeta[clave_agrupacion]["imagenes"].append(imagen)

    # Separar resultados en dos listas: ordenes (leyenda sin guion) y lecturas (leyenda con guion)
    ordenes = []
    lecturas_subgrupos = []
    for k, v in resultados_por_carpeta.items():
        if " - " in v["leyenda"]:
            lecturas_subgrupos.append(v)
        else:
            ordenes.append(v)

    ordenes.sort(key=lambda x: x["leyenda"])
    lecturas_subgrupos.sort(key=lambda x: x["leyenda"])

    resultados = []
    if lecturas_subgrupos:
        resultados.append({
            "leyenda": "LECTURAS",
            "subgrupos": lecturas_subgrupos
        })
    resultados.extend(ordenes)


    print(f"[LOG] Resultados finales ordenados: {resultados}")
    return resultados

def buscar_imagenes_por_codigo(codigo):
    print(f"[LOG] Buscando imágenes en BD para código: {codigo}")

    resultados_por_carpeta = {}

    codigo = codigo.strip()
    codigo_sin_ceros = codigo.lstrip("0")

    # Query a la base de datos
    imagenes = (
        db.session.query(Imagen)
        .filter(
            Imagen.suministro.in_([codigo, codigo_sin_ceros])
            # Imagen.filename.ilike(f"%{codigo}%") |
            # Imagen.filename.ilike(f"%{codigo_sin_ceros}%")
        )
        .all()
    )

    for img in imagenes:
        carpeta = img.carpeta
        imagen = img.filename

        if carpeta.isdigit() and len(carpeta) == 6:
            import re
            match = re.search(r'_C(\d+)_', imagen)
            if match:
                subcarpeta = match.group(1)
                carpeta_completa = os.path.normpath(os.path.join(carpeta, subcarpeta))
            else:
                carpeta_completa = carpeta

            leyenda = formato_fecha_lectura(carpeta)
            clave_agrupacion = carpeta
        else:
            carpeta_completa = carpeta
            leyenda = img.leyenda or LEYENDA.get(carpeta) or carpeta
            clave_agrupacion = carpeta_completa

        if clave_agrupacion not in resultados_por_carpeta:
            resultados_por_carpeta[clave_agrupacion] = {
                "carpeta": carpeta_completa,
                "leyenda": leyenda,
                "imagenes": []
            }

        resultados_por_carpeta[clave_agrupacion]["imagenes"].append(imagen)

    # Separar ordenes y lecturas
    ordenes = []
    lecturas_subgrupos = []

    for v in resultados_por_carpeta.values():
        if " - " in v["leyenda"]:
            lecturas_subgrupos.append(v)
        else:
            ordenes.append(v)

    ordenes.sort(key=lambda x: x["leyenda"])
    lecturas_subgrupos.sort(key=lambda x: x["leyenda"])

    resultados = []
    if lecturas_subgrupos:
        resultados.append({
            "leyenda": "LECTURAS",
            "subgrupos": lecturas_subgrupos
        })

    resultados.extend(ordenes)

    print(f"[LOG] Resultados finales: {resultados}")
    return resultados


@app.route("/buscar", methods=["POST"])
def buscar():
    data = request.get_json()
    codigo = data.get("codigo", "").strip()

    # completar con ceros a la izquierda hasta 11 dígitos
    codigo = codigo.zfill(11)

    print(f"[LOG] Request recibido para buscar código: {codigo}")

    try:
        resultados = buscar_imagenes_por_codigo(codigo)
        print(f"[LOG] Resultados obtenidos para código {codigo}: {resultados}")
        return jsonify({"resultados": resultados})
    except Exception as e:
        print(f"[ERROR] Error durante búsqueda: {e}")
        return jsonify({"error": str(e)}), 500
    

@app.route('/imagen/<path:subpath>/<archivo>')
def servir_imagen(subpath, archivo):
    # Detectar base_folder según subpath
    # Ejemplo: si subpath empieza con 6 dígitos, es lecturas
    if re.match(r'^\d{6}', subpath):
        base_folder = r"\\192.168.1.201\images\lecturas"
    else:
        base_folder = r"\\192.168.1.201\images\ordenes"

    ruta_absoluta = os.path.join(base_folder, subpath, archivo)
    print("Buscando archivo en:", ruta_absoluta)
    try:
        return send_file(ruta_absoluta)
    except FileNotFoundError:
        print("Archivo no encontrado:", ruta_absoluta)
        return "Imagen no encontrada", 404


## CODIGO PARA VER IMAGENES EN MAPA
@app.route("/buscar-multiples-coincidencias-v1", methods=["POST"])
def buscar_multiples_coincidencias_v1():
    import time
    inicio = time.time()

    data = request.get_json()
    pares = data.get("pares", [])

    if not pares or not isinstance(pares, list):
        return jsonify({"error": "Se requiere una lista de pares de códigos"}), 400

    print(f"[LOG] Búsqueda múltiple de coincidencias para {len(pares)} pares")

    index_actual = obtener_index_actualizado()
    resultados_por_carpeta = {}

    # 🔄 Preparamos todas las combinaciones de búsqueda válidas
    combinaciones = set()
    for par in pares:
        s = par.get("suministro", "").strip()
        i = par.get("inspeccion", "").strip()
        if not s or not i:
            continue
        combinaciones.add((s, i))
        combinaciones.add((s.lstrip("0"), i))
        combinaciones.add((s, i.lstrip("0")))
        combinaciones.add((s.lstrip("0"), i.lstrip("0")))

    for entry in index_actual:
        filename = entry["filename"]
        carpeta = entry["carpeta"]

        for suministro, inspeccion in combinaciones:
            if suministro in filename and inspeccion in filename:
                # Procesar carpeta (no se modifica)
                if carpeta.isdigit() and len(carpeta) == 6:
                    import re
                    match = re.search(r'_C(\d+)_', filename)
                    if match:
                        subcarpeta = match.group(1)
                        carpeta_completa = os.path.normpath(os.path.join(carpeta, subcarpeta))
                    else:
                        carpeta_completa = carpeta
                    leyenda = formato_fecha_lectura(carpeta)
                    clave_agrupacion = carpeta
                else:
                    carpeta_completa = carpeta
                    leyenda = entry.get("leyenda") or LEYENDA.get(carpeta) or carpeta
                    clave_agrupacion = carpeta_completa

                if clave_agrupacion not in resultados_por_carpeta:
                    resultados_por_carpeta[clave_agrupacion] = {
                        "carpeta": carpeta_completa,
                        "leyenda": leyenda,
                        "imagenes": []
                    }

                resultados_por_carpeta[clave_agrupacion]["imagenes"].append(filename)
                break  # no necesitas seguir más con este filename

    # Ordenar resultados
    ordenes = []
    lecturas_subgrupos = []
    for k, v in resultados_por_carpeta.items():
        if " - " in v["leyenda"]:
            lecturas_subgrupos.append(v)
        else:
            ordenes.append(v)

    ordenes.sort(key=lambda x: x["leyenda"])
    lecturas_subgrupos.sort(key=lambda x: x["leyenda"])

    resultados = []
    if lecturas_subgrupos:
        resultados.append({
            "leyenda": "LECTURAS",
            "subgrupos": lecturas_subgrupos
        })
    resultados.extend(ordenes)

    duracion = round(time.time() - inicio, 2)
    print(f"[LOG] Tiempo total de respuesta: {duracion}s")

    return jsonify({"resultados": resultados})

@app.route("/buscar-multiples-coincidencias", methods=["POST"])
def buscar_multiples_coincidencias():
    import time
    import re
    from sqlalchemy import and_, or_

    inicio = time.time()

    data = request.get_json()
    pares = data.get("pares", [])

    if not pares or not isinstance(pares, list):
        return jsonify({"error": "Se requiere una lista de pares de códigos"}), 400

    print(f"[LOG] Búsqueda múltiple BD para {len(pares)} pares")

    # 🔄 combinaciones válidas
    combinaciones = set()
    for par in pares:
        s = par.get("suministro", "").strip()
        i = par.get("inspeccion", "").strip()
        if not s or not i:
            continue

        combinaciones.add((s, i))
        combinaciones.add((s.lstrip("0"), i))
        combinaciones.add((s, i.lstrip("0")))
        combinaciones.add((s.lstrip("0"), i.lstrip("0")))

    if not combinaciones:
        return jsonify({"resultados": []})

    # 🔍 filtros SQL
    filtros = [
        and_(
            Imagen.filename.ilike(f"%{suministro}%"),
            Imagen.filename.ilike(f"%{inspeccion}%")
        )
        for suministro, inspeccion in combinaciones
    ]

    imagenes = (
        db.session.query(Imagen)
        .filter(or_(*filtros))
        .all()
    )

    resultados_por_carpeta = {}

    for img in imagenes:
        filename = img.filename
        carpeta = img.carpeta

        # Lógica original intacta
        if carpeta.isdigit() and len(carpeta) == 6:
            match = re.search(r'_C(\d+)_', filename)
            if match:
                subcarpeta = match.group(1)
                carpeta_completa = os.path.normpath(os.path.join(carpeta, subcarpeta))
            else:
                carpeta_completa = carpeta

            leyenda = formato_fecha_lectura(carpeta)
            clave_agrupacion = carpeta
        else:
            carpeta_completa = carpeta
            leyenda = img.leyenda or LEYENDA.get(carpeta) or carpeta
            clave_agrupacion = carpeta_completa

        if clave_agrupacion not in resultados_por_carpeta:
            resultados_por_carpeta[clave_agrupacion] = {
                "carpeta": carpeta_completa,
                "leyenda": leyenda,
                "imagenes": []
            }

        resultados_por_carpeta[clave_agrupacion]["imagenes"].append(filename)

    # 🔢 Ordenar resultados
    ordenes = []
    lecturas_subgrupos = []

    for v in resultados_por_carpeta.values():
        if " - " in v["leyenda"]:
            lecturas_subgrupos.append(v)
        else:
            ordenes.append(v)

    ordenes.sort(key=lambda x: x["leyenda"])
    lecturas_subgrupos.sort(key=lambda x: x["leyenda"])

    resultados = []
    if lecturas_subgrupos:
        resultados.append({
            "leyenda": "LECTURAS",
            "subgrupos": lecturas_subgrupos
        })

    resultados.extend(ordenes)

    duracion = round(time.time() - inicio, 2)
    print(f"[LOG] Tiempo total BD: {duracion}s")

    return jsonify({"resultados": resultados})


#############

UPLOAD_FOLDER = 'uploads'
PLANTILLAS = {
    'CIERRESIMPLE': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\CIERRESIMPLE.pdf',
    'CIERREALCANTARILLADO': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\CIERREALCANTARILLADO.pdf',
    'REAPERTURASIMPLE': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\REAPERTURASIMPLE.pdf',
    'REAPERTURAALCANTARILLADO': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\REAPERTURAALCANTARILLADO.pdf',
    'REAPERTURADRASTICA_CONPAVIMENTO': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\REAPERTURADRASTICA_CONPAVIMENTO.pdf',
    'REAPERTURADRASTICA_SINPAVIMENTO': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\REAPERTURADRASTICA_SINPAVIMENTO.pdf',
    'CIERREDRASTICOCONPAVIMENTO': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\CIERREDRASTICOCONPAVIMENTO.pdf',
    'CIERREDRASTICOSINPAVIMENTO': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\CIERREDRASTICOSINPAVIMENTO.pdf',
    'VERIFICACIONACCIONCOARCITIVA': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\VERIFICACIONACCIONCOARCITIVA.pdf',
    'LEVANTAMIENTO': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\LEVANTAMIENTO.pdf',
    'SELLADOALCANTARILLADO': r'C:\RADIAN\ASISTENCIAS\app\static\plantillas\SELLADOALCANTARILLADO.pdf',

}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

FIRMA_DIGITAL_PATH = r'C:\RADIAN\ALDAIR\DOCUMENTOS ECMAN-RADIAN\RADIAN\persuasivas\2_ cierre simple\FIRMA DIGITAL'

MAPEO_REAPERTURA = {
    "NEX CLI": "SUMINISTRO",
    "NEX NRO": "NRCX_NRO",
    "NEX DIR": "DIRECCIÓN",
    "NEX MED": "MEDIDOR",
    "NEX CIC": "CICLO",
    "NEX NOM": "USUARIO",
    "NEX FEC": "FECHA",
    "NEX HRA": "HORA",
    "NEX OBS": "NEX OBS",          # 👈 Para las marcas rojas
    "NEX DEU": "DEUDA",            # 👈 si lo usas en cierres
    "NEX MES": "MES DEUDA",
    "NRCX OPECX": "CODIGO",
    "CARGA": "CARGARD",
    "ORDEN": "ORDENRD",
    "NOMBRE OPERADOR": "TECNICO",
    "DNI OPERADOR": "DNI",
    "MATERIAL": "MATERIALES",
    "LECTURA": "LECTURA",          # 👈 si lo dibujas en otra plantilla
    "CANTIDAD": "CANTIDAD"
}


def insertar_imagen_en_pdf(pdf_input_path, imagen_path, pdf_output_path, pagina=0, x=0, y=0, ancho=None, alto=None):
    from PIL import Image
    
    # USAR BLOQUE WITH PARA ASEGURAR EL CIERRE DEL ARCHIVO
    with open(pdf_input_path, 'rb') as f_in:
        lector = PdfReader(f_in)
        pagina_original = lector.pages[pagina]
        ancho_pagina = float(pagina_original.mediabox.width)
        alto_pagina = float(pagina_original.mediabox.height)

        # Crear un PDF temporal con la imagen
        temp_pdf_img = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        temp_pdf_img.close()

        c = canvas.Canvas(temp_pdf_img.name, pagesize=(ancho_pagina, alto_pagina))

        img = Image.open(imagen_path)
        if ancho is None and alto is None:
            ancho, alto = img.size
        elif ancho is None:
            proporción = float(alto) / img.height
            ancho = int(img.width * proporción)
        elif alto is None:
            proporción = float(ancho) / img.width
            alto = int(img.height * proporción)

        x = float(x)
        y = float(y)

        c.drawImage(imagen_path, x, y, width=float(70), height=float(120), preserveAspectRatio=True, mask='auto')
        c.save()

        # Leer PDF con imagen (también con with open)
        with open(temp_pdf_img.name, 'rb') as f_img:
            lector_img = PdfReader(f_img)
            pagina_img = lector_img.pages[0]
            
            # Fusionar página original con la página que contiene la imagen
            pagina_original.merge_page(pagina_img)

            # Guardar PDF resultante
            escritor = PdfWriter()
            escritor.add_page(pagina_original)

            # Copiar el resto de páginas
            for i in range(len(lector.pages)):
                if i != pagina:
                    escritor.add_page(lector.pages[i])

            with open(pdf_output_path, "wb") as f_out:
                escritor.write(f_out)

    # Borrar temporal de imagen con seguridad
    try:
        if os.path.exists(temp_pdf_img.name):
            os.remove(temp_pdf_img.name)
    except OSError as e:
        print(f"⚠️ No se pudo borrar temporal de imagen: {e}")


def generar_codigo_barras(valor):
    codigo = barcode.Code128(valor, writer=ImageWriter())
    output = io.BytesIO()
    codigo.write(output, {"module_width": 0.5, "module_height": 10.0})
    output.seek(0)

    # Opcional: Verifica si la fuente está disponible (puedes eliminar esto si no usas PIL aquí)
    try:
        font_path = "arial.ttf"
        ImageFont.truetype(font_path, 14)
    except OSError:
        print("⚠️ No se encontró la fuente 'arial.ttf', intenta incluirla en el ejecutable.")

    return output


def crear_capa_texto(datos, ancho, alto, actividad):
    ancho = float(ancho)
    alto = float(alto)

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=(ancho, alto))

    can.setFont("Helvetica", 8)

    # --- AÑADE ESTE DICCIONARIO ---
    observaciones_map = {
        '0': 'NINGUNA',
        '1': 'CORTE EJECUTADO',
        '4': 'CONEXIÓN INUBICABLE',
        '3': 'FALTA DE SEGURIDAD',
        '6': 'CONEXIÓN INACTIVA',
        '18': 'CONEXIÓN ACTIVA',
        '8': 'DOBLE CONEXIÓN',
        '10': 'PAGO NO REPORTADO',
        '11': 'RECLAMO',
        '12': 'SUSPENDIDO',
        '13': 'PAGO AL DÍA',
        '17': 'OPOSICIÓN AL CORTE',
        '22': 'DIRECCIÓN INUBICABLE'
    }

    plantilla_actual = datos.get("__PLANTILLA__", "")

    if plantilla_actual in {
        "REAPERTURASIMPLE",
        "REAPERTURAALCANTARILLADO",
        "REAPERTURADRASTICA_CONPAVIMENTO",
        "REAPERTURADRASTICA_SINPAVIMENTO"
    }:
        print("Claves originales en datos:", list(datos.keys()))

        # Mapeo automático
        datos_mapeados = {}
        # En este punto los datos ya están mapeados, así que solo los limpiamos
        datos_mapeados = {k: str(v).strip() if v is not None else "" for k, v in datos.items()}


        # Sobrescribir datos originales con los mapeados
        datos.update(datos_mapeados)
        
        posiciones = {
            'USUARIO': (84, alto - 112),
            'DIRECCIÓN': (90, alto - 124),
            'MEDIDOR': (84, alto - 134),
            'FECHA': (478, alto - 112),
            'HORA': (424, alto - 123),
            'CICLO': (504, alto - 100),
            'MATERIALES': (125, alto - 322),
            'SUMINISTRO': (276, alto - 76),
            'NRCX_NRO': (422, alto - 89),
            'CARGARD': (450, alto - 89),
            'ORDENRD': (520, alto - 89),
            'CODIGO': (80, alto - 350),
            'DNI': (138, alto - 350),
            'TECNICO': (240, alto - 350),
        }
    
    elif plantilla_actual in {
        "CIERREDRASTICOCONPAVIMENTO",
        "CIERREDRASTICOSINPAVIMENTO"
    }:
        posiciones = {
            "USUARIO": (85, alto - 96),
            "DIRECCIÓN": (90, alto - 113),
            "FECHA": (398, alto - 80),
            "HORA": (485, alto - 80),
            "SUMINISTRO": (88, alto - 80),
            "MEDIDOR": (495, alto - 96),
            "NEX_NRO": (80, alto - 129),
            "CARGARD": (115, alto - 129),
            "ORDENRD": (398, alto - 129),
            "CICLO": (485, alto - 113),
            "DEUDA": (230, alto - 144),
            "MES DEUDA": (340, alto - 144),
            "MATERIALES": (128, alto - 176),
            "CODIGO": (73, alto - 351),
            "TECNICO": (220, alto - 351),
            "DNI": (135, alto - 351),
        }
    
    elif plantilla_actual == "VERIFICACIONACCIONCOARCITIVA":
        posiciones = {
            "USUARIO": (80, alto - 103),
            "DIRECCIÓN": (86, alto - 116),
            "SUMINISTRO": (88, alto - 91),
            "MEDIDOR": (515, alto - 103),
            "CARGARD": (80, alto - 131),
            "ORDENRD": (405, alto - 131),
            "CICLO": (500, alto - 116),
        }

    elif plantilla_actual == "LEVANTAMIENTO":
        posiciones = {
            "USUARIO": (79, alto - 103),
            "DIRECCIÓN": (88, alto - 116),
            "FECHA": (398, alto - 91),
            "HORA": (480, alto - 91),
            "SUMINISTRO": (88, alto - 91),
            "MEDIDOR": (495, alto - 103),
            "CARGARD": (79, alto - 129),
            "ORDENRD": (398, alto - 129),
            "CICLO": (480, alto - 116),
            "DEUDA": (225, alto - 142),
            "MES DEUDA": (315, alto - 142),
            "MATERIALES": (128, alto - 176),
            "CODIGO": (73, alto - 355),
            "TECNICO": (150, alto - 355),
            "DNI": (364, alto - 355),
        }

    elif plantilla_actual == "SELLADOALCANTARILLADO":
        posiciones = {
            "USUARIO": (79, alto - 103),
            "DIRECCIÓN": (88, alto - 117),
            "FECHA": (398, alto - 92),
            "HORA": (480, alto - 92),
            "SUMINISTRO": (88, alto - 92),
            "MEDIDOR": (495, alto - 103),
            "CARGARD": (80, alto - 131),
            "ORDENRD": (398, alto - 131),
            "CICLO": (480, alto - 116),
            "DEUDA": (225, alto - 143),
            "MES DEUDA": (315, alto - 143),
            "MATERIALES": (128, alto - 176),
            "CODIGO": (70, alto - 359),
            "TECNICO": (130, alto - 359),
            "DNI": (320, alto - 359),
        }

    else:
        posiciones = {
            'USUARIO': (82, alto - 128),
            'DIRECCIÓN': (88, alto - 138),
            'FECHA': (475, alto - 128),
            'MEDIDOR': (84, alto - 149),
            'HORA': (425, alto - 138),
            'SUMINISTRO': (290, alto - 81),
            'NEX_NRO': (380, alto - 104),
            'CARGARD': (410, alto - 104),
            'ORDENRD': (515, alto - 104),
            'CICLO': (420, alto - 116),
            'CODIGO': (75, alto - 360),
            'DNI': (125, alto - 360),
            'TECNICO': (215, alto - 360),
        }
    

    for campo, (x, y) in posiciones.items():
        texto = datos.get(campo, '')
        if campo == 'SUMINISTRO':
            can.setFont("Helvetica-Bold", 10)
        else:
            can.setFont("Helvetica", 8)
        can.drawString(x, y, texto)

    # Código de barras
    valor_codigo = datos.get('SUMINISTRO', '')
    if valor_codigo:
        imagen_codigo_barras_io = generar_codigo_barras(valor_codigo)
        imagen_codigo_barras = ImageReader(imagen_codigo_barras_io)
        barcode_x = 225
        barcode_y = alto - 45
        barcode_width = 144
        barcode_height = 24
        can.drawImage(imagen_codigo_barras, barcode_x, barcode_y, width=barcode_width, height=barcode_height)

    # Definir aquí coordenadas_x_roja para que 'alto' exista
        # Seleccionar coordenadas según la actividad
    if actividad in ['CIERREALCANTARILLADO', 'CIERRESIMPLE']:
        coordenadas_x_roja = {
            '1': (110, alto - 176),
            '2': (215, alto - 176),
            '3': (318, alto - 176),
            '4': (425, alto - 176),
            '6': (520, alto - 176),
            '7': (120, alto - 193),
            '8': (225, alto - 193),
            '9': (320, alto - 193),
            '10': (435, alto - 193),
            '11': (510, alto - 193),
            '12': (112, alto - 211),
            '13': (208, alto - 211),
            '17': (322, alto - 211),
            '22': (424, alto - 211),
            '29': (523, alto - 211),
            '0': (520, alto - 211),
        }
    elif actividad in [
        'REAPERTURASIMPLE',
        'REAPERTURAALCANTARILLADO',
        'REAPERTURADRASTICA_CONPAVIMENTO',
        'REAPERTURADRASTICA_SINPAVIMENTO'
    ]:
        coordenadas_x_roja = {
            '19': (78, alto - 161),
            '20': (156, alto - 161),
            '23': (275, alto - 161),
            '24': (350, alto - 161),
            '21': (435, alto - 161),
            '5': (520, alto - 161),
        }
    
    elif actividad in [
        'CIERREDRASTICOCONPAVIMENTO',
        'CIERREDRASTICOSINPAVIMENTO'
    ]:
        coordenadas_x_roja = {
            '1': (160, alto - 192),
        }
    else:
        coordenadas_x_roja = {}  # Por defecto
    
    x_marca_pavimento, y_marca_pavimento = None, None
    
    # 1. Definir coordenadas basadas en la actividad
    if actividad == 'CIERREDRASTICOCONPAVIMENTO':
        # 📌 COORDENADAS PARA "CON PAVIMENTO" (EJEMPLO - AJUSTA ESTOS VALORES)
        x_marca_pavimento = 92
        y_marca_pavimento = alto - 222
        
    elif actividad == 'CIERREDRASTICOSINPAVIMENTO':
        # 📌 COORDENADAS PARA "SIN PAVIMENTO" (EJEMPLO - AJUSTA ESTOS VALORES)
        x_marca_pavimento = 178
        y_marca_pavimento = alto - 222

    # 2. Dibujar la marca si se definieron coordenadas
    if x_marca_pavimento is not None:
        print(f"→ Marcando X por actividad de Cierre Drástico: {actividad}")
        
        can.setStrokeColorRGB(1, 0, 0)  # rojo
        can.setLineWidth(2)
        size = 8
        
        # Dibujar la 'X'
        can.line(x_marca_pavimento, y_marca_pavimento, x_marca_pavimento + size, y_marca_pavimento + size)
        can.line(x_marca_pavimento, y_marca_pavimento + size, x_marca_pavimento + size, y_marca_pavimento)
    
    

    # Dibujar X en coordenadas si valor_nex_obs está definido
    valor_nex_obs = str(datos.get('NEX OBS', '')).strip()

    print(f"DEBUG: Actividad actual: {actividad}")
    print(f"DEBUG: Valor de NEX OBS recibido: '{valor_nex_obs}'")
    print(f"DEBUG: Tipo de dato de NEX OBS: {type(valor_nex_obs)}")

    if plantilla_actual == "SELLADOALCANTARILLADO":
        
        if valor_nex_obs == '1':
            # ⚠️ ¡ACCIÓN REQUERIDA! ⚠️
            # Define las coordenadas (X, Y) donde debe ir la 'X' para la obs '1'
            x_mark, y_mark = (100, alto - 190)  # <-- CAMBIA ESTAS COORDENADAS
            
            print(f"→ Dibujando X para NEX OBS '1' en {x_mark}, {y_mark}")
            can.setStrokeColorRGB(1, 0, 0) # Rojo
            can.setLineWidth(2)
            size = 8
            can.line(x_mark, y_mark, x_mark + size, y_mark + size)
            can.line(x_mark, y_mark + size, x_mark + size, y_mark)

        elif valor_nex_obs in observaciones_map:
            # ⚠️ ¡ACCIÓN REQUERIDA! ⚠️
            # Define las coordenadas (X, Y) donde debe ir el TEXTO de la observación
            x_texto, y_texto = (200, alto - 295)  # <-- CAMBIA ESTAS COORDENADAS
            
            # Construye el texto: ej. "30 CARGO X LEVANTAMIENTO"
            texto_obs = f"{valor_nex_obs} {observaciones_map[valor_nex_obs]}"
            
            print(f"→ Escribiendo texto de OBS '{texto_obs}' en {x_texto}, {y_texto}")
            can.setFont("Helvetica", 8) # Define el tamaño de letra
            can.setStrokeColorRGB(0, 0, 0) # Color negro para el texto
            can.drawString(x_texto, y_texto, texto_obs)

    if valor_nex_obs in coordenadas_x_roja:
        x_mark, y_mark = coordenadas_x_roja[valor_nex_obs]
        print(f"→ Actividad: {actividad}")
        print(f"→ valor_nex_obs: '{valor_nex_obs}' en coordenadas_x_roja")

        can.setStrokeColorRGB(1, 0, 0)  # rojo
        can.setLineWidth(2)
        size = 8
        can.line(x_mark, y_mark, x_mark + size, y_mark + size)
        can.line(x_mark, y_mark + size, x_mark + size, y_mark)

    if plantilla_actual == "LEVANTAMIENTO":
        print(f"DEBUG (crear_capa): SÍ es LEVANTAMIENTO.")
        
        # --- CAMBIO: Buscamos 'DESCRIPCION NUEVO' y normalizamos a mayúsculas ---
        valor_descripcion = str(datos.get('DESCRIPCION NUEVO', 'NO_RECIBIDO')).strip().upper()
        print(f"DEBUG (crear_capa): Valor DESCRIPCION NUEVO a comparar: '{valor_descripcion}'")
        
        x_mark_niv, y_mark_niv = None, None
        
        # --- CAMBIO: Comparamos con el TEXTO en mayúsculas ---
        if valor_descripcion == 'LEVANTAMIENTO ACUEDUCTO CON PAVIMENTO':
            print("DEBUG (crear_capa): ¡COINCIDENCIA CON PAVIMENTO!")
            x_mark_niv = 146
            y_mark_niv = alto - 225
            
        elif valor_descripcion == 'LEVANTAMIENTO ACUEDUCTO SIN PAVIMENTO':
            print("DEBUG (crear_capa): ¡COINCIDENCIA SIN PAVIMENTO!")
            x_mark_niv = 198
            y_mark_niv = alto - 225
        
        else:
            # Esto es lo que seguirá pasando si no arreglas los datos:
            print(f"DEBUG (crear_capa): Sin coincidencia para '{valor_descripcion}'. No se dibuja 'X' de NIV.")

        # El resto del código para dibujar (que ya estaba bien)
        if x_mark_niv is not None:
            print(f"DEBUG (crear_capa): Dibujando X de NIV en X={x_mark_niv}, Y={y_mark_niv}")
            can.setStrokeColorRGB(1, 0, 0)  # rojo
            can.setLineWidth(2)
            size = 8
            can.line(x_mark_niv, y_mark_niv, x_mark_niv + size, y_mark_niv + size)
            can.line(x_mark_niv, y_mark_niv + size, x_mark_niv + size, y_mark_niv)
        else:
            print("DEBUG (crear_capa): x_mark_niv es None. No se dibuja.")
    
    # Dibujar marcas basadas en "MATERIAL" solo si la actividad lo permite
    if actividad in [
        'CIERRESIMPLE',
        'CIERREALCANTARILLADO'
    ]:
        materiales = str(datos.get("MATERIAL", "")).strip().upper()

        if "CONO" in materiales:
            x_cono, y_cono = 198, alto - 256  # Coordenadas para 'CONO'
            can.setStrokeColorRGB(1, 0, 0)  # rojo
            can.setLineWidth(1)
            size = 10
            can.line(x_cono, y_cono, x_cono + size, y_cono + size)
            can.line(x_cono, y_cono + size, x_cono + size, y_cono)

        if "TAPON" in materiales:
            x_tapon, y_tapon = 244, alto - 256  # Coordenadas para 'TAPON'
            can.setStrokeColorRGB(1, 0, 0)  # rojo
            can.setLineWidth(1)
            size = 10
            can.line(x_tapon, y_tapon, x_tapon + size, y_tapon + size)
            can.line(x_tapon, y_tapon + size, x_tapon + size, y_tapon)
        
        if "FICHA CIEGA" in materiales:
            x_tapon, y_tapon = 152, alto - 256  # Coordenadas para 'TAPON'
            can.setStrokeColorRGB(1, 0, 0)  # rojo
            can.setLineWidth(1)
            size = 10
            can.line(x_tapon, y_tapon, x_tapon + size, y_tapon + size)
            can.line(x_tapon, y_tapon + size, x_tapon + size, y_tapon)
        


    can.save()
    packet.seek(0)
    return packet


def generar_ficha_pdf(plantilla_path, datos, output_path, actividad):
    temp_sin_imagen = None # Inicializar variable
    try:
        if not os.path.exists(plantilla_path):
            raise FileNotFoundError(f"Plantilla no encontrada: {plantilla_path}")

        lector = PdfReader(plantilla_path)
        pagina = lector.pages[0]

        ancho = pagina.mediabox.width
        alto = pagina.mediabox.height

        print("🟢 Generando ficha con datos:", datos)
        datos['__PLANTILLA__'] = os.path.basename(plantilla_path).replace(".pdf", "")
        
        capa_texto = crear_capa_texto(datos, ancho, alto, actividad)
        
        # Leer capa texto con 'with' es difícil porque es BytesIO, pero PdfReader lo maneja bien en memoria
        lector_capa = PdfReader(capa_texto)
        pagina_capa = lector_capa.pages[0]

        escritor = PdfWriter()
        pagina.merge_page(pagina_capa)
        escritor.add_page(pagina)

        # Guardar PDF temporal sin firma
        temp_sin_imagen = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_sin_imagen.close()
        
        # IMPORTANTE: No marques para borrar AÚN si lo vas a usar inmediatamente
        # marcar_para_borrar(temp_sin_imagen.name) 

        with open(temp_sin_imagen.name, "wb") as f_salida:
            escritor.write(f_salida)

        # Buscar firma
        valor_imagen = datos.get('CODIGO', '')
        imagen_path = None
        if valor_imagen:
            for ext in ['.png', '.jpg', '.jpeg']:
                posible_path = os.path.join(FIRMA_DIGITAL_PATH, f"{valor_imagen}{ext}")
                if os.path.exists(posible_path):
                    imagen_path = posible_path
                    break

        if imagen_path:
            # Coordenadas personalizadas
            if actividad in ['CIERREDRASTICOCONPAVIMENTO', 'CIERREDRASTICOSINPAVIMENTO']:
                x_firma, y_firma = 475, alto - 360
            else:
                x_firma, y_firma = 475, alto - 390

            insertar_imagen_en_pdf(temp_sin_imagen.name, imagen_path, output_path,
                                   pagina=0, x=x_firma, y=y_firma, ancho=100)
        else:
            shutil.copy(temp_sin_imagen.name, output_path)
            
        return output_path

    except Exception as e:
        print(f"❌ Error en generar_ficha_pdf: {e}")
        # Si falló, intentamos borrar el archivo destino corrupto
        try:
            if os.path.exists(output_path):
                os.remove(output_path)
        except: pass
        return None
        
    finally:
        # LIMPIEZA SEGURA EN EL FINALLY
        if temp_sin_imagen and os.path.exists(temp_sin_imagen.name):
            try:
                os.remove(temp_sin_imagen.name)
            except OSError:
                print(f"⚠️ No se pudo eliminar temporal: {temp_sin_imagen.name}")


@app.route("/procesar-levantamiento", methods=["POST"])
def procesar_levantamiento():
    try:
        # 1. Validar si se subió el archivo
        if "reporte" not in request.files:
            return jsonify({"error": "No se subió el archivo"}), 400

        archivo = request.files["reporte"]

        # 2. Cargar Excel en memoria
        in_memory_file = io.BytesIO(archivo.read())
        wb = load_workbook(in_memory_file, data_only=True)
        ws = wb.active  # Primera hoja

        # 3. Buscar cabeceras
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # 🔎 DEBUG: Imprimir cabeceras detectadas en consola
        print("📊 Encabezados detectados en el Excel:", headers, flush=True)

        try:
            idx_nex = headers.index("NEX CLI")
            idx_cod = headers.index("CODIGO INSP")
        except ValueError:
            return jsonify({
                "error": "El archivo no contiene las columnas requeridas (NEX CLI, CODIGO INSP)",
                "headers_detectados": headers  # también lo devolvemos en la respuesta
            }), 400

        # 4. Recorrer filas y guardar datos
        registros = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            nex_cli = str(row[idx_nex]).strip() if row[idx_nex] else None
            codigo_insp = str(row[idx_cod]).strip() if row[idx_cod] else None

            if nex_cli and codigo_insp:
                registros.append({
                    "NEX_CLI": nex_cli,
                    "CODIGO_INSP": codigo_insp
                })

        return jsonify(registros)

    except Exception as e:
        print("❌ Error en procesar_levantamiento:", str(e), flush=True)
        return jsonify({"error": str(e)}), 500





# ======================= SUBIR EXCEL =======================
@app.route('/subir-excel', methods=['POST'])
def subir_excel():
    archivo = request.files.get('reporte')
    actividad = request.form.get("actividad")

    if not actividad or actividad not in PLANTILLAS:
        return "Actividad no válida", 400

    filename = secure_filename(archivo.filename)
    ruta = os.path.join(UPLOAD_FOLDER, filename)
    archivo.save(ruta)

    # Leer Excel
    df = pd.read_excel(ruta)
    if "CODIGO" in df.columns:
        df["CODIGO"] = df["CODIGO"].apply(
            lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x)
        )

    # 🔍 Nuevo filtro por DESCRIPCION NUEVO si aplica
    filtros_descripcion = {
        'REAPERTURADRASTICA_CONPAVIMENTO': 'REAPERTURA DRASTICA CON PAVIMIENTO',
        'REAPERTURADRASTICA_SINPAVIMENTO': 'REAPERTURA DRASTICA SIN PAVIMENTO',
        'CIERREDRASTICOCONPAVIMENTO': 'CIERRE DRASTICO CON PAVIMIENTO',
        'CIERREDRASTICOSINPAVIMENTO': 'CIERRE DRASTICO SIN PAVIMIENTO'
    }

    if actividad in filtros_descripcion:
        valor_filtro = filtros_descripcion[actividad]
        if 'DESCRIPCION NUEVO' in df.columns:
            df = df[df['DESCRIPCION NUEVO'].astype(str).str.strip().str.upper() == valor_filtro.upper()]
        else:
            return jsonify({"error": "El Excel no contiene la columna 'DESCRIPCION NUEVO'"}), 400

    columnas = list(df.columns)
    app.config['DATAFRAME_TEMP'] = df

    return jsonify({"columnas": columnas})


def filtrar_por_actividad(df, actividad):
    """
    Filtra el DataFrame según la actividad seleccionada y la columna 'DESCRIPCION NUEVO'
    """
    filtros_descripcion = {
        'REAPERTURADRASTICA_CONPAVIMENTO': 'REAPERTURA DRASTICA CON PAVIMIENTO',
        'REAPERTURADRASTICA_SINPAVIMENTO': 'REAPERTURA DRASTICA SIN PAVIMENTO',
        'CIERREDRASTICOCONPAVIMENTO': 'CIERRE DRASTICO CON PAVIMIENTO',
        'CIERREDRASTICOSINPAVIMENTO': 'CIERRE DRASTICO SIN PAVIMIENTO'
    }

    if actividad in filtros_descripcion and 'DESCRIPCION NUEVO' in df.columns:
        valor_filtro = filtros_descripcion[actividad]
        # Normalizamos para evitar problemas de espacios y mayúsculas
        df_filtrado = df[df['DESCRIPCION NUEVO'].astype(str).str.strip().str.upper() == valor_filtro.upper()]
        return df_filtrado
    return df



def pdf_a_imagenes(ruta_pdf):
    doc = fitz.open(ruta_pdf)
    imagenes = []
    zoom = 2  # zoom 3x para mejorar resolución (puedes ajustar entre 2 y 4)
    mat = fitz.Matrix(zoom, zoom)

    for pagina_num in range(len(doc)):
        pagina = doc.load_page(pagina_num)
        pix = pagina.get_pixmap(matrix=mat, alpha=False)  # RGB sin canal alfa, mejor para JPG
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        imagenes.append(img)
    
    doc.close()
    return imagenes

# Guardaremos el PDF generado temporalmente para reutilizarlo en /descargar-zip
pdf_generado_path = None


@app.route('/generar-fichas', methods=['POST'])
def generar_fichas():
    global pdf_generado_path

    actividad = request.form.get('actividad')
    if actividad not in PLANTILLAS:
        return "Actividad no válida", 400

    plantilla_path = PLANTILLAS[actividad]
    archivo_excel = request.files.get('archivo_excel')
    if not archivo_excel:
        return "No se subió archivo Excel", 400

    try:
        mapeo_str = request.form.get('mapeo')
        mapeo = json.loads(mapeo_str)
    except Exception as e:
        return f"Error leyendo mapeo: {str(e)}", 400

    # ✅ Leer Excel original
    df_original = pd.read_excel(archivo_excel)

    # ✅ Limpiar columna CODIGO si existe
    if "CODIGO" in df_original.columns:
        df_original["CODIGO"] = df_original["CODIGO"].apply(
            lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x)
        )

    # ✅ Guardamos el DataFrame original para referencia
    app.config['DATAFRAME_TEMP_ORIGINAL'] = df_original.copy()

    # ✅ Aplicar filtro según la actividad
    df_filtrado = filtrar_por_actividad(df_original, actividad)

    # ✅ Validar si hay datos después del filtro
    if df_filtrado.empty:
        return jsonify({"error": f"No se encontraron registros para la actividad seleccionada ({actividad})"}), 400

    # ✅ Guardar el DataFrame filtrado para usarlo en descargar ZIP
    app.config['DATAFRAME_TEMP'] = df_filtrado.reset_index(drop=True)

    campos_plantilla = list(mapeo.keys())
    pdf_temporales = []
    filas_fallidas = []

    print(f"🔄 Iniciando generación de {len(df_filtrado)} fichas...", flush=True)

    # ✅ Recorrer registros filtrados
    for idx, fila in df_filtrado.iterrows():
        datos = {}
        texto_renombrar = '' # Variable para personalizar el nombre del archivo si fuera necesario

        # --- PREPARACIÓN DE DATOS ---
        for campo in campos_plantilla:
            columna = mapeo.get(campo, '')
            valor = fila.get(columna, '')

            # Formateo de fechas
            if isinstance(valor, datetime):
                valor = valor.strftime('%d/%m/%Y')

            if 'FEC' in campo.upper():
                datos[campo] = formatear_fecha(valor)
            
            # Formateo NEX CLI / NRCX CLI (padding de ceros)
            elif columna in ['NEX CLI', 'NRCX CLI']:
                if pd.notna(valor):
                    texto = str(int(valor))
                    if len(texto) < 7:
                        texto = texto.zfill(7)
                    elif 7 < len(texto) < 11:
                        texto = texto.zfill(11)
                else:
                    texto = ''
                datos[campo] = texto
            
            # Formateo general
            else:
                if pd.isna(valor):
                    datos[campo] = ''
                elif columna == "DNI OPERADOR":
                    if isinstance(valor, float) and valor.is_integer():
                        datos[campo] = str(int(valor))
                    else:
                        datos[campo] = str(valor).strip()
                elif campo == "CODIGO":
                    if isinstance(valor, float) and valor.is_integer():
                        datos[campo] = str(int(valor))
                    else:
                        datos[campo] = str(valor).strip()
                else:
                    datos[campo] = str(valor)

        # --- LIMPIEZA NEX OBS ---
        valor_nex_obs_crudo = fila.get('NEX OBS', '')
        valor_nex_obs_limpio = ''
        if pd.notna(valor_nex_obs_crudo):
            try:
                valor_entero = int(float(valor_nex_obs_crudo))
                valor_nex_obs_limpio = str(valor_entero)
            except (ValueError, TypeError):
                valor_nex_obs_limpio = str(valor_nex_obs_crudo).strip()
        datos['NEX OBS'] = valor_nex_obs_limpio

        # --- LIMPIEZA NEX NIV ---
        columna_nex_niv = 'NEX NIV'
        valor_nex_niv_crudo = fila.get(columna_nex_niv, 'COLUMNA_NO_ENCONTRADA')
        valor_nex_niv_limpio = ''
        if pd.notna(valor_nex_niv_crudo) and valor_nex_niv_crudo != 'COLUMNA_NO_ENCONTRADA':
            try:
                valor_entero = int(float(valor_nex_niv_crudo))
                valor_nex_niv_limpio = str(valor_entero)
            except (ValueError, TypeError):
                valor_nex_niv_limpio = str(valor_nex_niv_crudo).strip()
        datos['NEX NIV'] = valor_nex_niv_limpio

        # Otros campos
        datos['MATERIAL'] = str(fila.get('MATERIAL', '')).strip() if 'MATERIAL' in df_filtrado.columns else ''
        datos['DESCRIPCION NUEVO'] = str(fila.get('DESCRIPCION NUEVO', '')).strip()

        # Re-formatear fechas internas si es necesario
        for campo in datos:
            if 'FEC' in campo.upper() and datos[campo]:
                datos[campo] = formatear_fecha(datos[campo])

        # --- GENERACIÓN DE PDF (CORREGIDO) ---
        
        # 1. Crear temporal inicial (cerrado inmediatamente)
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_pdf.close() # Importante cerrar para liberar el lock de Windows
        
        # NOTA: No llamamos a marcar_para_borrar(temp_pdf.name) aquí porque lo vamos a renombrar/mover
        
        try:
            # Generar el PDF en la ruta temporal
            generar_ficha_pdf(plantilla_path, datos, temp_pdf.name, actividad)
            
            # Validación básica
            if not os.path.exists(temp_pdf.name) or os.path.getsize(temp_pdf.name) == 0:
                raise ValueError("PDF generado está vacío o no existe")

            # 2. Construir nombre final y ruta
            nuevo_nombre_pdf = os.path.join(
                os.path.dirname(temp_pdf.name),
                f"ficha_{idx+1}.pdf" # Simplificado para evitar caracteres raros, puedes agregar texto_renombrar si quieres
            )

            # 3. Limpiar destino si ya existe (para evitar error en move)
            if os.path.exists(nuevo_nombre_pdf):
                try:
                    os.remove(nuevo_nombre_pdf)
                except OSError:
                    pass # Si falla, shutil intentará manejarlo o lanzará error controlable

            # 4. Mover (Renombrar) seguro
            shutil.move(temp_pdf.name, nuevo_nombre_pdf)
            
            # 5. Registrar el nuevo archivo para limpieza y fusión
            marcar_para_borrar(nuevo_nombre_pdf)
            pdf_temporales.append(nuevo_nombre_pdf)

        except Exception as e:
            print(f"❌ Error generando PDF en fila {idx + 1}: {e}")
            print(f"📄 Datos: {datos}")
            filas_fallidas.append(idx + 1)
            
            # Limpieza de emergencia del temporal inicial si falló
            try:
                if os.path.exists(temp_pdf.name):
                    os.remove(temp_pdf.name)
            except OSError:
                pass
            continue

    # --- FUSIÓN DE ARCHIVOS ---
    salida_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    salida_pdf.close()
    marcar_para_borrar(salida_pdf.name)

    merger = PdfMerger()
    archivos_validos = 0

    for pdf in pdf_temporales:
        try:
            if os.path.exists(pdf) and os.path.getsize(pdf) > 0:
                merger.append(pdf)
                archivos_validos += 1
            else:
                print(f"⚠️ Archivo omitido (vacío o no existe): {pdf}")
        except Exception as e:
            print(f"❌ Error al añadir {pdf} al merge: {e}")

    print(f"✅ PDFs válidos añadidos al merge: {archivos_validos}")

    if archivos_validos == 0:
        return "No se pudo generar ninguna ficha válida.", 500

    try:
        merger.write(salida_pdf.name)
        merger.close()
    except Exception as e:
        print(f"❌ Error al escribir el PDF final: {e}")
        return "Error al generar el PDF final compilado", 500

    # --- LIMPIEZA DE INDIVIDUALES ---
    for p in pdf_temporales:
        try:
            if os.path.exists(p):
                os.remove(p)
        except OSError:
            # Ignorar errores de borrado en limpieza, no detener el flujo
            pass

    # --- GESTIÓN DEL GLOBAL PDF ---
    if pdf_generado_path and os.path.exists(pdf_generado_path):
        try:
            os.remove(pdf_generado_path)
        except OSError:
            pass
            
    pdf_generado_path = salida_pdf.name
    marcar_para_borrar(pdf_generado_path)

    if filas_fallidas:
        print(f"⚠️ No se generaron fichas para las filas: {filas_fallidas}")
    else:
        print("✅ Todas las fichas se generaron correctamente.")

    return send_file(salida_pdf.name, mimetype='application/pdf', as_attachment=False)




@app.route('/descargar-zip', methods=['POST'])
def descargar_zip():
    global pdf_generado_path

    if not pdf_generado_path or not os.path.exists(pdf_generado_path):
        print("Error: No hay PDF generado o no existe en disco")
        return "No hay fichas generadas para descargar. Primero genera las fichas.", 400

    df = app.config.get('DATAFRAME_TEMP')
    if df is None:
        print("Error: No se encontró DataFrame temporal con datos del Excel")
        return "No se encontró el DataFrame con los datos del Excel. Primero sube el Excel.", 400

    imagenes = pdf_a_imagenes(pdf_generado_path)

    if len(imagenes) != len(df):
        print(f"Error: Número de imágenes {len(imagenes)} no coincide con filas Excel {len(df)}")
        return f"El número de imágenes ({len(imagenes)}) no coincide con el número de filas en Excel ({len(df)}).", 400

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for i, img in enumerate(imagenes):
            img_io = io.BytesIO()

            # dato columna 2 (índice 1)
            nombre_base_raw = str(df.iloc[i, 1]).strip()
            if nombre_base_raw.endswith('.0'):
                nombre_base_raw = nombre_base_raw[:-2]

            if len(nombre_base_raw) == 7 or len(nombre_base_raw) == 11:
                nombre_base = nombre_base_raw
            elif len(nombre_base_raw) < 7:
                nombre_base = nombre_base_raw.zfill(7)
            elif 7 < len(nombre_base_raw) < 11:
                nombre_base = nombre_base_raw.zfill(11)
            else:
                nombre_base = nombre_base_raw

            nombre_base = nombre_base.replace(" ", "_").replace("/", "_")

            # dato columna 1 (índice 0)
            dato_col1 = str(df.iloc[i, 0]).strip()
            if dato_col1.endswith('.0'):
                dato_col1 = dato_col1[:-2]

            dato_col1 = dato_col1.replace(" ", "_").replace("/", "_")

            # Concatenar con guion bajo
            nombre_archivo = f"{nombre_base}_{dato_col1}.jpg"

            img.save(img_io, format='JPEG', quality=90, optimize=True)
            img_io.seek(0)
            zip_file.writestr(nombre_archivo, img_io.read())



    zip_buffer.seek(0)

    return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name='fichas_jpg.zip')


def formatear_fecha(fecha_input):
    from datetime import datetime

    if not fecha_input or pd.isna(fecha_input):
        return ""

    # Si ya es un datetime
    if isinstance(fecha_input, datetime):
        return fecha_input.strftime("%d/%m/%Y")

    valor_str = str(fecha_input).strip()

    # Si contiene hora, quedarse solo con la parte de fecha
    if " " in valor_str:
        valor_str = valor_str.split(" ")[0]

    # Diccionario para traducir meses abreviados en español a inglés
    meses_es = {
        "ene": "jan", "feb": "feb", "mar": "mar", "abr": "apr", "may": "may", "jun": "jun",
        "jul": "jul", "ago": "aug", "sep": "sep", "oct": "oct", "nov": "nov", "dic": "dec"
    }

    # Reemplazar mes si está en español
    valor_lower = valor_str.lower()
    for mes_es, mes_en in meses_es.items():
        if f"-{mes_es}-" in valor_lower:
            valor_str = valor_lower.replace(f"-{mes_es}-", f"-{mes_en}-")
            break

    # Intentar varios formatos comunes
    formatos = ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d")

    for fmt in formatos:
        try:
            fecha = datetime.strptime(valor_str, fmt)
            return fecha.strftime("%d/%m/%Y")
        except ValueError:
            continue

    return ""



@app.route('/procesar-coordenadas', methods=['POST'])
def procesar_coordenadas():
    archivo_excel = request.files.get('reporte')
    if not archivo_excel:
        return "No se subió archivo Excel", 400

    try:
        # Leer archivo sin encabezados
        df = pd.read_excel(archivo_excel, header=None)

        # Eliminar primeras 5 filas y primera columna (A) y tercera columna (C -> índice 2)
        df = df.iloc[5:, 1:].reset_index(drop=True)
        df = df.drop(df.columns[[1, 3, 11, 12]], axis=1)  # Eliminar columna C después del recorte

        # Fila 0 se convierte en cabecera
        df.columns = df.iloc[0]
        df = df.drop(df.index[0]).reset_index(drop=True)

        # Renombrar columnas para el PDF
        renombres = {
            'FECHA EMPADRONAMIENTO': 'FECHA',
            'CODIGO CLIENTE': 'SUMINISTRO',
            'COD CAT NUEVO CLIENTE': 'NUEVO SUM'
        }
        df = df.rename(columns={k: v for k, v in renombres.items() if k in df.columns})

        # Validación de columnas necesarias
        columnas_requeridas = ['CODIGO EMPADRONADOR', 'SECTOR', 'MANZANA', 'LOTE', 'FECHA']
        for col in columnas_requeridas:
            if col not in df.columns:
                return f"Falta columna requerida: {col}", 400


        # Ordenar SECTOR, MANZANA, LOTE descendente
        df[['SECTOR', 'MANZANA', 'LOTE']] = df[['SECTOR', 'MANZANA', 'LOTE']].apply(pd.to_numeric, errors='coerce')
        df = df.sort_values(by=['SECTOR', 'MANZANA', 'LOTE'], ascending=True)

        empadronadores = df['CODIGO EMPADRONADOR'].dropna().unique()
        pdf_paths = []

        for nombre in empadronadores:
            df_filtrado = df[df['CODIGO EMPADRONADOR'] == nombre].copy()
            df_filtrado = df_filtrado.drop(columns=['CODIGO EMPADRONADOR'])

            pdf = FPDF(orientation='P', unit='mm', format='A4')
            pdf.add_page()

            # Título de empadronador
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, f"EMPADRONADOR: {nombre}", ln=True, align='C')

            orden_deseado = ['FECHA', 'SUMINISTRO', 'NUEVO SUM', 'ESTE', 'NORTE', 'CICLO','SECTOR', 'MANZANA', 'LOTE']
            cols = [col for col in orden_deseado if col in df_filtrado.columns]

            n_cols = len(cols)
            page_width = 210 - 20  # A4 horizontal con márgenes
            cell_width = page_width / n_cols
            cell_height = 6

            # Cabecera
            pdf.set_fill_color(41, 128, 185)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", 'B', 8)

            for col in cols:
                pdf.cell(cell_width, cell_height, str(col).upper(), border=1, align='C', fill=True)
            pdf.ln(cell_height)

            # Cuerpo de la tabla
            pdf.set_font("Arial", '', 8)
            pdf.set_text_color(0, 0, 0)

            for _, row in df_filtrado.iterrows():
                for col in cols:
                    val = row[col]
                    text = str(val).strip() if pd.notna(val) else ""
                    pdf.cell(cell_width, cell_height, text[:40], border=1, align='L')
                pdf.ln(cell_height)

            # Guardar PDF
            nombre_sanitizado = nombre.replace(" ", "_").replace("/", "_")
            output_path = os.path.join(tempfile.gettempdir(), f"{nombre_sanitizado}.pdf")
            pdf.output(output_path)
            pdf_paths.append(output_path)



        # Guardar rutas para descarga posterior
        app.config['COORDENADAS_PDFS'] = pdf_paths

        nombres = [os.path.splitext(os.path.basename(p))[0] for p in pdf_paths]
        return jsonify({'pdfs': nombres})

    except Exception as e:
        return f"Error procesando archivo: {str(e)}", 500


@app.route('/ver-pdf-generado/<nombre>')
def ver_pdf_generado(nombre):
    path = os.path.join(tempfile.gettempdir(), f"{nombre}.pdf")
    if os.path.exists(path):
        return send_file(path, mimetype='application/pdf')
    return "Archivo no encontrado", 404


@app.route('/descargar-coordenadas-zip')
def descargar_coordenadas_zip():
    pdf_paths = app.config.get('COORDENADAS_PDFS', [])
    if not pdf_paths:
        return "No hay PDFs generados", 400

    zip_path = os.path.join(tempfile.gettempdir(), "coordenadas_empadronadores.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for pdf in pdf_paths:  # ← aquí corregido
            if os.path.exists(pdf):
                zipf.write(pdf, os.path.basename(pdf))

    return send_file(zip_path, as_attachment=True)




# Modificación: La función ahora retorna una lista de diccionarios
def renombrar_con_zxing(folder_path):
    log_messages = []
    
    log_messages.append({"status": "INFO", "message": f"--- Escaneando con el motor ZXing-CPP (cargador OpenCV): {folder_path} ---"})
    
    renombrados, no_encontrados, errores = 0, 0, 0

    try:
        for filename in os.listdir(folder_path):
            if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.tiff')):
                original_path = os.path.join(folder_path, filename)
                
                try:
                    img = cv2.imread(original_path)
                    if img is None:
                        log_messages.append({"status": "WARNING", "message": f"⚠️ ADVERTENCIA: OpenCV no pudo leer el archivo '{filename}'."})
                        errores += 1
                        continue

                    results = zxingcpp.read_barcodes(img)
                    
                    if results:
                        barcode_data = results[0].text
                        safe_barcode_data = "".join(c for c in barcode_data if c.isalnum() or c in ('-', '_')).rstrip()

                        if not safe_barcode_data:
                            log_messages.append({"status": "WARNING", "message": f"⚠️ ADVERTENCIA: Código de barras en '{filename}' vacío."})
                            errores += 1
                            continue
                        
                        file_extension = os.path.splitext(filename)[1]
                        new_filename = f"{safe_barcode_data}{file_extension}"
                        new_path = os.path.join(folder_path, new_filename)

                        if os.path.exists(new_path):
                            log_messages.append({"status": "WARNING", "message": f"⚠️ OMITIENDO: Ya existe un archivo '{new_filename}'."})
                            errores += 1
                        else:
                            os.rename(original_path, new_path)
                            log_messages.append({"status": "SUCCESS", "message": f"✅ ÉXITO: '{filename}' renombrado a '{new_filename}'"})
                            renombrados += 1
                    else:
                        log_messages.append({"status": "INFO", "message": f"❌ INFO: No se encontró código en '{filename}'."})
                        no_encontrados += 1
                
                except Exception as e:
                    log_messages.append({"status": "ERROR", "message": f"⛔ ERROR: No se pudo procesar el archivo '{filename}'. Causa: {e}"})
                    errores += 1
    
    except FileNotFoundError:
        log_messages.append({"status": "ERROR", "message": "⛔ ERROR: La ruta de la carpeta no existe."})

    log_messages.append({"status": "INFO", "message": "--- Resumen del Proceso ---"})
    log_messages.append({"status": "INFO", "message": f"Archivos renombrados exitosamente: {renombrados}"})
    log_messages.append({"status": "INFO", "message": f"Imágenes sin código de barras detectable: {no_encontrados}"})
    log_messages.append({"status": "INFO", "message": f"Archivos con advertencias o errores: {errores}"})
    
    return log_messages


@app.route('/renombrar', methods=['POST'])
def handle_renombrar():
    data = request.get_json()
    folder_path = data.get('path')
    
    if folder_path and os.path.isdir(folder_path):
        resultados = renombrar_con_zxing(folder_path)
        return jsonify(resultados)
    else:
        return jsonify([{"status": "ERROR", "message": "⛔ ERROR: La ruta ingresada no es una carpeta válida."}])
    



@app.route('/api/guardar_carga_dia', methods=['POST'])
def guardar_carga_dia():
    try:
        data = request.get_json()
        
        if not data or 'fecha_ejecutar' not in data or 'cargas' not in data:
            return jsonify({'error': 'Faltan datos requeridos (fecha_ejecutar o cargas).'}), 400

        fecha_ejecutar_str = data.get('fecha_ejecutar')
        cargas_data = data.get('cargas')
        
        # Convertir la fecha de string a objeto Date para la DB
        # El formato esperado del input HTML es 'YYYY-MM-DD'
        fecha_ejecutar = datetime.strptime(fecha_ejecutar_str, '%Y-%m-%d').date()

        registros_guardados = 0
        
        for carga in cargas_data:
            # Crea una nueva instancia de CargaDia
            nueva_carga = CargaDia(
                suministro=carga.get('suministro'),
                direccion=carga.get('direccion'),
                actividad=carga.get('actividad'),
                # El campo fecha_ejecutar viene del formulario, no del Excel
                fecha_ejecutar=fecha_ejecutar,
                # El campo operario viene del Excel y ahora es un VARCHAR
                operario=carga.get('operario') 
            )
            
            db.session.add(nueva_carga)
            registros_guardados += 1
        
        # Confirmar todos los cambios
        db.session.commit()
        
        return jsonify({
            'message': 'Datos guardados con éxito.', 
            'count': registros_guardados
        }), 200

    except Exception as e:
        db.session.rollback() # Revierte cualquier cambio en caso de error
        print(f"Error al guardar datos de carga_dia: {e}")
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500
    


@app.route('/api/guardar_trabajo_diario', methods=['POST'])
def guardar_trabajo_diario():
    try:
        data = request.get_json()
        
        fecha_ejecucion_str = data.get('fecha_ejecucion')
        trabajos_data = data.get('trabajos')

        if not fecha_ejecucion_str or not trabajos_data or not isinstance(trabajos_data, list):
            return jsonify({'error': 'Faltan datos requeridos para el trabajo diario.'}), 400

        try:
            fecha_ejecucion_date = datetime.strptime(fecha_ejecucion_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de fecha de ejecución no válido. Se espera YYYY-MM-DD.'}), 400
        
        registros_guardados = 0
        
        for trabajo in trabajos_data:
            material = str(trabajo.get('material', '')).strip()
            cantidad_str = str(trabajo.get('cantidad', '')).strip() # Usamos string vacío como default
            actividad_excel = str(trabajo.get('actividad', '')).strip() # DESCRIPCION NUEVO
            
            # 1. Validaciones básicas de datos:
            if not material or not cantidad_str or not actividad_excel:
                print(f"[SKIP] Fila omitida: faltan datos clave (Material/Cantidad/Actividad). Material: '{material}', Actividad: '{actividad_excel}'.")
                continue

            try:
                # Usamos Decimal para mayor precisión y manejo robusto de strings
                cantidad = decimal.Decimal(cantidad_str)
                if cantidad <= 0:
                     print(f"[SKIP] Fila omitida: Cantidad es cero o negativa: {cantidad_str}.")
                     continue
            except Exception:
                print(f"[SKIP] Fila omitida: Cantidad '{cantidad_str}' no es un número válido.")
                continue
            
            # 2. BÚSQUEDA DE COINCIDENCIA EN CARGA_DIA (MÁS TOLERANTE)
            # Usamos ILIKE y LOWER() para ignorar mayúsculas/minúsculas y espacios
            
            # Nota: Algunos ORM (como Flask-SQLAlchemy) requieren el uso de .ilike() o .lower() para MySQL
            # Asegúrate de que el campo 'actividad' en la BD no sea NULL.
            
            carga_dia_match = CargaDia.query.filter(
                CargaDia.fecha_ejecutar == fecha_ejecucion_date,
                # Comparación flexible: Ignora mayúsculas/minúsculas y espacios extra en la búsqueda
                db.func.lower(CargaDia.actividad) == actividad_excel.lower() 
            ).first()

            if carga_dia_match:
                # 3. BÚSQUEDA DE MATERIAL_ASIGNADO
                # Si CargaDia coincide, buscamos cualquier asignación de material para esa carga
                material_asignado_match = MaterialAsignado.query.filter_by(
                    id_carga=carga_dia_match.id_carga
                ).first()
                
                if material_asignado_match:
                    # 4. INSERCIÓN EN CARGA_EJECUTADA
                    nueva_ejecucion = CargaEjecutada(
                        id_asignado=material_asignado_match.id_asignado,
                        fecha_ejecucion=fecha_ejecucion_date,
                        material_u=material,
                        cantidad_u=cantidad
                    )
                    
                    db.session.add(nueva_ejecucion)
                    registros_guardados += 1
                else:
                    print(f"[FAIL] CargaDia: '{actividad_excel}' ({carga_dia_match.id_carga}) encontrada, pero NO tiene material asignado. Omite guardar ejecución.")
            else:
                print(f"[FAIL] No se encontró CargaDia planeada para la actividad '{actividad_excel}' en la fecha {fecha_ejecucion_str}. Omite guardar ejecución.")
                
        
        db.session.commit()
        
        if registros_guardados == 0:
             # Este mensaje se enviará si la lógica de filtrado ha descartado todos los registros
             return jsonify({'message': 'Proceso completado. Ningún registro se guardó (falla en coincidencia de planificación/asignación).', 'saved_count': 0}), 200
             
        return jsonify({
            'message': 'Datos de ejecución validados y guardados con éxito.', 
            'saved_count': registros_guardados
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"ERROR: Fallo grave al guardar Trabajo Diario: {e}")
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500
    




# TEMPLATE_DBF_PATH = r'\templates_excel\PLANTILLA_VL229082023.dbf' 
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

TEMPLATE_DBF_PATH = os.path.join(
    BASE_DIR,
    'app',
    'templates_excel',
    'PLANTILLA_VL229082023.dbf'
)

# 2. Carpeta temporal para guardar los DBF antes de zippear
OUTPUT_DIR = 'temp_dbf_output'
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

# --- RUTA 1 (NUEVA VERSIÓN): PREVISUALIZAR CON GRUPOS ---
@app.route('/previsualizar-dbf', methods=['POST'])
def previsualizar_dbf():
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se encontró el archivo Excel.'}), 400
        
        file = request.files['archivo']
        fecha_seleccionada_str = request.form['fecha'] 
        
        # Leemos el Excel, forzando que TODO se lea como STRING (texto)
        df = pd.read_excel(file, skiprows=5, dtype=str)
        
        df = df.drop(columns=['T', 'U', 'V'], errors='ignore')
        df = df.dropna(how='all')

        # Convertimos la fecha seleccionada por el usuario a un string dd/mm/YYYY
        fecha_obj_seleccionada = datetime.strptime(fecha_seleccionada_str, '%Y-%m-%d')
        df['FECNOTIMED'] = fecha_obj_seleccionada.strftime('%d/%m/%Y')
        
        # --- LÓGICA DE AGRUPACIÓN (COMO EN LA FUNCIÓN DE DESCARGA) ---
        columna_filtro = 'CICLOREAL' # Basado en tus logs
        
        if columna_filtro not in df.columns:
             return jsonify({'error': f'No se encontró la columna de filtro "{columna_filtro}" en el Excel.'}), 400
        
        grupos_unicos = df[columna_filtro].dropna().unique()
        
        # ¡IMPORTANTE! Creamos un diccionario para guardar los datos por grupo
        data_por_grupo = {}

        for grupo in grupos_unicos:
            df_partido = df[df[columna_filtro] == grupo].copy()
            
            # --- Preparamos los datos para JSON ---
            df_partido['LECTURA'] = pd.to_numeric(df_partido['LECTURA'], errors='coerce')
            
            # Convertimos el DataFrame DE ESTE GRUPO a JSON
            data_para_json = df_partido.fillna('').to_dict(orient='records')
            
            # Lo añadimos al diccionario principal
            data_por_grupo[str(grupo)] = data_para_json # Usamos str(grupo) para la clave JSON
        
        # Devolvemos el diccionario de grupos
        # Ejemplo: {"21": [...datos...], "22": [...datos...]}
        return jsonify(data_por_grupo)

    except Exception as e:
        print(f"Error grave en /previsualizar-dbf: {e}")
        return jsonify({'error': str(e)}), 500




# --- RUTA 2 (MODIFICADA): TU CÓDIGO FUNCIONAL PARA DESCARGAR EL ZIP ---
@app.route('/descargar-dbf', methods=['POST'])
def descargar_dbf():
    # Este es el código que me enviaste y que funciona perfectamente
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se encontró el archivo Excel.'}), 400
        
        file = request.files['archivo']
        fecha_seleccionada_str = request.form['fecha'] 
        
        if not os.path.exists(TEMPLATE_DBF_PATH):
             return jsonify({'error': 'No se encontró el archivo de plantilla DBF en el servidor.'}), 500

        # --- 1. PROCESAMIENTO DE EXCEL ---
        df = pd.read_excel(file, skiprows=5, dtype=str)
        
        print("\nDEBUG 1: Nombres de columna leídos del EXCEL:", list(df.columns))

        df = df.drop(columns=['T', 'U', 'V'], errors='ignore')
        df = df.dropna(how='all')

        fecha_obj_seleccionada = datetime.strptime(fecha_seleccionada_str, '%Y-%m-%d').date()
        df['FECNOTIMED'] = fecha_obj_seleccionada
        
        # --- 2. DIVISIÓN DEL EXCEL ---
        columna_filtro = 'CICLOREAL' 
        
        if columna_filtro not in df.columns:
             return jsonify({'error': f'No se encontró la columna de filtro "{columna_filtro}" en el Excel.'}), 400
        
        grupos_unicos = df[columna_filtro].dropna().unique()
        dataframes_partidos = {}
        for grupo in grupos_unicos:
            dataframes_partidos[grupo] = df[df[columna_filtro] == grupo].copy()

        archivos_dbf_generados = []

        # --- 3. LÓGICA DBF CON CORRECCIÓN DE TIPO ---
        with dbf.Table(TEMPLATE_DBF_PATH) as plantilla:
            lista_campos_dbf = [f.lower() for f in plantilla.field_names]

        print(f"DEBUG 2 (Simplificado): Campos detectados en DBF: {lista_campos_dbf}")

        for grupo, df_partido in dataframes_partidos.items():
            
            dbf_filename = f'resultado_{grupo}.dbf'
            dbf_filepath = os.path.join(OUTPUT_DIR, dbf_filename)
            
            shutil.copy(TEMPLATE_DBF_PATH, dbf_filepath)
            
            dbf_table = dbf.Table(dbf_filepath)
            dbf_table.open(dbf.READ_WRITE)
            dbf_table.zap()
            
            col_map = {}
            for col_excel in df_partido.columns:
                if col_excel.lower() in lista_campos_dbf:
                    col_map[col_excel.lower()] = col_excel

            print(f"\n--- DEBUG 3 (Grupo {grupo}): Mapeo de columnas (DBF -> Excel) ---")
            print(col_map)

            for index, fila_excel in df_partido.iterrows():
                
                nuevo_registro = {}
                
                for campo_dbf_lower, col_excel_original in col_map.items():
                    
                    valor = fila_excel[col_excel_original]
                    
                    if pd.isna(valor) or valor in (None, 'None', ''):
                        nuevo_registro[campo_dbf_lower] = None
                        continue

                    try:
                        if campo_dbf_lower == 'fecnotimed':
                            nuevo_registro[campo_dbf_lower] = valor
                        elif campo_dbf_lower == 'fchinsreal':
                            nuevo_registro[campo_dbf_lower] = datetime.strptime(str(valor), '%d/%m/%Y').date()
                        elif campo_dbf_lower == 'lectura':
                            nuevo_registro[campo_dbf_lower] = float(valor)
                        else:
                            # --- CORRECCIÓN AQUÍ ---
                            val_str = str(valor)
                            # Si el texto supera los 254 caracteres, lo cortamos
                            if len(val_str) > 254:
                                val_str = val_str[:254] 
                            
                            nuevo_registro[campo_dbf_lower] = val_str

                    except Exception as e:
                        print(f"  -> Advertencia: No se pudo convertir '{valor}' para el campo '{campo_dbf_lower}'. Error: {e}")
                        nuevo_registro[campo_dbf_lower] = None

                if nuevo_registro:
                    try:
                        dbf_table.append(nuevo_registro)
                    except Exception as e:
                        print(f"¡ERROR AL ANEXAR! {e}. Datos: {nuevo_registro}")

            dbf_table.close()
            archivos_dbf_generados.append(dbf_filepath)

        # 4. Comprimir y enviar
        if not archivos_dbf_generados:
            return jsonify({'error': 'No se generaron archivos, revise los datos del Excel.'}), 500

        zip_io = io.BytesIO()
        with zipfile.ZipFile(zip_io, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f_path in archivos_dbf_generados:
                zf.write(f_path, os.path.basename(f_path))
                os.remove(f_path)
                
        zip_io.seek(0)

        return send_file(zip_io,
                         mimetype='application/zip',
                         as_attachment=True,
                         download_name='conversiones_dbf.zip')

    except Exception as e:
        print(f"Error grave en /descargar-dbf: {e}")
        return jsonify({'error': str(e)}), 500
    

###### GENERAR CARGAS ######
# --- RUTA 1 (ACTUALIZADA): PREVISUALIZAR LA CARGA (JSON Único) ---
@app.route('/previsualizar-carga', methods=['POST'])
def previsualizar_carga():
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se encontró el archivo Excel.'}), 400
        
        file = request.files['archivo']
        actividad = request.form.get('actividad')

        if actividad == 'PERSUASIVAS':
            
            # --- SOLUCIÓN REQ 2 y 4: LEER TODO COMO TEXTO ---
            df = pd.read_excel(file, dtype=str)
            df = df.fillna('') # Reemplazamos NaN por strings vacíos

            # 1. Insertar columnas al inicio
            df.insert(0, 'TECNICO', '')
            df.insert(1, 'CARGA', '')
            df.insert(2, 'ORDEN', '')

            # 2. Duplicar columnas 'obs'
            new_obs_cols = []
            if 'nex_obs' in df.columns:
                idx = df.columns.get_loc('nex_obs')
                new_col_name = 'nex_obs_COPIA' # Nuevo nombre
                df.insert(idx + 1, new_col_name, '')
                new_obs_cols.append(new_col_name)

            if 'nrcx_obs' in df.columns:
                idx = df.columns.get_loc('nrcx_obs')
                new_col_name = 'nrcx_obs_COPIA' # Nuevo nombre
                df.insert(idx + 1, new_col_name, '')
                new_obs_cols.append(new_col_name)

            # 3. Añadir SECTOR y MANZANA
            df['SECTOR'] = ''
            df['MANZANA'] = ''
            
            # 4. Búsqueda en Base de Datos
            key_col = None
            if 'nex_cli' in df.columns: key_col = 'nex_cli'
            elif 'nrcx_cli' in df.columns: key_col = 'nrcx_cli'
            elif 'CODIGO' in df.columns: key_col = 'CODIGO'
            
            if not key_col:
                return jsonify({'error': 'No se encontró columna de suministro (nex_cli, nrcx_cli, o CODIGO)'}), 400

            keys_to_lookup = df[key_col].dropna().unique()
            
            lookup_results = db.session.query(
                DataCatastroV2.suministro_p,
                DataCatastroV2.sector,
                DataCatastroV2.manzana
            ).filter(
                DataCatastroV2.suministro_p.in_(keys_to_lookup)
            ).all()
            
            lookup_map = {res.suministro_p: (res.sector, res.manzana) for res in lookup_results}

            # 5. Poblar los datos en el DataFrame
            def get_data_from_map(key, part): # 0 = sector, 1 = manzana
                return lookup_map.get(str(key), ('', ''))[part]
            
            df['SECTOR'] = df[key_col].apply(lambda x: get_data_from_map(x, 0))
            df['MANZANA'] = df[key_col].apply(lambda x: get_data_from_map(x, 1))

            # 6. Ordenar por SECTOR y MANZANA
            df = df.sort_values(by=['SECTOR', 'MANZANA'], ascending=True)

            # 7. Definir las columnas que quieres MANTENER VISIBLES (ORDENADAS)
            visible_cols_list = [
                'TECNICO', 'CARGA', 'ORDEN', 
                'nex_cli', 'nex_dir', 'nex_med', 'nex_nom', 'nex_obs', 'nex_cic',
                'ITEM', 'CODIGO', 'NOMBRE', 'DIRECCION', 'MEDIDOR', 'CICLO',
                'nrcx_cli', 'nrcx_niv', 'nrcx_nom', 'nrcx_dir', 'nrcx_med', 'nrcx_cic', 'nrcx_obs',
                'acciones', 'servidor'
            ]
            visible_cols_list.extend(new_obs_cols) 
            visible_cols_list.extend(['SECTOR', 'MANZANA']) # <-- SECTOR Y MANZANA AL FINAL
            
            final_visible_cols = [col for col in visible_cols_list if col in df.columns]
            
            # 8. Crear el DataFrame de previsualización
            df_preview = df[final_visible_cols]

            # --- INICIO DE LA MODIFICACIÓN ---
            # Devolvemos un objeto con la lista de columnas (en orden)
            # y las filas de datos.
            data_to_send = {
                'columns': final_visible_cols,
                'rows': df_preview.to_dict(orient='records')
            }
            return jsonify(data_to_send)
            # --- FIN DE LA MODIFICACIÓN ---
        
        else:
            return jsonify({'error': f'Lógica para {actividad} aún no implementada.'}), 400

    except Exception as e:
        print(f"Error grave en /previsualizar-carga: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/descargar-carga', methods=['POST'])
def descargar_carga():
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se encontró el archivo Excel.'}), 400
        
        file = request.files['archivo']
        actividad = request.form.get('actividad')
        
        # --- MODIFICACIÓN 1: Obtener el nombre original ---
        # Si no se envía, usamos 'carga_generada.xlsx' como respaldo
        nombre_original = request.form.get('nombre_original', 'carga_generada.xlsx')
        
        if actividad == 'PERSUASIVAS':
            
            # --- SOLUCIÓN REQ 2 y 4: LEER TODO COMO TEXTO ---
            df = pd.read_excel(file, dtype=str)
            df = df.fillna('') # Reemplazamos NaN por strings vacíos

            # ... (TODA TU LÓGICA DE PROCESAMIENTO VA AQUÍ) ...
            
            # 1. Insertar columnas
            df.insert(0, 'TECNICO', '')
            df.insert(1, 'CARGA', '')
            df.insert(2, 'ORDEN', '')

            # 2. Duplicar 'obs'
            new_obs_cols = []
            yellow_cols = [] 
            
            if 'nex_obs' in df.columns:
                idx = df.columns.get_loc('nex_obs')
                new_col_name = 'nex_obs_COPIA'
                df.insert(idx + 1, new_col_name, '')
                new_obs_cols.append(new_col_name)
                yellow_cols.append(new_col_name)

            if 'nrcx_obs' in df.columns:
                idx = df.columns.get_loc('nrcx_obs')
                new_col_name = 'nrcx_obs_COPIA'
                df.insert(idx + 1, new_col_name, '')
                new_obs_cols.append(new_col_name)
                yellow_cols.append(new_col_name)

            # 3. Añadir SECTOR y MANZANA
            df['SECTOR'] = ''
            df['MANZANA'] = ''
            yellow_cols.extend(['SECTOR', 'MANZANA'])
            
            # 4. Búsqueda en Base de Datos
            key_col = None
            if 'nex_cli' in df.columns: key_col = 'nex_cli'
            elif 'nrcx_cli' in df.columns: key_col = 'nrcx_cli'
            elif 'CODIGO' in df.columns: key_col = 'CODIGO'
            else:
                return jsonify({'error': 'No se encontró columna de suministro'}), 400

            keys_to_lookup = df[key_col].dropna().unique()
            lookup_results = db.session.query(DataCatastroV2.suministro_p, DataCatastroV2.sector, DataCatastroV2.manzana).filter(DataCatastroV2.suministro_p.in_(keys_to_lookup)).all()
            lookup_map = {res.suministro_p: (res.sector, res.manzana) for res in lookup_results}
            df['SECTOR'] = df[key_col].apply(lambda x: lookup_map.get(str(x), ('', ''))[0])
            df['MANZANA'] = df[key_col].apply(lambda x: lookup_map.get(str(x), ('', ''))[1])

            # 5. Ordenar
            df = df.sort_values(by=['SECTOR', 'MANZANA'], ascending=True)

            # 6. Definir columnas VISIBLES
            visible_cols_list = [
                'TECNICO', 'CARGA', 'ORDEN', 'SECTOR', 'MANZANA',
                'nex_cli', 'nex_dir', 'nex_med', 'nex_nom', 'nex_obs', 'nex_cic',
                'ITEM', 'CODIGO', 'NOMBRE', 'DIRECCION', 'MEDIDOR', 'CICLO',
                'nrcx_cli', 'nrcx_niv', 'nrcx_nom', 'nrcx_dir', 'nrcx_med', 'nrcx_cic', 'nrcx_obs',
                'acciones', 'servidor'
            ]
            visible_cols_list.extend(new_obs_cols) # Solución Req 3
            final_visible_cols = [col for col in visible_cols_list if col in df.columns]

            # --- 7. CREACIÓN DEL EXCEL FORMATEADO ---
            
            output_stream = BytesIO()
            wb = Workbook()
            ws = wb.active
            
            # Escribir los datos en la hoja
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # --- SOLUCIÓN REQ 5: APLICAR ESTILOS ---
            
            # Definir estilos
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Azul oscuro
            header_font = Font(color="FFFFFF", bold=True)
            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid") # Amarillo
            thin_border_side = Side(border_style="thin", color="000000")
            cell_border = Border(top=thin_border_side, left=thin_border_side, right=thin_border_side, bottom=thin_border_side)
            
            col_map = {name: get_column_letter(idx) for idx, name in enumerate(df.columns, 1)}

            # Aplicar estilos y formatos
            for col_name in df.columns:
                col_letter = col_map[col_name]
                
                # Ocultar columnas que NO están en la lista visible
                if col_name not in final_visible_cols:
                    ws.column_dimensions[col_letter].hidden = True
                else:
                    # Auto-ajustar las columnas visibles
                    max_length = max((len(str(s)) for s in df[col_name].dropna()), default=0)
                    max_length = max(len(col_name), max_length) + 3
                    ws.column_dimensions[col_letter].width = max_length

                # Pintar las columnas amarillas
                if col_name in yellow_cols:
                    for cell in ws[col_letter]:
                        cell.fill = yellow_fill

            # Aplicar estilo de cabecera y bordes a todo
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    # Aplicar bordes a todas las celdas
                    cell.border = cell_border
                    # Aplicar estilo de cabecera a la fila 1
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font

            # Congelar la fila superior y añadir filtro
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
            
            # --- FIN DE ESTILOS ---
            
            # 8. Guardar el Excel en memoria
            wb.save(output_stream)
            output_stream.seek(0)
            
            # --- MODIFICACIÓN 2: Procesar y usar el nombre original ---
            
            # (Ej: "mi_archivo.xls" -> "mi_archivo_procesado.xlsx")
            nombre_base, _ = os.path.splitext(nombre_original)
            nuevo_nombre_descarga = f"{nombre_base}_procesado.xlsx"

            return send_file(
                output_stream,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=nuevo_nombre_descarga # <-- Usamos el nuevo nombre
            )
        
        else:
            return jsonify({'error': f'La descarga para {actividad} aún no está implementada.'}), 400

    except Exception as e:
        print(f"Error grave en /descargar-carga: {e}")
        return jsonify({'error': str(e)}), 500
    

### DASHBOARD ###
@app.route('/api/dashboard-data')
def get_dashboard_data():
    # 1. FILTROS DINÁMICOS (Recibe fecha del selector HTML)
    fecha_str = request.args.get('fecha')
    zona_filtro = request.args.get('zona') # Si quieres filtrar por zona

    # Si no envían fecha, usamos HOY por defecto
    if not fecha_str:
        fecha_str = str(date.today())

    # Consulta Base de Datos
    query = RegistroTrabajo.query.filter(RegistroTrabajo.fecha_ini_ejecucion == fecha_str)
    
    # Si quisieras filtrar por zona (asumiendo que tienes columna 'zona' o 'localidad')
    if zona_filtro and zona_filtro != "Todas":
        query = query.filter(RegistroTrabajo.localidad == zona_filtro)
        
    registros = query.all()

    # Si no hay datos, retornamos ceros
    if not registros:
        return jsonify({
            'kpis': {'total': 0, 'sla': 0, 'efectividad': 0},
            'timeline': [],
            'anomalies': []
        })

    # --- PROCESAMIENTO PANDAS (Igual que antes pero robusto) ---
    data_list = []
    for r in registros:
        if r.hora_ini_ejecucion and r.hora_fin_ejecucion:
            # Calcular duración real para eficiencia
            inicio = datetime.combine(r.fecha_ini_ejecucion, r.hora_ini_ejecucion)
            fin = datetime.combine(r.fecha_fin_ejecucion, r.hora_fin_ejecucion)
            duracion_min = (fin - inicio).total_seconds() / 60
            
            data_list.append({
                'operario': r.operario,
                'actividad': r.actividad,
                'fecha_cargue': r.fecha_cargue,
                'inicio_dt': inicio,
                'fin_dt': fin,
                'duracion': duracion_min
            })

    df = pd.DataFrame(data_list)
    if df.empty:
         return jsonify({'kpis': {'total': 0}, 'timeline': [], 'anomalies': []})

    # KPI 1: Total
    total_ordenes = len(df)
    
    # KPI 2: SLA (Promedio días)
    df['fecha_cargue'] = pd.to_datetime(df['fecha_cargue'])
    avg_sla = (df['inicio_dt'] - df['fecha_cargue']).dt.days.mean()

    # KPI 3: Tiempo Efectivo (Simplificado para el ejemplo)
    # Suponemos turno de 8 horas (480 min). Qué % del tiempo estuvieron en órdenes?
    tiempo_total_trabajado = df.groupby('operario')['duracion'].sum().mean()
    efectividad = (tiempo_total_trabajado / 480) * 100 

    # TIMELINE
    timeline_data = []
    colores = {'Corte': '#D7263D', 'Inspección': '#008FFB', 'Lectura': '#00E396', 'Mantenimiento': '#FEB019'}
    
    for _, row in df.iterrows():
        actividad = row['actividad'] or 'Generico'
        timeline_data.append({
            'x': row['operario'],
            'y': [row['inicio_dt'].timestamp() * 1000, row['fin_dt'].timestamp() * 1000],
            'fillColor': colores.get(actividad, '#775DD0'),
            'meta': actividad # Para el tooltip
        })

    # ANOMALÍAS (Huecos > 20 min)
    anomalias = []
    df = df.sort_values(by=['operario', 'inicio_dt'])
    for operario, grupo in df.groupby('operario'):
        grupo['fin_anterior'] = grupo['fin_dt'].shift(1)
        grupo['tiempo_muerto'] = grupo['inicio_dt'] - grupo['fin_anterior']
        
        huecos = grupo[grupo['tiempo_muerto'] > timedelta(minutes=20)] # Umbral 20 min
        for _, row in huecos.iterrows():
            anomalias.append({
                'operario': operario,
                'hora_fin': row['fin_anterior'].strftime('%H:%M'),
                'hora_inicio': row['inicio_dt'].strftime('%H:%M'),
                'duracion': int(row['tiempo_muerto'].total_seconds() / 60)
            })

    return jsonify({
        'kpis': {
            'total': total_ordenes,
            'sla': round(avg_sla, 1) if not pd.isna(avg_sla) else 0,
            'efectividad': int(efectividad) if not pd.isna(efectividad) else 0
        },
        'timeline': timeline_data,
        'anomalies': anomalias
    })



#### CARTAS ####
# --- FUNCIÓN AUXILIAR ROBUSTA PARA IMÁGENES ---
def procesar_imagen_para_pdf(nombre_archivo_sin_extension):
    """
    Busca la imagen directamente en la carpeta 'static' de la app.
    Soporta .png, .jpg, .jpeg.
    """
    try:
        # CORRECCIÓN: Apuntamos directamente a 'static', no a 'static/img'
        base_folder = os.path.join(current_app.root_path, 'static')
        
        ruta_encontrada = None
        # Busca probando extensiones comunes
        for ext in ['.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG']:
            posible_ruta = os.path.join(base_folder, f"{nombre_archivo_sin_extension}{ext}")
            if os.path.exists(posible_ruta):
                ruta_encontrada = posible_ruta
                break
        
        if not ruta_encontrada:
            print(f"ERROR: No se encontró '{nombre_archivo_sin_extension}' en {base_folder}")
            return None

        # Procesar con Pillow (Quitar transparencia y convertir a RGB)
        img = Image.open(ruta_encontrada)
        if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
            alpha = img.convert('RGBA').split()[-1]
            bg = Image.new("RGB", img.size, (255, 255, 255))
            bg.paste(img, mask=alpha)
            img = bg
        else:
            img = img.convert('RGB')

        # Guardar en temporal
        temp = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
        img.save(temp, format='JPEG', quality=95)
        temp.close()
        return temp.name

    except Exception as e:
        print(f"Error procesando imagen: {str(e)}")
        return None

def normalizar_cols(df):
    df.columns = df.columns.astype(str).str.strip().str.upper()
    return df

# --- RUTA PRINCIPAL ---
@app.route('/generar_cartas_pdf', methods=['POST'])
def generar_cartas_pdf():
    temps_to_clean = [] 

    try:
        # 1. VALIDACIÓN
        files = request.files.getlist('excelFiles')
        fecha_emision = request.form.get('fechaEmision')

        if not files or len(files) < 2 or not fecha_emision:
            return jsonify({'error': 'Faltan archivos o fecha'}), 400

        # 2. PROCESAR EXCEL
        df_madre = None
        df_sec = None

        for f in files:
            df_temp = pd.read_excel(f, dtype=str, keep_default_na=False)
            df_temp = normalizar_cols(df_temp)
            cols = df_temp.columns.tolist()
            
            if 'SUM. ENTIDAD' in cols and 'SUMINISTRO' in cols:
                df_madre = df_temp
            elif 'CLICODFAX' in cols and 'CARGARD' in cols:
                df_sec = df_temp

        if df_madre is None or df_sec is None:
            return jsonify({'error': 'No se identificaron los archivos requeridos.'}), 400

        df_madre['SUMINISTRO'] = df_madre['SUMINISTRO'].str.strip()
        df_sec['CLICODFAX'] = df_sec['CLICODFAX'].str.strip()

        merged = pd.merge(df_madre, df_sec, left_on='SUMINISTRO', right_on='CLICODFAX', how='inner')
        
        # --- [AGREGAR ESTA LÍNEA] ---
        # 1. Convertimos las columnas a NÚMEROS en columnas temporales
        # 'errors=coerce' transforma en 0 cualquier texto que no sea número para que no falle
        merged['SORT_CARGA'] = pd.to_numeric(merged['CARGARD'], errors='coerce').fillna(0)
        merged['SORT_ORDEN'] = pd.to_numeric(merged['ORDENRD'], errors='coerce').fillna(0)

        # 2. Ordenamos usando esas columnas numéricas
        merged = merged.sort_values(by=['SORT_CARGA', 'SORT_ORDEN'], ascending=[True, True])# ----------------------------

        grupos = merged.groupby('SUM. ENTIDAD')

        if grupos.ngroups == 0:
            return jsonify({'error': 'No hay coincidencias.'}), 400

        # 3. PREPARAR IMÁGENES
        ruta_logo_tmp = procesar_imagen_para_pdf('logo')
        ruta_firma_tmp = procesar_imagen_para_pdf('firma_distribucion')
        
        if ruta_logo_tmp: temps_to_clean.append(ruta_logo_tmp)
        if ruta_firma_tmp: temps_to_clean.append(ruta_firma_tmp)

        # 4. CONFIGURAR PDF
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)
        
        fecha_dt = datetime.strptime(fecha_emision, '%Y-%m-%d')
        dias = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
        meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        fecha_texto = f"Trujillo, {dias[fecha_dt.weekday()]}, {fecha_dt.day} de {meses[fecha_dt.month-1]} de {fecha_dt.year}"

        for nombre_grupo, datos in grupos:
            pdf.add_page()
            cabecera = datos.iloc[0]
            
            # --- LOGO (SE MANTIENE ARRIBA: y=10) ---
            if ruta_logo_tmp:
                pdf.image(ruta_logo_tmp, x=10, y=10, w=20)

            # --- CÓDIGO DE BARRAS (BAJADO UN POCO: y=15) ---
            barcode_val = str(nombre_grupo)
            rv = io.BytesIO()
            Code128 = barcode.get_barcode_class('code128')
            Code128(barcode_val, writer=ImageWriter()).write(rv, options={"write_text": False}) 
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_bar:
                tmp_bar.write(rv.getvalue())
                path_bar = tmp_bar.name
                temps_to_clean.append(path_bar)

            ancho_pag = 210
            ancho_bar = 50
            x_bar = (ancho_pag - ancho_bar) / 2
            
            # Ajuste de posición Y para bajarlo un poco (antes y=10)
            pdf.image(path_bar, x=x_bar, y=25, w=ancho_bar, h=10)
            
            # --- TEXTO DEBAJO BARCODE (BAJADO: y=27) ---
            pdf.set_y(37) # Antes 21
            pdf.set_font("Arial", 'B', size=8) 
            pdf.cell(0, 4, txt=barcode_val, ln=True, align='C')

            # --- CARTA N° Y FECHA (BAJADO: y=45) ---
            pdf.set_y(55) # Antes 35, bajamos 10mm más para separar del encabezado
            # CARTA EN NEGRITA
            pdf.set_font("Arial", 'B', size=10)
            pdf.cell(0, 4, txt=f"CARTA N° DRC {cabecera.get('ID.ENT.', 'S/N')} - {fecha_dt.year}", ln=True)
            
            # FECHA NORMAL
            pdf.set_font("Arial", size=9)
            pdf.ln(1)
            pdf.cell(0, 4, txt=fecha_texto, ln=True)
            pdf.ln(5)

            # --- DATOS DEL DESTINATARIO (Minimalista Size 9) ---
            
            def escribir_campo(etiqueta, valor):
                # Etiqueta en Negrita
                pdf.set_font("Arial", 'B', size=9)
                ancho_etiqueta = pdf.get_string_width(etiqueta) + 2
                pdf.cell(ancho_etiqueta, 4, txt=etiqueta, ln=0) 
                
                # Valor en Normal
                pdf.set_font("Arial", size=9)
                pdf.cell(0, 4, txt=str(valor), ln=1)

            escribir_campo("Señores:", cabecera.get('NOMBRE ENTIDAD', ''))
            escribir_campo("Asunto:", "Entrega de Recibos")
            escribir_campo("Dirección:", cabecera.get('DIRECCION ENTIDAD', ''))
            
            pdf.ln(5)
            pdf.multi_cell(0, 4, txt="De nuestra mayor consideración, hacemos entrega de los recibos de agua potable, perteneciente a SEDALIB S.A.")

            # ... (Línea ~139) ...
            pdf.ln(6)
            pdf.set_font("Arial", 'B', size=10) # Título un poco más grande
            pdf.cell(0, 4, txt="RELACIÓN DE RECIBOS", ln=True, align='C')
            pdf.ln(2)

            # --- CABECERA DE LA TABLA ---
            # Cambio 1: Letra tamaño 9 en negrita
            pdf.set_font("Arial", 'B', size=9) 
            pdf.set_fill_color(173, 216, 230)
            
            # Cambio 2: Columnas más anchas para que el texto respire
            # Antes sumaban 117mm, ahora suman 140mm (entra perfecto en A4)
            w_cols = [15, 45, 40, 40] 
            
            x_start = (ancho_pag - sum(w_cols)) / 2
            
            pdf.set_x(x_start)
            headers = ['N°', 'SUMINISTRO', 'CARGA', 'ORDEN']
            for i, h in enumerate(headers):
                pdf.cell(w_cols[i], 6, h, border=1, fill=True, align='C') # Altura 6 en cabecera
            pdf.ln()

            # --- CONTENIDO DE LA TABLA ---
            # Cambio 3: Letra tamaño 9 normal
            pdf.set_font("Arial", size=9)
            
            for _, row in datos.iterrows():
                pdf.set_x(x_start)
                
                # Cambio 4: Altura de fila a 6mm (antes era 4)
                # Esto es necesario porque la letra es más grande
                altura_fila = 6 
                
                pdf.cell(w_cols[0], altura_fila, str(row.get('ITEM', '')), border=1, align='C')
                pdf.cell(w_cols[1], altura_fila, str(row.get('SUMINISTRO', '')), border=1, align='C')
                pdf.cell(w_cols[2], altura_fila, str(row.get('CARGARD', '')), border=1, align='C')
                pdf.cell(w_cols[3], altura_fila, str(row.get('ORDENRD', '')), border=1, align='C')
                pdf.ln()

            # --- DESPEDIDA ---
            pdf.ln(6)
            pdf.set_font("Arial", size=9)
            pdf.cell(0, 4, txt="Sin otro particular, reciba usted las muestras de mi mayor estima personal.", ln=True)
            pdf.cell(0, 4, txt="Atentamente,", ln=True)

            # Verificar salto de página para firma
            if pdf.get_y() > 250: pdf.add_page()
            
            y_firma = pdf.get_y() + 20
            
            # --- FIRMA (Compacta) ---
            if ruta_firma_tmp:
                ancho_firma = 25
                x_firma = (ancho_pag - ancho_firma) / 2
                pdf.image(ruta_firma_tmp, x=x_firma, y=y_firma-20, w=ancho_firma)

            x_linea_start = (ancho_pag - 70) / 2 
            x_linea_end = x_linea_start + 70
            pdf.line(x_linea_start, y_firma, x_linea_end, y_firma)
            
            pdf.set_y(y_firma + 2)
            pdf.set_font("Arial", 'B', size=8)
            pdf.cell(0, 3, txt="PALACIOS RISCO JOSÉ MIGUEL", ln=True, align='C')
            pdf.set_font("Arial", size=6)
            pdf.cell(0, 3, txt="SUPERVISOR DE IMPRESIÓN Y DISTRIBUCIÓN DE RECIBOS Y COMUNICACIONES", ln=True, align='C')
            pdf.cell(0, 3, txt="CONSORCIO ECMAN - RADIAN", ln=True, align='C')

        out = pdf.output(dest='S').encode('latin-1')
        buffer = io.BytesIO(out)
        
        return send_file(buffer, mimetype='application/pdf', as_attachment=False, download_name=f'cartas_{fecha_emision}.pdf')

    except Exception as e:
        print(f"ERROR: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
    finally:
        for tmp in temps_to_clean:
            if os.path.exists(tmp):
                try: os.unlink(tmp)
                except: pass


#### GENERAR ASIGNACIÓN ####
@app.route('/procesar_asignacion', methods=['POST'])
def procesar_asignacion():
    try:
        file = request.files['archivo']
        f_cal_input = request.form.get('fecha_calendario', '')
        f_exe_input = request.form.get('fecha_ejecucion', '')
        # 1. CAPTURAR EL NUEVO PARÁMETRO (Por defecto 'CONTINUOS' por seguridad)
        sub_actividad = request.form.get('sub_actividad', 'CONTINUOS') 
        mapping = json.loads(request.form.get('mapping', '{}'))

        # LEER EXCEL (dtype=str para no perder ceros en clicodfax o cargard)
        df = pd.read_excel(file, dtype=str)
        
        # Normalizar cabeceras a minúsculas inmediatamente
        df.columns = df.columns.str.lower().str.strip()

        # --- 2. LÓGICA CONDICIONAL SEGÚN SUB-ACTIVIDAD ---
        if sub_actividad == 'CONTINUOS':
            # LÓGICA DE RE-PARTICIÓN (528 REGISTROS)
            if 'fciclo' in df.columns and 'cargard' in df.columns:
                # Asegurar que cargard sea numérico para poder sumar
                df['cargard'] = pd.to_numeric(df['cargard'], errors='coerce').fillna(0).astype(int)
                
                nuevo_limite = 528
                df_reestructurado = pd.DataFrame()
                ciclos = df['fciclo'].unique()

                for ciclo in ciclos:
                    df_ciclo = df[df['fciclo'] == ciclo].copy().reset_index(drop=True)
                    
                    if not df_ciclo.empty:
                        # Tomamos la carga inicial del ciclo para mantener la correlación
                        carga_inicial = df_ciclo['cargard'].iloc[0]
                        # Aplicamos la nueva distribución
                        df_ciclo['cargard'] = df_ciclo.index.map(lambda x: carga_inicial + (x // nuevo_limite))
                    
                    df_reestructurado = pd.concat([df_reestructurado, df_ciclo], ignore_index=True)
                
                df = df_reestructurado

        # Si es 'DISPERSOS', el código ignora el bloque if anterior y el DataFrame 
        # mantiene sus cargas originales intactas.

        # --- 3. PROCESAR FECHAS Y ASIGNACIÓN ---
        def formatear_fecha(fecha_str):
            try:
                if not fecha_str: return ""
                return datetime.strptime(fecha_str, '%Y-%m-%d').strftime('%d/%m/%Y')
            except: return fecha_str

        df['fecha_cal'] = formatear_fecha(f_cal_input)
        df['fecha_Ejecucion'] = formatear_fecha(f_exe_input)

        df['cargard'] = df['cargard'].astype(str).str.replace(r'\.0$', '', regex=True)
        df['operario'] = df['cargard'].map(mapping).fillna("")

        # --- 4. ORDENAR COLUMNAS Y EXPORTAR ---
        nuevas_cols = ['fecha_cal', 'fecha_Ejecucion', 'operario']
        cols_base = [c for c in df.columns if c not in nuevas_cols]
        df = df[cols_base + nuevas_cols]

        # Generar archivos en memoria
        output_excel = io.BytesIO()
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        
        output_csv = io.BytesIO()
        # utf-8-sig ayuda a que Excel reconozca tildes en el CSV automáticamente
        df.to_csv(output_csv, index=False, sep=';', encoding='utf-8-sig') 

        # --- 5. ZIPPEAR ---
        mem_zip = io.BytesIO()
        with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("asignacion_procesada.xlsx", output_excel.getvalue())
            zf.writestr("asignacion_procesada.csv", output_csv.getvalue())
        
        mem_zip.seek(0)
        return send_file(mem_zip, mimetype='application/zip', as_attachment=True, download_name='Asignacion_Cargas.zip')

    except Exception as e:
        return {"error": str(e)}, 500


# ==========================================
# FUNCIONES AUXILIARES PARA LIMPIAR DATOS
# ==========================================
def to_date(f): 
    return datetime.strptime(f, '%Y-%m-%d').date() if f and str(f).strip() != "" else None

def to_num(n): 
    return float(n) if n and str(n).strip() != "" else 0.0

def to_time(t): 
    # Asegurarnos de que solo tome la parte de la hora (HH:MM) si viene con segundos o en otro formato
    if t and str(t).strip() != "":
        try:
            return datetime.strptime(t[:5], '%H:%M').time()
        except ValueError:
            return None
    return None


@app.route('/descargar_transformado', methods=['POST'])
def descargar_transformado():
    try:
        # 1. Obtener los datos enviados desde el frontend (FormData)
        datos_json = request.form.get('datos_limpios')
        nombre_base = request.form.get('nombre_base', 'Datos_Transformados')

        if not datos_json:
            return {"error": "No se recibieron datos"}, 400

        # Convertir el texto JSON a una lista de diccionarios en Python
        datos = json.loads(datos_json)

        # 2. Convertir directamente a DataFrame de Pandas
        df = pd.DataFrame(datos)

        # Garantizar el orden exacto de las columnas que solicitaste
        columnas_esperadas = [
            'ITEM', 'SUMINISTRO', 'NOMBRE', 'LOCALIDAD', 'URBANIZACION',
            'CALLE', 'NUMERO', 'CICLO', 'MEDIDOR', 'N° DOCUMENTO',
            'FECHA EMISION', 'FECHA ENVIO', 'CARGA_RD', 'ORDEN_RD', 'TIPO_ORDEN'
        ]
        
        # Filtramos y ordenamos el DataFrame
        # (Si por algún motivo llega una columna extra se descarta, y si falta se crea vacía)
        for col in columnas_esperadas:
            if col not in df.columns:
                df[col] = ""
                
        df = df[columnas_esperadas]

        # 3. Generar archivos en memoria (Excel y CSV)
        output_excel = io.BytesIO()
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        
        output_csv = io.BytesIO()
        # utf-8-sig asegura que los acentos y las ñ (como en N° o URBANIZACIÓN) se vean bien en Excel
        df.to_csv(output_csv, index=False, sep=';', encoding='utf-8-sig') 

        # 4. Crear el archivo ZIP en memoria
        mem_zip = io.BytesIO()
        with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            # Insertamos ambos archivos generados con el nombre de la sub-actividad y fecha
            zf.writestr(f"{nombre_base}.xlsx", output_excel.getvalue())
            zf.writestr(f"{nombre_base}.csv", output_csv.getvalue())
        
        mem_zip.seek(0)

        # 5. Enviar el archivo ZIP descargable al cliente
        return send_file(
            mem_zip, 
            mimetype='application/zip', 
            as_attachment=True, 
            download_name=f'{nombre_base}.zip'
        )

    except Exception as e:
        return {"error": str(e)}, 500

###### GESTIÓN DE EMPLEADOS #####
@app.route('/api/empleados/listar', methods=['GET'])
def listar_empleados():
    empleados = Empleado.query.order_by(Empleado.apellidos.asc()).all()
    lista_datos = []
    
    for emp in empleados:
        # Función interna para manejar el formato de horas (TIME/timedelta)
        def format_time(t):
            if not t: return '-'
            # Si es timedelta (común en MySQL), lo convertimos a string y tomamos HH:MM
            if hasattr(t, 'seconds'): 
                total_seconds = int(t.total_seconds())
                horas = total_seconds // 3600
                minutos = (total_seconds % 3600) // 60
                return f"{horas:02d}:{minutos:02d}"
            # Si ya es un objeto time de python
            return t.strftime('%H:%M')

        lista_datos.append({
            'id': emp.id_empleado,
            'dni': emp.dni,
            'apellidos_nombres': emp.nombres,
            'area': emp.area,
            'cargo': emp.cargo,
            'estado': emp.estado,
            'fecha_ingreso': emp.fecha_ingreso.strftime('%d/%m/%Y') if emp.fecha_ingreso else '-',
            'fecha_cese': emp.fecha_cese.strftime('%d/%m/%Y') if emp.fecha_cese else '-',
            'fecha_nacimiento': emp.fecha_nacimiento.strftime('%d/%m/%Y') if emp.fecha_nacimiento else '-',
            # ✅ Usamos la función format_time para evitar el crash
            'hora_ingreso': format_time(emp.hora_ingreso),
            'hora_salida': format_time(emp.hora_salida),
            'telefono': emp.telefono or '-',
            'correo': emp.correo or '-'
        })
        
    return jsonify(lista_datos)

# Configuración de carpeta de subida (colócalo arriba en tu archivo)
UPLOAD_FOLDER = 'uploads/documentos_empleados'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route('/api/empleados/registrar_todo', methods=['POST'])
def registrar_empleado_completo():
    try:
        # --- CRUCIAL: Cambiamos get_json() por request.form ---
        # request.form recibe los textos y request.files recibe los archivos
        datos = request.form 

        # --- 1. VALIDACIÓN DE OBLIGATORIOS ---
        nombres_f = datos.get('nombres', '').strip()
        apellidos_f = datos.get('apellidos', '').strip()
        dni = datos.get('dni')
        cargo = datos.get('cargo')
        area = datos.get('area')
        fecha_ingreso = datos.get('fecha_ingreso')
        fecha_cese=datos.get('fecha_cese')

        if not nombres_f or not dni or not cargo or not area or not fecha_ingreso:
            return jsonify({'error': 'Faltan campos obligatorios (Nombres, DNI, Cargo, Área o Ingreso)'}), 400

        # --- 2. FUNCIONES DE LIMPIEZA ---
        def to_date(f): return datetime.strptime(f, '%Y-%m-%d').date() if f else None
        def to_num(n): return float(n) if n and str(n).strip() != "" else 0.0
        def to_time(t): return datetime.strptime(t, '%H:%M').time() if t and str(t).strip() != "" else None

        # --- 3. CREAR INSTANCIA DE EMPLEADO ---
        nombre_completo = f"{nombres_f} {apellidos_f}".strip()
        
        nuevo_empleado = Empleado(
            nombres=nombre_completo,
            apellidos=None, 
            dni=dni,
            cargo=cargo,
            area=area,
            fecha_ingreso=to_date(fecha_ingreso),
            fecha_cese=to_date(fecha_cese),
            fecha_nacimiento=to_date(datos.get('fecha_nacimiento')),
            sexo=datos.get('sexo'),
            estado_civil=datos.get('estado_civil'),
            direccion=datos.get('direccion'),
            telefono=datos.get('telefono'),
            correo=datos.get('correo'),
            tipo_contrato=datos.get('tipo_contrato'),
            jornada_laboral=datos.get('jornada_laboral'),
            # Usando el formato TIME que definimos antes
            hora_ingreso=to_time(datos.get('hora_ingreso')),
            hora_salida=to_time(datos.get('hora_salida')),
            refrigerio_inicio=to_time(datos.get('refrigerio_inicio')),
            refrigerio_fin=to_time(datos.get('refrigerio_fin')),
            regimen_laboral=datos.get('regimen_laboral'),
            estado='ACTIVO'
        )

        db.session.add(nuevo_empleado)
        db.session.flush() 

        # ==========================================
        # 📎 NUEVA SECCIÓN: GUARDADO DE DOCUMENTO
        # ==========================================
        archivo = request.files.get('archivo') # Viene del input type="file"
        tipo_doc_id = datos.get('tipo_documento') # Viene del select

        if archivo and archivo.filename != '' and tipo_doc_id:
            nombre_original = secure_filename(archivo.filename)
            # Creamos un nombre único: emp_ID_FECHA_NOMBRE.ext
            nombre_final = f"emp_{nuevo_empleado.id_empleado}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{nombre_original}"
            
            ruta_relativa = os.path.join(UPLOAD_FOLDER, nombre_final)
            archivo.save(ruta_relativa)
            
            # Guardamos en la tabla de documentos del empleado
            nuevo_doc = DocumentoEmpleado(
                empleado_id=nuevo_empleado.id_empleado,
                tipo_documento_id=int(tipo_doc_id),
                ruta_archivo=ruta_relativa
            )
            db.session.add(nuevo_doc)
        # ==========================================

        # --- 4. TABLA REMUNERACIONES ---
        if datos.get('sueldo_basico'):
            nueva_rem = Remuneracion(
                empleado_id=nuevo_empleado.id_empleado,
                sueldo_basico=to_num(datos.get('sueldo_basico')),
                asignacion_familiar=to_num(datos.get('asignacion_familiar')),
                bonificacion=to_num(datos.get('bonificacion')),
                comisiones=to_num(datos.get('comisiones')),
                horas_extras=to_num(datos.get('horas_extras')),
                moneda=datos.get('moneda', 'PEN')
            )
            db.session.add(nueva_rem)

        # --- 5. TABLA DATOS BANCARIOS ---
        if datos.get('numero_cuenta') or datos.get('banco'):
            nuevo_banco = DatosBancarios(
                empleado_id=nuevo_empleado.id_empleado,
                banco=datos.get('banco'),
                tipo_cuenta=datos.get('tipo_cuenta'),
                numero_cuenta=datos.get('numero_cuenta'),
                cci=datos.get('cci')
            )
            db.session.add(nuevo_banco)

        # --- 6. TABLA BENEFICIOS SOCIALES ---
        if datos.get('cts') or datos.get('gratificacion'):
            nuevo_ben = BeneficioSocial(
                empleado_id=nuevo_empleado.id_empleado,
                cts=to_num(datos.get('cts')),
                gratificacion=to_num(datos.get('gratificacion')),
                vacaciones_truncas=to_num(datos.get('vacaciones_truncas')),
                liquidacion=to_num(datos.get('liquidacion'))
            )
            db.session.add(nuevo_ben)

        # --- 7. FINALIZAR TRANSACCIÓN ---
        db.session.commit()

        return jsonify({
            'mensaje': 'Registro integral con documento completado',
            'id_empleado': nuevo_empleado.id_empleado
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"ERROR: {str(e)}")
        return jsonify({'error': 'Error al procesar el registro', 'detalle': str(e)}), 500
    

# ==============================================================
# 2. RUTA: OBTENER DATOS PARA EL MODAL (GET)
# ==============================================================
@app.route('/api/empleados/<int:id_empleado>', methods=['GET'])
def obtener_empleado(id_empleado):
    try:
        empleado = Empleado.query.get_or_404(id_empleado)
        rem = Remuneracion.query.filter_by(empleado_id=id_empleado).first()
        
        # 1. Traemos TODOS los documentos del empleado (No solo el primero)
        documentos_db = DocumentoEmpleado.query.filter_by(empleado_id=id_empleado).order_by(DocumentoEmpleado.fecha_subida.desc()).all()
        
        # 2. Diccionario para mapear el ID del tipo con su nombre real
        nombres_tipos = {1: 'DNI Escaneado', 2: 'Currículum Vitae (CV)', 3: 'Contrato Firmado', 4: 'Foto Carnet'}
        
        # 3. Armamos la lista de documentos para el JavaScript
        lista_docs = []
        for d in documentos_db:
            lista_docs.append({
                # OJO: Cambia 'd.id' por el nombre real de tu llave primaria en la tabla DocumentoEmpleado si es diferente (ej. d.id_documento)
                'id_doc': d.id if hasattr(d, 'id') else getattr(d, 'id_documento', 0), 
                'tipo': nombres_tipos.get(d.tipo_documento_id, 'Documento Adjunto'),
                'ruta': f"/{d.ruta_archivo.replace('\\', '/')}",
                'fecha': d.fecha_subida.strftime('%d/%m/%Y') if hasattr(d, 'fecha_subida') and d.fecha_subida else 'Reciente'
            })
        
        datos = {
            'id_empleado': empleado.id_empleado,
            'nombres': empleado.nombres.split(' ')[0] if empleado.nombres else '', 
            'apellidos': ' '.join(empleado.nombres.split(' ')[1:]) if empleado.nombres else '', 
            'dni': empleado.dni,
            'cargo': empleado.cargo,
            'area': empleado.area,
            'fecha_ingreso': empleado.fecha_ingreso.strftime('%Y-%m-%d') if empleado.fecha_ingreso else '',
            'fecha_cese': empleado.fecha_cese.strftime('%Y-%m-%d') if empleado.fecha_cese else '',
            'fecha_nacimiento': empleado.fecha_nacimiento.strftime('%Y-%m-%d') if empleado.fecha_nacimiento else '',
            'sexo': empleado.sexo,
            'estado_civil': empleado.estado_civil,
            'direccion': empleado.direccion,
            'telefono': empleado.telefono,
            'correo': empleado.correo,
            'tipo_contrato': empleado.tipo_contrato,
            'jornada_laboral': empleado.jornada_laboral,
            'hora_ingreso': empleado.hora_ingreso.strftime('%H:%M') if empleado.hora_ingreso else '',
            'hora_salida': empleado.hora_salida.strftime('%H:%M') if empleado.hora_salida else '',
            'refrigerio_inicio': empleado.refrigerio_inicio.strftime('%H:%M') if empleado.refrigerio_inicio else '',
            'refrigerio_fin': empleado.refrigerio_fin.strftime('%H:%M') if empleado.refrigerio_fin else '',
            'regimen_laboral': empleado.regimen_laboral,
            'estado': empleado.estado,
            
            'sueldo_basico': rem.sueldo_basico if rem else '',
            'moneda': rem.moneda if rem else 'PEN',
            
            # 4. Pasamos la lista completa al JSON en lugar de un solo objeto
            'documentos': lista_docs
        }
        return jsonify(datos), 200
    except Exception as e:
        return jsonify({'error': 'Error al cargar empleado', 'detalle': str(e)}), 500


# ==============================================================
# 3. RUTA: ACTUALIZAR EMPLEADO EXISTENTE (PUT)
# ==============================================================
@app.route('/api/empleados/actualizar/<int:id_empleado>', methods=['PUT'])
def actualizar_empleado_completo(id_empleado):
    try:
        empleado = Empleado.query.get_or_404(id_empleado)
        datos = request.form

        # Actualizamos datos básicos
        nombres_f = datos.get('nombres', '').strip()
        apellidos_f = datos.get('apellidos', '').strip()
        empleado.nombres = f"{nombres_f} {apellidos_f}".strip()
        
        empleado.dni = datos.get('dni')
        empleado.cargo = datos.get('cargo')
        empleado.area = datos.get('area')
        
        # Guardamos directamente los datos sin conversiones extra
        empleado.fecha_ingreso = datos.get('fecha_ingreso') if datos.get('fecha_ingreso') else None
        empleado.fecha_cese = datos.get('fecha_cese') if datos.get('fecha_cese') else None
        empleado.fecha_nacimiento = datos.get('fecha_nacimiento') if datos.get('fecha_nacimiento') else None
        empleado.sexo = datos.get('sexo')
        empleado.estado_civil = datos.get('estado_civil')
        empleado.direccion = datos.get('direccion')
        empleado.telefono = datos.get('telefono')
        empleado.correo = datos.get('correo')
        empleado.tipo_contrato = datos.get('tipo_contrato')
        empleado.jornada_laboral = datos.get('jornada_laboral')
        
        empleado.hora_ingreso = datos.get('hora_ingreso') if datos.get('hora_ingreso') else None
        empleado.hora_salida = datos.get('hora_salida') if datos.get('hora_salida') else None
        empleado.refrigerio_inicio = datos.get('refrigerio_inicio') if datos.get('refrigerio_inicio') else None
        empleado.refrigerio_fin = datos.get('refrigerio_fin') if datos.get('refrigerio_fin') else None
        
        empleado.regimen_laboral = datos.get('regimen_laboral')
        empleado.estado = datos.get('estado')

        # --- ACTUALIZAR DOCUMENTO (Si sube uno nuevo) ---
        archivo = request.files.get('archivo')
        tipo_doc_id = datos.get('tipo_documento')

        # Esto solo agregará un documento nuevo a la lista, no borrará los anteriores
        if archivo and archivo.filename != '' and tipo_doc_id:
            nombre_original = secure_filename(archivo.filename)
            nombre_final = f"emp_{empleado.id_empleado}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{nombre_original}"
            ruta_relativa = os.path.join(UPLOAD_FOLDER, nombre_final).replace('\\', '/')
            archivo.save(ruta_relativa)
            
            nuevo_doc = DocumentoEmpleado(
                empleado_id=empleado.id_empleado,
                tipo_documento_id=int(tipo_doc_id),
                ruta_archivo=ruta_relativa
            )
            db.session.add(nuevo_doc) 

        # --- ACTUALIZAR REMUNERACIÓN ---
        if datos.get('sueldo_basico'):
            rem = Remuneracion.query.filter_by(empleado_id=empleado.id_empleado).first()
            if rem:
                rem.sueldo_basico = datos.get('sueldo_basico')
                rem.moneda = datos.get('moneda', 'PEN')
            else:
                nueva_rem = Remuneracion(empleado_id=empleado.id_empleado, sueldo_basico=datos.get('sueldo_basico'), moneda=datos.get('moneda', 'PEN'))
                db.session.add(nueva_rem)

        db.session.commit()
        return jsonify({'mensaje': 'Empleado actualizado correctamente'}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error al actualizar', 'detalle': str(e)}), 500


@app.route('/api/documentos/<int:id_doc>', methods=['DELETE'])
def eliminar_documento(id_doc):
    try:
        # 1. Buscar el registro en la base de datos (usando DocumentoEmpleado)
        documento = DocumentoEmpleado.query.get(id_doc)

        if not documento:
            return jsonify({'success': False, 'message': 'No se encontró el documento'}), 404

        # ¡NOMBRES CORREGIDOS SEGÚN TU MODELO!
        ruta_archivo_fisico = documento.ruta_archivo
        id_empleado = documento.empleado_id 
        tipo_doc_id = documento.tipo_documento_id # Guarda el ID del tipo de documento

        # Obtener nombre del empleado para la auditoría
        empleado = Empleado.query.get(id_empleado)
        nombre_empleado = empleado.nombres if empleado else f'ID {id_empleado}'

        # 2. Eliminar el registro de la base de datos
        db.session.delete(documento)
        db.session.commit()

        # 3. Eliminar el archivo físico del servidor
        if ruta_archivo_fisico:
            ruta_fisica = os.path.join(current_app.root_path, ruta_archivo_fisico.lstrip('/'))
            if os.path.exists(ruta_fisica):
                os.remove(ruta_fisica)
                print(f"Archivo físico eliminado: {ruta_fisica}")

        # 4. ✅ Registrar en la auditoría
        if 'user_id' in session:
            registrar_evento(
                user_id=session['user_id'],
                usuario=session['user_name'],
                evento='eliminar_documento',
                modulo=f"Documentos | Tipo ID: {tipo_doc_id} | Empleado: {nombre_empleado}"
            )

        return jsonify({
            'success': True, 
            'message': 'Documento eliminado correctamente'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False, 
            'message': f'Error al eliminar: {str(e)}'
        }), 500

@app.route('/empleado/cesar/<int:id>', methods=['POST'])
def cesar_empleado(id):
    try:
        empleado = Empleado.query.get_or_404(id)
        empleado.estado = 'CESADO'
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})
    
    
@app.route('/api/reportes/asistencia', methods=['GET'])
def api_reporte_asistencia():
    inicio = request.args.get('inicio')
    fin = request.args.get('fin')
    empleado_id = request.args.get('empleado')

    if not inicio or not fin:
        return jsonify({'error': 'Faltan parámetros de fecha'}), 400

    tablas_asistencia = [
        EmpleadoLectura, EmpleadoDistribucion, EmpleadoInspecciones,
        EmpleadoCatastro, EmpleadoPersuasivas, EmpleadoMedidores,
        EmpleadoNorte, EmpleadoRecaudacion, EmpleadoAdministrativo
    ]

    # Diccionario con TODAS tus claves inicializadas en 0
    conteo = {
        'A': 0, 'DT': 0, 'FT': 0, 'LG': 0, 'DM': 0, 'V': 0, 'LSG': 0,
        'F': 0, 'R': 0, 'SU': 0, 'CE': 0, 'FG': 0, 'LD': 0, 'DC': 0,
        'AP': 0, 'LP': 0, 'TC': 0
    }

    for modelo in tablas_asistencia:
        query = db.session.query(modelo.estado, func.count(modelo.estado)) \
            .filter(modelo.fec_asist >= inicio, modelo.fec_asist <= fin)

        if empleado_id:
            try:
                empleado_id = int(empleado_id)
                query = query.filter(modelo.id_empleado == empleado_id)
            except ValueError:
                pass

        resultados = query.group_by(modelo.estado).all()

        for estado_db, cantidad in resultados:
            estado = (estado_db or '').strip().upper()
            # Si el estado de la BD existe en nuestro diccionario, le sumamos la cantidad
            if estado in conteo:
                conteo[estado] += cantidad

    # Enviamos el JSON con los 17 datos
    return jsonify(conteo)



@app.route('/api/reportes/gastos_incidencias', methods=['GET'])
def api_gastos_incidencias():
    inicio = request.args.get('inicio')
    fin = request.args.get('fin')
    empleado_id = request.args.get('empleado')

    if not inicio or not fin:
        return jsonify({'error': 'Faltan fechas'}), 400

    tablas = [
        ('Lectura', EmpleadoLectura), ('Distribucion', EmpleadoDistribucion),
        ('Inspecciones', EmpleadoInspecciones), ('Catastro', EmpleadoCatastro),
        ('Persuasivas', EmpleadoPersuasivas), ('Medidores', EmpleadoMedidores),
        ('Norte', EmpleadoNorte), ('Recaudacion', EmpleadoRecaudacion),
        ('Administrativo', EmpleadoAdministrativo)
    ]

    gastos_data = {'areas': [], 'viaticos': [], 'pasajes': []}
    incidencias_dia = {'Lunes': 0, 'Martes': 0, 'Miércoles': 0, 'Jueves': 0, 'Viernes': 0, 'Sábado': 0, 'Domingo': 0}
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

    # --- NUEVOS DICCIONARIOS PARA RANKING ---
    ranking_inc_map = {} # {id: {'nombre': '', 'faltas': 0, 'dm': 0}}
    ranking_gastos_map = {} # {id: {'nombre': '', 'pasajes': 0, 'viaticos': 0}}

    for nombre_area, modelo in tablas:
        # Hacemos un JOIN con la tabla Empleado para traer el nombre
        query = db.session.query(
            modelo.id_empleado,
            Empleado.nombres.label('nombre_empleado'),
            Empleado.area.label('area_real'), 
            modelo.fec_asist,
            modelo.estado,
            modelo.viaticos,
            modelo.pasajes
        ).join(Empleado, Empleado.id_empleado == modelo.id_empleado).filter( # <--- AQUÍ EL CAMBIO
            modelo.fec_asist >= inicio, 
            modelo.fec_asist <= fin
        )

        if empleado_id:
            try:
                emp_id_int = int(empleado_id)
                query = query.filter(modelo.id_empleado == emp_id_int)
            except ValueError: pass

        registros = query.all()
        suma_viaticos_area = 0
        suma_pasajes_area = 0

        for reg in registros:
            e_id = reg.id_empleado
            nombre = reg.nombre_empleado or f"Emp. {e_id}"
            area_empleado = reg.area_real
            
            # --- 1. PROCESAMIENTO DE GASTOS ---
            v_val = 0
            p_val = 0
            try: v_val = float(reg.viaticos or 0)
            except: pass
            try: p_val = float(reg.pasajes or 0)
            except: pass

            suma_viaticos_area += v_val
            suma_pasajes_area += p_val

            # Acumular para el ranking global de gastos
            if e_id not in ranking_gastos_map:
                # Agregamos 'area': nombre_area aquí
                ranking_gastos_map[e_id] = {'nombre': nombre, 'area': area_empleado, 'pasajes': 0, 'viaticos': 0}
            
            ranking_gastos_map[e_id]['pasajes'] += p_val
            ranking_gastos_map[e_id]['viaticos'] += v_val

            # --- 2. PROCESAMIENTO DE INCIDENCIAS ---
            estado = (reg.estado or '').strip().upper()
            
            if estado in ['F', 'DM']:
                if e_id not in ranking_inc_map:
                    # Agregamos 'area': nombre_area aquí
                    ranking_inc_map[e_id] = {'nombre': nombre, 'area': area_empleado, 'faltas': 0, 'dm': 0}
                
                if estado == 'F': ranking_inc_map[e_id]['faltas'] += 1
                if estado == 'DM': ranking_inc_map[e_id]['dm'] += 1

                # Gráfico por día
                if reg.fec_asist:
                    try:
                        dia_idx = reg.fec_asist.weekday()
                        incidencias_dia[dias_semana[dia_idx]] += 1
                    except: pass

        gastos_data['areas'].append(nombre_area)
        gastos_data['viaticos'].append(suma_viaticos_area)
        gastos_data['pasajes'].append(suma_pasajes_area)
    
    incidencias_final = {dia: incidencias_dia[dia] for dia in dias_semana}

    # --- 3. ORDENAR Y OBTENER TOP 10 ---
    # Ordenar incidencias por la suma de (faltas + dm)
    lista_incidencias = sorted(
        ranking_inc_map.values(), 
        key=lambda x: (x['faltas'] + x['dm']), 
        reverse=True
    )[:10]

    # Ordenar gastos por la suma de (pasajes + viaticos)
    lista_gastos = sorted(
        ranking_gastos_map.values(), 
        key=lambda x: (x['pasajes'] + x['viaticos']), 
        reverse=True
    )[:10]

    return jsonify({
        'gastos': gastos_data,
        'incidencias': incidencias_final,
        'ranking_incidencias': lista_incidencias,
        'ranking_gastos': lista_gastos
    })





# =====================================================================
# 1. EXCEL ULTRA PROFESIONAL (Con Membrete, Filtros y Tablas Nativas)
# =====================================================================
@app.route('/api/exportar/excel', methods=['GET'])
def exportar_excel():
    inicio = request.args.get('inicio')
    fin = request.args.get('fin')
    tipo = request.args.get('tipo', 'todos')
    empleado_id = request.args.get('empleado')

    if not inicio or not fin:
        return "Faltan fechas", 400

    if tipo in ['asistencias', 'todos']:
        return generar_excel_record_asistencias(inicio, fin, empleado_id)
    
    return "Exportación para este tipo en desarrollo", 200


def generar_excel_record_asistencias(inicio, fin, empleado_id=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Récord de Asistencias"

    # ==========================================
    # 1. DISEÑO Y ESTILOS PROFESIONALES
    # ==========================================
    ws.merge_cells('A1:W1') 
    ws['A1'] = f"RÉCORD GENERAL DE ASISTENCIAS ({inicio} al {fin})"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_grupo = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    fill_sub = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    font_blanca = Font(bold=True, color="FFFFFF", size=10)
    font_negra = Font(bold=True, color="000000", size=9)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ==========================================
    # 2. ESTRUCTURA DE CABECERAS
    # ==========================================
    cabeceras_info = ['N°', 'DNI', 'NOMBRES Y APELLIDOS', 'CARGO', 'ÁREA', 'FECHA INGRESO', 'FECHA CESE']
    for i, titulo in enumerate(cabeceras_info, start=1):
        col = ws.cell(row=3, column=i)
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
        col.value = titulo
        col.fill = fill_grupo
        col.font = font_blanca
        col.alignment = align_center
        col.border = borde_fino

    grupos = [
        ('H3:J3', 'TAREO'),
        ('K3:N3', 'DESCANSOS'),
        ('O3:O3', 'FALTAS'),
        ('P3:Q3', 'TRABAJOS EXTRAS'),
        ('R3:V3', 'OTROS MOTIVOS'),
        ('W3:W4', 'TOTAL DÍAS')
    ]
    for rango, titulo in grupos:
        ws.merge_cells(rango)
        celda = ws[rango.split(':')[0]]
        celda.value = titulo
        celda.fill = PatternFill(start_color="244062", end_color="244062", fill_type="solid")
        celda.font = font_blanca
        celda.alignment = align_center
        celda.border = borde_fino

    sub_cabeceras = [
        ('H4', 'Días Laborados', 'A'), ('I4', 'Lic. Sin Goce', 'LSG'), ('J4', 'Suspensión', 'SU'),
        ('K4', 'Médicos', 'DM'), ('L4', 'Días Subsidiados', 'DS'), ('M4', 'Lic. Paternidad', 'LP'), ('N4', 'Fallecimiento', 'LD'),
        ('O4', 'Faltas', 'F'),
        ('P4', 'Feriados Trab.', 'FT'), ('Q4', 'Domingos Trab.', 'DT'), 
        ('R4', 'Lic. Con Goce', 'LG'), ('S4', 'Vacaciones', 'V'), ('T4', 'Día Compensado', 'DC'), 
        ('U4', 'Feriado Ganado', 'FG'), ('V4', 'Domingo Ganado', 'DG')
    ]
    
    mapa_columnas = {clave: ws[celda].column for celda, nombre, clave in sub_cabeceras}

    for celda, nombre, clave in sub_cabeceras:
        c = ws[celda]
        c.value = nombre
        c.fill = fill_sub
        c.font = font_negra
        c.alignment = align_center
        c.border = borde_fino

    # ==========================================
    # 3. PROCESAMIENTO DE DATOS EN BD Y FILTRADO
    # ==========================================
    # Convertimos las fechas del reporte a objetos 'date' desde el inicio
    inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    fin_dt = datetime.strptime(fin, "%Y-%m-%d").date()

    query_empleados = Empleado.query
    if empleado_id and empleado_id != "":
        query_empleados = query_empleados.filter(Empleado.id_empleado == empleado_id)
    empleados_db = query_empleados.all()
    
    empleados = []
    for emp in empleados_db:
        if not emp.dni or not emp.fecha_ingreso:
            continue # Si no tiene DNI o fecha de ingreso registrada, lo omitimos
            
        # Aseguramos que la fecha sea objeto date para poder compararla
        emp_ingreso = emp.fecha_ingreso.date() if isinstance(emp.fecha_ingreso, datetime) else emp.fecha_ingreso
        
        # REGLA 1: Si ingresó DESPUÉS del fin del reporte, no debe salir en el Excel
        if emp_ingreso > fin_dt:
            continue
            
        # REGLA 2: Si tiene fecha de cese y cesó ANTES del inicio del reporte, tampoco sale
        if emp.fecha_cese:
            emp_cese = emp.fecha_cese.date() if isinstance(emp.fecha_cese, datetime) else emp.fecha_cese
            if emp_cese < inicio_dt:
                continue
                
        # Si sobrevive a los filtros, es un empleado válido para este mes
        empleados.append(emp)
    
    # Inicializamos el diccionario SOLO para los empleados válidos del periodo
    datos_asistencia = {
        emp.dni: {k: 0 for k in mapa_columnas.keys()} for emp in empleados
    }

    tablas_asistencia = [
        EmpleadoLectura, EmpleadoDistribucion, EmpleadoInspecciones,
        EmpleadoCatastro, EmpleadoPersuasivas, EmpleadoMedidores,
        EmpleadoNorte, EmpleadoRecaudacion, EmpleadoAdministrativo
    ]

    registros_diarios = defaultdict(dict)

    for modelo in tablas_asistencia:
        query = db.session.query(
            modelo.dni, 
            modelo.fec_asist, 
            modelo.estado
        ).filter(
            modelo.fec_asist >= inicio,
            modelo.fec_asist <= fin
        )

        if empleado_id and empleado_id != "":
            query = query.filter(modelo.id_empleado == empleado_id)

        resultados = query.all()

        for row in resultados:
            if not row.dni or not row.fec_asist:
                continue
            
            fecha_dt = row.fec_asist.date() if isinstance(row.fec_asist, datetime) else row.fec_asist
            estado_raw = row.estado if row.estado is not None else ""
            estado_limpio = estado_raw.strip().upper()
            
            registros_diarios[row.dni][fecha_dt] = estado_limpio

    # ==========================================
    # 3.1 CONTEO DE COLUMNAS (A, F, V, etc.)
    # ==========================================
    for dni, fechas in registros_diarios.items():
        if dni in datos_asistencia:
            for fecha, estado in fechas.items():
                if estado not in ['', '-', 'NULL'] and estado in datos_asistencia[dni]:
                    datos_asistencia[dni][estado] += 1

    # ==========================================
    # 3.2 DETECCIÓN INTELIGENTE DE FERIADOS
    # ==========================================
    inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    fin_dt = datetime.strptime(fin, "%Y-%m-%d").date()
    
    feriados_dinamicos = set()
    dia_evaluacion = inicio_dt
    
    while dia_evaluacion <= fin_dt:
        if dia_evaluacion.weekday() < 6: # Solo Lunes a Sábado
            asistencias_normales = sum(
                1 for emp_fechas in registros_diarios.values() 
                if str(emp_fechas.get(dia_evaluacion, '')).strip() == 'A'
            )
            
            if asistencias_normales == 0: 
                feriados_dinamicos.add(dia_evaluacion)
                
        dia_evaluacion += timedelta(days=1)
        
    print(f"[SISTEMA] Feriados deducidos automáticamente: {feriados_dinamicos}")

    # ==========================================
    # 3.3 CÁLCULO DE DOMINGOS Y FERIADOS GANADOS
    # ==========================================
    for emp in empleados:
        if not emp.dni: continue

        emp_inicio = max(inicio_dt, emp.fecha_ingreso) if emp.fecha_ingreso else inicio_dt
        emp_fin = min(fin_dt, emp.fecha_cese) if emp.fecha_cese else fin_dt

        domingos_ganados = 0
        feriados_ganados = 0
        actual = emp_inicio

        while actual <= emp_fin:
            registro_dia = registros_diarios.get(emp.dni, {}).get(actual)
            es_vacio = registro_dia is None or str(registro_dia).strip() in ['', '-', 'NULL']

            if actual.weekday() == 6:  # Es domingo
                if es_vacio:
                    domingos_ganados += 1
            
            elif actual.weekday() < 6 and actual in feriados_dinamicos: # Es feriado inteligente
                if es_vacio:
                    feriados_ganados += 1
                    
            actual += timedelta(days=1)

        datos_asistencia[emp.dni]['DG'] = domingos_ganados
        datos_asistencia[emp.dni]['FG'] = feriados_ganados

    # ==========================================
    # 4. LLENADO DE FILAS
    # ==========================================
    fila_actual = 5
    for idx, emp in enumerate(empleados, start=1):
        if not emp.dni:
            continue
            
        nombres_completos = f"{emp.nombres or ''} {emp.apellidos or ''}".strip()
        asist = datos_asistencia.get(emp.dni, {k: 0 for k in mapa_columnas.keys()})
        
        ws.cell(row=fila_actual, column=1, value=idx) 
        ws.cell(row=fila_actual, column=2, value=emp.dni)
        ws.cell(row=fila_actual, column=3, value=nombres_completos.upper())
        ws.cell(row=fila_actual, column=4, value=emp.cargo or '')
        ws.cell(row=fila_actual, column=5, value=emp.area or '')
        ws.cell(row=fila_actual, column=6, value=emp.fecha_ingreso.strftime("%d/%m/%Y") if emp.fecha_ingreso else "")
        ws.cell(row=fila_actual, column=7, value=emp.fecha_cese.strftime("%d/%m/%Y") if emp.fecha_cese else "")

        total_dias = 0
        for clave, col_idx in mapa_columnas.items():
            valor = asist.get(clave, 0)
            celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
            celda.alignment = Alignment(horizontal="center")
            total_dias += valor

        celda_total = ws.cell(row=fila_actual, column=23, value=total_dias) # Columna W (23)
        celda_total.font = Font(bold=True)
        celda_total.alignment = Alignment(horizontal="center")

        for col in range(1, 24):
            ws.cell(row=fila_actual, column=col).border = borde_fino

        fila_actual += 1

    # ==========================================
    # 5. AJUSTES VISUALES
    # ==========================================
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(i) 
        
        for cell in col[4:]: 
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        
        adjusted_width = max(max_length + 3, 10)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    ws.freeze_panes = 'D5'
    ws.auto_filter.ref = f"A4:W{fila_actual-1}"

    # ==========================================
    # 6. DESCARGA
    # ==========================================
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Record_Asistencias_{inicio}_al_{fin}.xlsx'
    )

# =====================================================================
# 2. GENERACIÓN DE PDF EJECUTIVO (Diseño Profesional)
# =====================================================================
@app.route('/api/exportar/pdf', methods=['GET'])
def exportar_pdf():
    inicio = request.args.get('inicio', 'N/A')
    fin = request.args.get('fin', 'N/A')
    tipo = request.args.get('tipo', 'todos')
    empleado_id = request.args.get('empleado')

    # Consultas a la base de datos
    total_activos = db.session.query(Empleado).filter(Empleado.estado == 'ACTIVO').count()
    total_cesados = db.session.query(Empleado).filter(Empleado.estado == 'CESADO').count()
    fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")

    # 1. HTML con diseño Corporativo (Optimizado para xhtml2pdf)
    html_texto = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <style>
            /* Configuración de la página y pie de página */
            @page {{
                size: A4;
                margin: 2cm;
                @frame footer {{
                    -pdf-frame-content: footerContent;
                    bottom: 1cm;
                    margin-left: 2cm;
                    margin-right: 2cm;
                    height: 1cm;
                }}
            }}
            
            /* Estilos generales */
            body {{ font-family: Helvetica, sans-serif; color: #333333; font-size: 12px; }}
            h1 {{ color: #1a5276; font-size: 20px; text-align: center; margin-bottom: 5px; }}
            hr {{ border: 0; border-top: 2px solid #1a5276; margin-bottom: 25px; }}
            
            .section-title {{ font-size: 14px; color: #1a5276; margin-top: 30px; margin-bottom: 10px; font-weight: bold; border-bottom: 1px solid #dddddd; padding-bottom: 5px; }}
            
            /* Tabla de Información General */
            .info-table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            .info-table td {{ padding: 6px 0; vertical-align: top; }}
            .info-label {{ font-weight: bold; color: #555555; width: 150px; }}
            
            /* Tabla de Resumen (Métricas) */
            .stats-table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            .stats-table th {{ background-color: #1a5276; color: #ffffff; padding: 10px; text-align: left; font-weight: bold; font-size: 13px; }}
            .stats-table td {{ background-color: #f8f9fa; padding: 12px 10px; border-bottom: 1px solid #e0e0e0; font-size: 13px; }}
            .stats-number {{ font-weight: bold; font-size: 15px; text-align: right; }}
            
            /* Pie de página */
            #footerContent {{ text-align: right; font-size: 10px; color: #777777; border-top: 1px solid #dddddd; padding-top: 5px; }}
        </style>
    </head>
    <body>
        <h1>REPORTE EJECUTIVO DE ASISTENCIAS</h1>
        <hr>

        <table class="info-table">
            <tr>
                <td class="info-label">Fecha de Generación:</td>
                <td>{fecha_generacion}</td>
            </tr>
            <tr>
                <td class="info-label">Período Analizado:</td>
                <td>{inicio} al {fin}</td>
            </tr>
            <tr>
                <td class="info-label">Tipo de Filtro:</td>
                <td>{tipo.capitalize()}</td>
            </tr>
        </table>

        <div class="section-title">Resumen de Personal</div>

        <table class="stats-table">
            <tr>
                <th>Estado del Empleado</th>
                <th style="text-align: right;">Cantidad Total</th>
            </tr>
            <tr>
                <td>Total Empleados Activos</td>
                <td class="stats-number" style="color: #27ae60;">{total_activos}</td>
            </tr>
            <tr>
                <td>Total Empleados Cesados</td>
                <td class="stats-number" style="color: #c0392b;">{total_cesados}</td>
            </tr>
        </table>

        <div id="footerContent">
            Sistema de Gestión Interno | Página <pdf:pagenumber> de <pdf:pagecount>
        </div>
    </body>
    </html>
    """

    pdf_buffer = io.BytesIO()

    pisa_status = pisa.CreatePDF(
        html_texto, 
        dest=pdf_buffer,
        encoding='utf-8'
    )

    if pisa_status.err:
        return "Hubo un error al generar el PDF internamente", 500

    response = make_response(pdf_buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=Informe_Ejecutivo_{inicio}.pdf'
    
    return response



@app.route('/api/empleados/select', methods=['GET'])
def listar_empleados_select():
    empleados = db.session.query(
        Empleado.id_empleado,
        Empleado.nombres,
        Empleado.apellidos,
        Empleado.dni
    ).order_by(Empleado.apellidos.asc()).all()

    data = []
    for emp in empleados:
        nombre_completo = f"{emp.apellidos or ''} {emp.nombres or ''}".strip()
        data.append({
            'id': emp.id_empleado,
            'nombre': nombre_completo,
            'dni': emp.dni or ''
        })

    return jsonify(data)


# ==========================================
# 🛠️ GESTIÓN DE ALMACEN
# ==========================================
@app.route('/almacen/eliminar-proveedor', methods=['POST'])
def eliminar_proveedor():
    try:
        data = request.get_json()
        proveedor = Proveedor.query.get(data.get('id_proveedor'))
        
        if not proveedor:
            return jsonify({'error': 'Proveedor no encontrado'}), 404

        # Borrado Lógico: Cambiamos el estado
        proveedor.estado = 'INACTIVO'
        db.session.commit()
        return jsonify({'mensaje': 'Proveedor inactivado correctamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/almacen/reactivar-proveedor', methods=['POST'])
def reactivar_proveedor():
    try:
        data = request.get_json()
        proveedor = Proveedor.query.get(data.get('id_proveedor'))
        
        if not proveedor:
            return jsonify({'error': 'Proveedor no encontrado'}), 404

        proveedor.estado = 'ACTIVO'
        db.session.commit()
        return jsonify({'mensaje': 'Proveedor reactivado correctamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    

@app.route('/almacen/api/unidades', methods=['POST'])
def agregar_unidad():
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip().upper()

        if not nombre:
            return jsonify({'error': 'Nombre inválido'}), 400

        # Comprobar si ya existe
        existe = UnidadMedida.query.filter_by(nombre_unidad=nombre).first()
        if existe:
            return jsonify({'error': 'La unidad ya existe'}), 400

        # Guardar en BD
        nueva_unidad = UnidadMedida(nombre_unidad=nombre)
        db.session.add(nueva_unidad)
        db.session.commit()

        return jsonify({'mensaje': 'Unidad guardada', 'unidad': nombre}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500



@app.route('/almacen/api/unidades', methods=['GET'])
def obtener_unidades():
    try:
        unidades = UnidadMedida.query.order_by(UnidadMedida.nombre_unidad.asc()).all()
        # Transformamos la consulta a una lista de diccionarios
        lista_unidades = [{"nombre": u.nombre_unidad} for u in unidades]
        return jsonify(lista_unidades), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    

def generar_prefijo_categoria(nombre_cat):
    """Genera prefijo de 3 letras ignorando conectores y manejando colisiones de forma dinámica"""
    texto = nombre_cat.upper().strip()
    # Eliminar conectores comunes de strings
    texto_limpio = re.sub(r'\b(DE|DEL|EL|LA|LOS|LAS|Y|EN|PARA|CON|POR)\b', '', texto)
    palabras = texto_limpio.split()
    
    if len(palabras) >= 2:
        prefijo = palabras[0][0] + palabras[1][:2] # Primera letra de la 1° palabra + dos de la 2°
    elif len(palabras) == 1:
        prefijo = palabras[0][:3] # Primeras 3 letras
    else:
        prefijo = "CAT"
        
    prefijo = prefijo.ljust(3, 'X')[:3] # Forzar longitud exacta de 3 caracteres
    
    # Resolver colisiones en la BD (ej. MEDIDORES vs MEDICINA)
    base_prefijo = prefijo[:2]
    int_caracter = 65 # Código ASCII para 'A'
    while Categoria.query.filter_by(codigo_prefijo=prefijo).first():
        prefijo = base_prefijo + chr(int_caracter)
        int_caracter += 1
        if int_caracter > 90: # Romper bucle si se excede la Z
            prefijo = base_prefijo + str(int_caracter)
            break
    return prefijo


# ==========================================
# 🚀 RUTAS API ENDPOINTS
# ==========================================
@app.route('/almacen/crear-categoria', methods=['POST'])
def crear_categoria():
    data = request.get_json()
    nombre = data.get('tipo_categoria', '').strip().upper()
    
    if not nombre:
        return jsonify({"success": False, "message": "El nombre de categoría es requerido"}), 400
        
    # VALIDACIÓN: Verificar si la categoría ya existe en la BD
    cat_existente = Categoria.query.filter_by(tipo_categoria=nombre).first()
    if cat_existente:
        return jsonify({
            "success": False, 
            "message": f"¡Alerta! La categoría '{nombre}' ya existe con el prefijo {cat_existente.codigo_prefijo}."
        }), 400
        
    prefijo = generar_prefijo_categoria(nombre)
    nueva_cat = Categoria(tipo_categoria=nombre, codigo_prefijo=prefijo)
    db.session.add(nueva_cat)
    db.session.commit()
    
    return jsonify({
        "success": True, 
        "prefijo": prefijo,
        "categoria": {
            "id": nueva_cat.id_categoria,
            "texto": f"{nombre} ({prefijo})"
        }
    })


# NUEVO ENDPOINT: Para autogenerar el código identificador en vivo
@app.route('/almacen/siguiente-codigo/<int:id_cat>', methods=['GET'])
def obtener_siguiente_codigo(id_cat):
    categoria = Categoria.query.get(id_cat)
    if not categoria:
        return jsonify({"success": False}), 404
        
    total_productos = Producto.query.filter_by(id_categoria=id_cat).count()
    codigo_final = f"{categoria.codigo_prefijo}-{str(total_productos + 1).zfill(4)}"
    
    return jsonify({"success": True, "codigo": codigo_final})


@app.route('/almacen/crear-producto', methods=['POST'])
def crear_producto():
    data = request.get_json()
    nombre = data.get('nombre_prod', '').strip()
    id_cat = data.get('id_categoria')
    unidad = data.get('unidad_medida')
    precio = float(data.get('precio_igv') or 0.00)

    # Validación global de nombres repetidos
    prod_existente = Producto.query.filter_by(nombre_prod=nombre).first()
    if prod_existente:
        cat_duplicada = Categoria.query.get(prod_existente.id_categoria)
        return jsonify({
            "success": False, 
            "message": f"¡Alerta! El nombre '{nombre}' ya fue creado en la categoría '{cat_duplicada.tipo_categoria}'."
        }), 400

    categoria = Categoria.query.get(id_cat)
    total_productos = Producto.query.filter_by(id_categoria=id_cat).count()
    codigo_final = f"{categoria.codigo_prefijo}-{str(total_productos + 1).zfill(4)}"

    nuevo_prod = Producto(id_categoria=id_cat, codigo_identificador=codigo_final, nombre_prod=nombre, unidad_medida=unidad, precio_igv=precio)
    db.session.add(nuevo_prod)
    db.session.commit()
    return jsonify({"success": True, "codigo": codigo_final})


@app.route('/almacen/editar-producto', methods=['POST'])
def editar_producto():
    data = request.get_json()
    id_prod = data.get('id_producto')
    id_cat = data.get('id_categoria')
    nombre = data.get('nombre_prod', '').strip().upper() # Forzado a mayúsculas
    unidad = data.get('unidad_medida')
    precio = float(data.get('precio_igv') or 0.00)

    if not all([id_prod, id_cat, nombre, unidad]):
        return jsonify({"success": False, "message": "Faltan datos obligatorios."}), 400

    producto = Producto.query.get(id_prod)
    if not producto:
        return jsonify({"success": False, "message": "Producto no encontrado en la base de datos."}), 404

    # Validar que el nuevo nombre no exista en OTRO producto diferente al que estamos editando
    prod_existente = Producto.query.filter(Producto.nombre_prod == nombre, Producto.id_producto != id_prod).first()
    if prod_existente:
        cat_duplicada = Categoria.query.get(prod_existente.id_categoria)
        return jsonify({
            "success": False, 
            "message": f"¡Alerta! El nombre '{nombre}' ya le pertenece a otro producto en la categoría '{cat_duplicada.tipo_categoria}'."
        }), 400

    # Actualizamos los campos (El código identificador NO se toca)
    producto.id_categoria = id_cat
    producto.nombre_prod = nombre
    producto.unidad_medida = unidad
    producto.precio_igv = precio
    
    db.session.commit()
    
    return jsonify({"success": True})

@app.route('/almacen/api/listar-datos', methods=['GET'])
def api_listar_datos():
    try:
        categorias = Categoria.query.order_by(Categoria.tipo_categoria.asc()).all()
        productos = Producto.query.order_by(Producto.nombre_prod.asc()).all()
        proveedores = Proveedor.query.order_by(Proveedor.razon_social.asc()).all()
        empleados = Empleado.query.filter_by(estado='ACTIVO').order_by(Empleado.nombres.asc()).all()
        ultimas_salidas = MovimientoDetalle.query.filter_by(tipo_movimiento='SALIDA').order_by(MovimientoDetalle.id_movimiento.desc()).limit(30).all()

        lista_productos = []
        for p in productos:
            # Buscar el último movimiento
            ultimo_mov = MovimientoDetalle.query.filter_by(id_producto=p.id_producto, tipo_movimiento='ENTRADA').order_by(MovimientoDetalle.id_movimiento.desc()).first()
            
            # Siempre intentamos obtener el nombre del proveedor si existe movimiento
            nombre_prov = "-"
            if ultimo_mov and ultimo_mov.entrada_rel and ultimo_mov.entrada_rel.proveedor:
                nombre_prov = ultimo_mov.entrada_rel.proveedor.razon_social
            
            # Forzamos que sea True para que SIEMPRE aparezca en la lista
            en_inventario = True 

            lista_productos.append({
                "id_producto": p.id_producto,
                "id_categoria": p.id_categoria,
                "codigo": p.codigo_identificador,
                "nombre": p.nombre_prod,
                "unidad": p.unidad_medida,
                "precio_igv": float(p.precio_igv or 0.00),
                "stock": float(p.stock or 0.00),
                "categoria_nombre": p.categoria.tipo_categoria,
                "ultimo_proveedor": nombre_prov,
                "en_inventario": en_inventario 
            })

        # ==========================================
        # NUEVA LISTA PARA EL INVENTARIO SIN SUMAR (POR LOTES)
        # ==========================================
        movimientos = MovimientoDetalle.query.filter_by(tipo_movimiento='ENTRADA').order_by(MovimientoDetalle.id_movimiento.desc()).all()
        
        lista_inventario = []
        for m in movimientos:
            # 🚨 NUEVO: Buscamos el último conteo físico registrado para este lote específico
            ultimo_conteo = InventarioAuditoria.query.filter(
                InventarioAuditoria.observaciones.like(f"%Lote ID: {m.id_movimiento}%")
            ).order_by(InventarioAuditoria.id_auditoria.desc()).first()
            
            # Si hay conteo, lo pasamos como float. Si no, lo mandamos vacío.
            conteo_val = float(ultimo_conteo.conteo_fisico) if ultimo_conteo else ""

            lista_inventario.append({
                "id_movimiento": m.id_movimiento,
                "codigo": m.producto_rel.codigo_identificador,
                "nombre": m.producto_rel.nombre_prod,
                "talla": m.talla if m.talla else "-",
                "categoria": m.producto_rel.categoria.tipo_categoria,
                "unidad": m.producto_rel.unidad_medida,
                "precio_igv": float(m.precio_unitario or m.producto_rel.precio_igv or 0.00),
                "cantidad": float(m.stock_restante or 0),
                "fecha_ingreso": m.entrada_rel.fecha_ingreso.strftime('%d-%m-%Y') if m.entrada_rel and m.entrada_rel.fecha_ingreso else "-",
                "proveedor": m.entrada_rel.proveedor.razon_social if (m.entrada_rel and m.entrada_rel.proveedor) else "-",
                "conteo_fisico": conteo_val
            })

        # ==========================================
        # LISTA HISTÓRICA DE ENTRADAS
        # ==========================================
        ultimas_entradas = MovimientoDetalle.query.filter_by(tipo_movimiento='ENTRADA').order_by(MovimientoDetalle.id_movimiento.desc()).limit(30).all()

        # ==========================================
        # 🚨 NUEVO: LOTES DISPONIBLES (CON STOCK) PARA EL SELECT DE SALIDAS 🚨
        # ==========================================
        lotes_vivos = MovimientoDetalle.query.filter(
            MovimientoDetalle.tipo_movimiento == 'ENTRADA',
            MovimientoDetalle.stock_restante > 0,
            MovimientoDetalle.estado == 'ACTIVO'
        ).order_by(MovimientoDetalle.id_movimiento.asc()).all()

        lista_lotes_salida = []
        for lote in lotes_vivos:
            prov_nombre = lote.entrada_rel.proveedor.razon_social if lote.entrada_rel and lote.entrada_rel.proveedor else "-"
            fecha_ing = lote.entrada_rel.fecha_ingreso.strftime('%d-%m-%Y') if lote.entrada_rel and lote.entrada_rel.fecha_ingreso else "-"
            
            lista_lotes_salida.append({
                "id_lote": lote.id_movimiento, 
                "id_producto": lote.id_producto,
                "codigo": lote.producto_rel.codigo_identificador,
                "nombre": lote.producto_rel.nombre_prod,
                
                # 🚨 ENVIAMOS LA TALLA AL FRONTEND PARA EL SELECT
                "talla": lote.talla if lote.talla else "-", 
                
                "proveedor": prov_nombre,
                "fecha_ingreso": fecha_ing,
                "stock_restante": float(lote.stock_restante or 0),
                "precio": float(lote.precio_unitario or lote.producto_rel.precio_igv or 0.00)
            })

        # ==========================================
        # ESTRUCTURA FINAL DE RESPUESTA
        # ==========================================
        data = {
            "categorias": [{ "id": c.id_categoria, "texto_select": f"{c.tipo_categoria} ({c.codigo_prefijo})", "nombre": c.tipo_categoria } for c in categorias],
            "productos": lista_productos,
            "proveedores": [{ "id_proveedor": pr.id_proveedor, "ruc": pr.ruc, "razon_social": pr.razon_social, "nombre_comercial": pr.nombre_comercial or "", "celular": pr.celular or "", "correo": pr.correo or "", "direccion": pr.direccion or "", "estado": pr.estado } for pr in proveedores],
            "inventario_fisico": lista_inventario, 
            
            "entradas": [{ 
                "id_mov": e.id_movimiento,
                "fecha_fac": e.entrada_rel.fecha_factura.strftime('%d-%m-%Y') if e.entrada_rel.fecha_factura else "-",
                "fecha_ing": e.entrada_rel.fecha_ingreso.strftime('%d-%m-%Y') if e.entrada_rel.fecha_ingreso else "-",
                "factura": e.entrada_rel.nro_factura,
                "guia": e.entrada_rel.nro_guia or "-",
                "codigo": e.producto_rel.codigo_identificador,
                "producto": e.producto_rel.nombre_prod,
                "talla": e.talla if e.talla else "-",
                "empleado_recupero": Empleado.query.get(e.id_empleado_recupero).nombres if e.id_empleado_recupero else "-",
                "cantidad": float(e.cantidad),
                "precio": float(e.precio_unitario or e.producto_rel.precio_igv or 0),
                "proveedor": e.entrada_rel.proveedor.razon_social if e.entrada_rel.proveedor else "-",
                "obs": e.entrada_rel.obs_entrada or "-"
            } for e in ultimas_entradas],

            "empleados": [{
                "id_empleado": emp.id_empleado, 
                "nombres": emp.nombres or "SIN NOMBRE",
                "area": emp.area or ""   
            } for emp in empleados],

            "salidas": [{
                "id_mov": s.id_movimiento,
                "fecha_salida": s.salida_rel.fecha_salida.strftime('%d-%m-%Y') if s.salida_rel and s.salida_rel.fecha_salida else "-",
                "cantidad": float(s.cantidad),
                "codigo": s.producto_rel.codigo_identificador,
                "producto": s.producto_rel.nombre_prod,
                
                # 🚨 ENVIAMOS LA TALLA TAMBIÉN AL HISTORIAL DE SALIDAS
                "talla": s.talla if s.talla else "-",
                
                "empleado": s.salida_rel.empleado.nombres if s.salida_rel and s.salida_rel.empleado else "-",
                "area": s.salida_rel.empleado.area if s.salida_rel and s.salida_rel.empleado else "-", 
                "obs": s.salida_rel.obs_salida or "-"
            } for s in ultimas_salidas],

            "lotes_disponibles": lista_lotes_salida
        }
        return jsonify(data)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/almacen/crear-proveedor', methods=['POST'])
def crear_proveedor():
    data = request.get_json()
    
    # Validar duplicado por RUC
    ruc_prov = data.get('ruc')
    if Proveedor.query.filter_by(ruc=ruc_prov).first():
        return jsonify({"success": False, "message": f"El proveedor con RUC {ruc_prov} ya está registrado."}), 400
        
    nuevo_prov = Proveedor(
        ruc=ruc_prov,
        razon_social=data.get('razon_social').strip().upper(),
        nombre_comercial=data.get('nombre_comercial').strip().upper() if data.get('nombre_comercial') else None,
        celular=data.get('celular'),
        correo=data.get('correo'),
        direccion=data.get('direccion')
    )
    db.session.add(nuevo_prov)
    db.session.commit()
    
    return jsonify({"success": True})

@app.route('/almacen/editar-proveedor', methods=['POST'])
def editar_proveedor():
    data = request.get_json()
    id_prov = data.get('id_proveedor')
    ruc = data.get('ruc')
    razon_social = data.get('razon_social').strip().upper()

    if not all([id_prov, ruc, razon_social]):
        return jsonify({"success": False, "message": "Faltan datos obligatorios (RUC o Razón Social)."}), 400

    # Validar que no se duplique el RUC con OTRO proveedor distinto
    prov_existente = Proveedor.query.filter(Proveedor.ruc == ruc, Proveedor.id_proveedor != id_prov).first()
    if prov_existente:
        return jsonify({"success": False, "message": f"¡Alerta! El RUC '{ruc}' ya pertenece a otro proveedor."}), 400

    proveedor = Proveedor.query.get(id_prov)
    if not proveedor:
        return jsonify({"success": False, "message": "Proveedor no encontrado."}), 404

    proveedor.ruc = ruc
    proveedor.razon_social = razon_social
    proveedor.nombre_comercial = data.get('nombre_comercial').strip().upper() if data.get('nombre_comercial') else None
    proveedor.celular = data.get('celular')
    proveedor.correo = data.get('correo')
    proveedor.direccion = data.get('direccion')
    
    db.session.commit()
    
    return jsonify({"success": True})


@app.route('/almacen/guardar-inventario', methods=['POST'])
def guardar_inventario():
    data = request.get_json()
    id_prod = data.get('id_producto')
    
    if not id_prod:
        return jsonify({"success": False, "message": "Debe seleccionar un producto."}), 400
        
    producto = Producto.query.get(id_prod)
    if not producto:
        return jsonify({"success": False, "message": "Producto no encontrado."}), 404

    # Convertimos a float, si viene vacío lo dejamos en 0.00
    try:
        nuevo_stock = float(data.get('stock') or 0)
        nuevo_precio = float(data.get('precio_igv') or 0)
    except ValueError:
        return jsonify({"success": False, "message": "Valores de stock o precio inválidos."}), 400

    producto.stock = nuevo_stock
    producto.precio_igv = nuevo_precio
    
    # Opcional: Aquí podrías registrar también en tu tabla InventarioAuditoria si envían conteo_fisico
    
    db.session.commit()
    
    return jsonify({"success": True, "message": "Inventario actualizado correctamente."})


@app.route('/almacen/guardar-conteo-fisico', methods=['POST'])
def guardar_conteo_fisico():
    data = request.get_json()
    # El JS envía el 'idMov' bajo el key 'id_producto'
    id_mov = data.get('id_producto') 
    conteo = data.get('conteo_fisico')
    
    if not id_mov or conteo is None:
        return jsonify({"success": False, "message": "Datos incompletos."}), 400
        
    # 🚨 LA CORRECCIÓN CLAVE: Buscamos en MovimientoDetalle, NO en Producto
    movimiento = MovimientoDetalle.query.get(id_mov)
    if not movimiento:
        return jsonify({"success": False, "message": "Lote no encontrado."}), 404

    try:
        conteo_float = float(conteo)
        # Comparamos contra el stock de ESTE lote específico
        stock_actual = float(movimiento.stock_restante or 0)
        diferencia = conteo_float - stock_actual
        
        # Guardamos la auditoría
        nueva_auditoria = InventarioAuditoria(
            id_producto=movimiento.id_producto, # Ahora sí sacamos el ID del producto real
            id_empleado_auditor=1, # Ojo: Asegúrate de poner tu session['user_id'] si tienes login
            stock_sistema=stock_actual,
            conteo_fisico=conteo_float,
            diferencia=diferencia,
            observaciones=f"Conteo físico desde tabla interactiva (Lote ID: {id_mov})"
        )
        db.session.add(nueva_auditoria)
        db.session.commit()
        
        return jsonify({"success": True})
        
    except ValueError:
        return jsonify({"success": False, "message": "El conteo debe ser un número válido."}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
        

@app.route('/almacen/guardar-entrada-lote', methods=['POST'])
def guardar_entrada_lote():
    data = request.get_json()
    
    print("\n\n====== 🚨 INICIO DEBUG GUARDAR LOTE 🚨 ======")
    print(f"JSON COMPLETO RECIBIDO: {data}")
    
    cabecera = data.get('cabecera')
    detalles = data.get('detalles')
    
    if not cabecera or not detalles or len(detalles) == 0:
        return jsonify({"success": False, "message": "No hay productos para guardar."}), 400

    try:
        f_fac = datetime.strptime(cabecera['fecha_fac'], '%Y-%m-%d').date() if cabecera.get('fecha_fac') else None
        f_ing = datetime.strptime(cabecera['fecha_ing'], '%Y-%m-%d') if cabecera.get('fecha_ing') else datetime.utcnow()
        nro_factura = cabecera.get('factura', '').strip().upper()
        id_prov = cabecera.get('id_proveedor')

        # 1. Buscamos si la cabecera (Factura) ya existe, sino la creamos
        entrada = Entrada.query.filter_by(nro_factura=nro_factura, id_proveedor=id_prov).first()
        if not entrada:
            entrada = Entrada(
                id_proveedor=id_prov,
                id_empleado_receptor=1, # Reemplaza con tu variable de sesión
                fecha_ingreso=f_ing,
                fecha_factura=f_fac,
                nro_factura=nro_factura,
                nro_guia=cabecera.get('guia', '').strip().upper(),
                obs_entrada="" 
            )
            db.session.add(entrada)
            db.session.flush() # Obtenemos el ID generado

        # 2. Recorremos el lote de productos
        for item in detalles:
            print(f"\n--- PROCESANDO PRODUCTO ID: {item.get('id_producto')} ---")
            
            cant_float = float(item['cantidad'])
            precio_float = float(item['precio'])
            
            # 🚨 CAPTURANDO Y VERIFICANDO LA TALLA
            talla_cruda = item.get('talla')
            print(f"1. Talla cruda que llegó de JavaScript: '{talla_cruda}' (Tipo: {type(talla_cruda)})")
            
            talla_val = talla_cruda
            if talla_val == '-' or not talla_val:
                talla_val = None  
                print("2. Decisión: La talla se guardará como NULL (Vació o Guion)")
            else:
                talla_val = str(talla_val).strip().upper()
                print(f"2. Decisión: La talla válida a guardar será: '{talla_val}'")
            
            id_emp_rec = item.get('id_empleado_recupero')
            if not id_emp_rec or id_emp_rec == '':
                id_emp_rec = None
            
            # Guardamos el detalle del movimiento
            movimiento = MovimientoDetalle(
                id_entrada=entrada.id_entrada,
                id_producto=item['id_producto'],
                tipo_movimiento='ENTRADA',
                cantidad=cant_float,
                precio_unitario=precio_float, 
                stock_restante=cant_float,    
                estado='ACTIVO',
                talla=talla_val,
                id_empleado_recupero=id_emp_rec
            )
            db.session.add(movimiento)
            
            print(f"3. Objeto MovimientoDetalle preparado en memoria. Talla asignada: {movimiento.talla}")

            producto = Producto.query.get(item['id_producto'])
            if producto:
                # Actualizamos stock global
                producto.stock = float(producto.stock or 0) + cant_float
                
                # 🚨 FOTO HISTÓRICA: Guardamos el stock resultante en el movimiento
                movimiento.stock_historico = producto.stock

        db.session.commit()
        print("====== ✅ FIN DEBUG: GUARDADO EXITOSO EN MYSQL ✅ ======\n")
        return jsonify({"success": True})
        
    except Exception as e:
        db.session.rollback() 
        print(f"====== ❌ FIN DEBUG: ERROR AL GUARDAR ❌ ======")
        print(f"MOTIVO DEL ERROR: {str(e)}\n")
        return jsonify({"success": False, "message": f"Error al procesar: {str(e)}"}), 500
    

@app.route('/almacen/guardar-salida-lote', methods=['POST'])
def guardar_salida_lote():
    data = request.get_json()
    cabecera = data.get('cabecera')
    detalles = data.get('detalles')
    
    if not cabecera or not detalles or len(detalles) == 0:
        return jsonify({"success": False, "message": "No hay productos para despachar."}), 400

    try:
        fecha_str = cabecera.get('fecha')
        fecha_salida = datetime.strptime(fecha_str, '%Y-%m-%d') if fecha_str else datetime.utcnow()
        id_emp = cabecera.get('id_empleado')

        # 1. Validar Stock del Lote Específico
        #for item in detalles:
            # Aquí 'item['id_lote']' ES EL ID DEL LOTE ESPECÍFICO (La fila del Kardex de entrada)
            #lote_seleccionado = MovimientoDetalle.query.get(int(item['id_lote']))
            #cant_req = float(item['cantidad'])
            
            #if not lote_seleccionado or cant_req > float(lote_seleccionado.stock_restante or 0):
                #return jsonify({
                    #"success": False, 
                    #"message": f"Stock insuficiente en el lote seleccionado. Solo quedan {lote_seleccionado.stock_restante} unidades."
                #}), 400

        # 2. Guardar Cabecera
        nueva_salida = Salida(
            id_empleado_solicitante=id_emp,
            fecha_salida=fecha_salida,
            obs_salida=cabecera.get('area', '') 
        )
        db.session.add(nueva_salida)
        db.session.flush() 

        # 3. Descontar del Lote Específico y del Producto Global
        for item in detalles:
            cant_a_restar = float(item['cantidad'])
            lote_especifico = MovimientoDetalle.query.get(int(item['id_lote']))
            producto = Producto.query.get(lote_especifico.id_producto)
            
            # Registrar movimiento
            movimiento = MovimientoDetalle(
                id_salida=nueva_salida.id_salida,
                id_producto=lote_especifico.id_producto,
                tipo_movimiento='SALIDA',
                cantidad=cant_a_restar,
                precio_unitario=lote_especifico.precio_unitario, # Mantiene el costo de esa entrada
                estado='ACTIVO',
                talla=lote_especifico.talla,
                id_lote_origen=lote_especifico.id_movimiento
            )
            db.session.add(movimiento)

            # Descontar del lote específico (El que el usuario eligió)
            lote_especifico.stock_restante = float(lote_especifico.stock_restante) - cant_a_restar
            
            # Descontar del inventario global
            if producto:
                producto.stock = float(producto.stock or 0) - cant_a_restar
                
                # 🚨 FOTO HISTÓRICA: Guardamos el stock resultante en el movimiento
                movimiento.stock_historico = producto.stock

        db.session.commit()
        return jsonify({"success": True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"Error interno: {str(e)}"}), 500


@app.route('/almacen/api/historico-kardex', methods=['GET'])
def api_historico_kardex():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('limit', 20, type=int)
        tipo = request.args.get('tipo', 'TODO', type=str)
        search = request.args.get('search', '', type=str)
        
        # 1. CAPTURAR LAS FECHAS
        fecha_inicio = request.args.get('fecha_inicio', '', type=str)
        fecha_fin = request.args.get('fecha_fin', '', type=str)
        
        # 2. INICIAR CONSULTA CON JOINS Y ALIAS
        EmpleadoSolicitante = aliased(Empleado)
        EmpleadoRetorno = aliased(Empleado)

        query = MovimientoDetalle.query.join(Producto)
        query = query.outerjoin(Categoria, Producto.id_categoria == Categoria.id_categoria)
        query = query.outerjoin(Entrada, MovimientoDetalle.id_entrada == Entrada.id_entrada)
        query = query.outerjoin(Proveedor, Entrada.id_proveedor == Proveedor.id_proveedor)
        query = query.outerjoin(Salida, MovimientoDetalle.id_salida == Salida.id_salida)
        
        # 🚨 UNIMOS LA TABLA EMPLEADO DOS VECES (Una para salida, otra para retorno)
        query = query.outerjoin(EmpleadoSolicitante, Salida.id_empleado_solicitante == EmpleadoSolicitante.id_empleado)
        query = query.outerjoin(EmpleadoRetorno, MovimientoDetalle.id_empleado_recupero == EmpleadoRetorno.id_empleado)

        # 3. FILTROS EXISTENTES
        if tipo != 'TODO':
            query = query.filter(MovimientoDetalle.tipo_movimiento == tipo)
            
        if search:
            palabras = search.split()
            for palabra in palabras:
                search_term = f"%{palabra}%"
                query = query.filter(or_(
                    Producto.codigo_identificador.ilike(search_term),
                    Producto.nombre_prod.ilike(search_term),
                    Categoria.tipo_categoria.ilike(search_term),
                    Proveedor.razon_social.ilike(search_term),
                    
                    # 🚨 BUSCAMOS EN EL EMPLEADO SOLICITANTE (Salidas)
                    EmpleadoSolicitante.nombres.ilike(search_term),
                    EmpleadoSolicitante.area.ilike(search_term),
                    
                    # 🚨 Y TAMBIÉN BUSCAMOS EN EL EMPLEADO DE RETORNO (Entradas)
                    EmpleadoRetorno.nombres.ilike(search_term)
                ))

        # 4. APLICAR FILTRO DE FECHAS (Cubriendo ambos tipos de movimiento)
        if fecha_inicio:
            # Añadimos 00:00:00 para asegurar que tome todo el día desde la madrugada
            query = query.filter(or_(
                Entrada.fecha_ingreso >= f"{fecha_inicio} 00:00:00",
                Salida.fecha_salida >= f"{fecha_inicio} 00:00:00"
            ))
            
        if fecha_fin:
            # Añadimos 23:59:59 para incluir hasta el último segundo del día seleccionado
            query = query.filter(or_(
                Entrada.fecha_ingreso <= f"{fecha_fin} 23:59:59",
                Salida.fecha_salida <= f"{fecha_fin} 23:59:59"
            ))

        # 5. ORDENAR Y PAGINAR
        movimientos_paginados = query.order_by(MovimientoDetalle.id_movimiento.desc()).paginate(page=page, per_page=per_page, error_out=False)
        
        lista_historial = []
        for m in movimientos_paginados.items:
            es_entrada = m.tipo_movimiento == 'ENTRADA'
            
            # 🚨 INICIAMOS LA VARIABLE TALLA
            talla_mostrar = m.talla 
            
            # Lógica para cruzar datos según si es Entrada o Salida
            if es_entrada:
                fecha = m.entrada_rel.fecha_ingreso
                f_fac = m.entrada_rel.fecha_factura.strftime('%d-%m-%Y') if m.entrada_rel.fecha_factura else "-"
                doc_ref = m.entrada_rel.nro_factura or "-"
                guia = m.entrada_rel.nro_guia or "-"
                prov = m.entrada_rel.proveedor.razon_social if m.entrada_rel.proveedor else "-"
                obs = m.entrada_rel.obs_entrada
                
                # 🚨 CAPTURAMOS EL EMPLEADO DE RECUPERO SI EXISTE
                emp_retorno = Empleado.query.get(m.id_empleado_recupero).nombres if m.id_empleado_recupero else "-"
                
            else:
                fecha = m.salida_rel.fecha_salida
                # Rastreamos la última entrada de este producto para "prestarle" la factura y guía a esta salida
                ultima_entrada = MovimientoDetalle.query.filter_by(id_producto=m.id_producto, tipo_movimiento='ENTRADA').order_by(MovimientoDetalle.id_movimiento.desc()).first()
                
                # 🚨 HEREDAMOS LA TALLA SI LA SALIDA NO LA TIENE GRABADA
                if not talla_mostrar and ultima_entrada and ultima_entrada.talla:
                     talla_mostrar = ultima_entrada.talla
                
                f_fac = ultima_entrada.entrada_rel.fecha_factura.strftime('%d-%m-%Y') if (ultima_entrada and ultima_entrada.entrada_rel.fecha_factura) else "-"
                doc_ref = ultima_entrada.entrada_rel.nro_factura if ultima_entrada else "-"
                guia = ultima_entrada.entrada_rel.nro_guia if ultima_entrada else "-"
                prov = "-"
                
                # 🚨 LAS SALIDAS NORMALES NO TIENEN EMPLEADO DE RECUPERO
                emp_retorno = "-" 
                
                obs = m.salida_rel.obs_salida

            # Datos del Empleado Solicitante (El que se lleva la salida)
            emp = m.salida_rel.empleado.nombres if (not es_entrada and m.salida_rel.empleado) else "-"
            area = m.salida_rel.empleado.area if (not es_entrada and m.salida_rel.empleado) else "-"
            cargo = m.salida_rel.empleado.cargo if (not es_entrada and m.salida_rel.empleado) else "-"

            lista_historial.append({
                "id_mov": m.id_movimiento,
                "fecha": fecha.strftime('%d-%m-%Y') if fecha else "-",
                "tipo": m.tipo_movimiento,
                "codigo": m.producto_rel.codigo_identificador,
                "producto": m.producto_rel.nombre_prod,
                "talla": talla_mostrar if talla_mostrar else "-", # 🚨 USAMOS TALLA MOSTRAR
                "unidad": m.producto_rel.unidad_medida or "-",
                "categoria": m.producto_rel.categoria.tipo_categoria,
                "cantidad": float(m.cantidad),
                "stock_actual": float(m.stock_historico if m.stock_historico is not None else m.producto_rel.stock), 
                "proveedor": prov,
                
                "empleado_recupero": emp_retorno, # 🚨 ENVIAMOS EL EMPLEADO DE RECUPERO AL FRONTEND
                
                "empleado": emp,
                "area": area,
                "cargo": cargo,
                "documento": doc_ref,
                "fecha_factura": f_fac,
                "guia": guia,
                "obs": obs or "-"
            })

        # 6. Devolver el JSON con metadata de paginación
        return jsonify({
            "success": True,
            "data": lista_historial,
            "pagination": {
                "total_records": movimientos_paginados.total,
                "current_page": movimientos_paginados.page,
                "total_pages": movimientos_paginados.pages,
                "per_page": movimientos_paginados.per_page
            }
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/almacen/eliminar-movimiento', methods=['POST'])
def eliminar_movimiento():
    data = request.get_json()
    id_mov = data.get('id_mov')
    
    movimiento = MovimientoDetalle.query.get(id_mov)
    if not movimiento:
        return jsonify({"success": False, "message": "Movimiento no encontrado."}), 404
        
    producto = Producto.query.get(movimiento.id_producto)
    
    try:
        cant_float = float(movimiento.cantidad)
        
        # 1. REVERTIMOS EL STOCK
        if movimiento.tipo_movimiento == 'SALIDA':
            # A. Regresamos el stock al total global (+)
            producto.stock = float(producto.stock or 0) + cant_float
            
            # B. 🚨 MAGIA OPCIÓN 1: Devolvemos el stock exacto al lote origen de donde salió
            if movimiento.id_lote_origen:
                lote_origen = MovimientoDetalle.query.get(movimiento.id_lote_origen)
                if lote_origen:
                    lote_origen.stock_restante = float(lote_origen.stock_restante or 0) + cant_float
            
            # C. Borramos el registro del detalle
            id_padre_salida = movimiento.id_salida
            db.session.delete(movimiento)
            
            # D. Limpieza: Si este era el último producto de esa Boleta de Salida, borramos la boleta.
            detalles_restantes = MovimientoDetalle.query.filter_by(id_salida=id_padre_salida).count()
            if detalles_restantes == 0:
                padre_salida = Salida.query.get(id_padre_salida)
                if padre_salida:
                    db.session.delete(padre_salida)
            
        elif movimiento.tipo_movimiento == 'ENTRADA':
            # Seguridad: Verificar si ya se despachó mercancía de este lote
            stock_restante = float(movimiento.stock_restante or 0)
            if stock_restante < cant_float:
                return jsonify({
                    "success": False, 
                    "message": f"No se puede eliminar la entrada. Ya se han despachado productos de este lote (Quedan {stock_restante} de {cant_float}). Elimine primero las salidas vinculadas."
                }), 400

            if float(producto.stock or 0) < cant_float:
                return jsonify({
                    "success": False, 
                    "message": f"No se puede eliminar. El stock actual de {producto.nombre_prod} es insuficiente para restar {cant_float}."
                }), 400
                
            producto.stock = float(producto.stock or 0) - cant_float

            # Borramos el detalle
            id_padre_entrada = movimiento.id_entrada
            db.session.delete(movimiento)
            
            # Limpieza: Si era el único producto de esta Factura, borramos la Factura.
            detalles_restantes = MovimientoDetalle.query.filter_by(id_entrada=id_padre_entrada).count()
            if detalles_restantes == 0:
                padre_entrada = Entrada.query.get(id_padre_entrada)
                if padre_entrada:
                    db.session.delete(padre_entrada)

        # 2. CONFIRMAMOS CAMBIOS
        db.session.commit()
        
        return jsonify({"success": True, "message": "Registro eliminado y stock recalculado correctamente."})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/almacen/api/exportar-excel-kardex', methods=['GET'])
def exportar_excel_kardex():
    tipo = request.args.get('tipo', 'TODO', type=str)
    search = request.args.get('search', '', type=str)
    fecha_inicio = request.args.get('fecha_inicio', '', type=str)
    fecha_fin = request.args.get('fecha_fin', '', type=str)

    # 1. REPLICAR LA CONSULTA EXACTA DE LOS FILTROS
    EmpleadoSolicitante = aliased(Empleado)
    EmpleadoRetorno = aliased(Empleado)

    query = MovimientoDetalle.query.join(Producto)
    query = query.outerjoin(Categoria, Producto.id_categoria == Categoria.id_categoria)
    query = query.outerjoin(Entrada, MovimientoDetalle.id_entrada == Entrada.id_entrada)
    query = query.outerjoin(Proveedor, Entrada.id_proveedor == Proveedor.id_proveedor)
    query = query.outerjoin(Salida, MovimientoDetalle.id_salida == Salida.id_salida)
        
    query = query.outerjoin(EmpleadoSolicitante, Salida.id_empleado_solicitante == EmpleadoSolicitante.id_empleado)
    query = query.outerjoin(EmpleadoRetorno, MovimientoDetalle.id_empleado_recupero == EmpleadoRetorno.id_empleado)

    if tipo != 'TODO':
            query = query.filter(MovimientoDetalle.tipo_movimiento == tipo)
                
    if search:
        palabras = search.split()
        for palabra in palabras:
            search_term = f"%{palabra}%"
            query = query.filter(or_(
                    Producto.codigo_identificador.ilike(search_term),
                    Producto.nombre_prod.ilike(search_term),
                    Categoria.tipo_categoria.ilike(search_term),
                    Proveedor.razon_social.ilike(search_term),
                    
                    # BUSCAMOS EN EL EMPLEADO SOLICITANTE (Salidas)
                    EmpleadoSolicitante.nombres.ilike(search_term),
                    EmpleadoSolicitante.area.ilike(search_term),
                    
                    # Y TAMBIÉN BUSCAMOS EN EL EMPLEADO DE RETORNO (Entradas)
                    EmpleadoRetorno.nombres.ilike(search_term)
                ))
                
    if fecha_inicio and fecha_inicio.strip() != "":
        query = query.filter(or_(
            Entrada.fecha_ingreso >= f"{fecha_inicio} 00:00:00",
            Salida.fecha_salida >= f"{fecha_inicio} 00:00:00"
        ))
        
    if fecha_fin and fecha_fin.strip() != "":
        query = query.filter(or_(
            Entrada.fecha_ingreso <= f"{fecha_fin} 23:59:59",
            Salida.fecha_salida <= f"{fecha_fin} 23:59:59"
        ))

    # Obtenemos TODOS los registros filtrados (sin .paginate())
    movimientos = query.order_by(MovimientoDetalle.id_movimiento.desc()).all()

    # 2. CREAR EL EXCEL Y ESTILOS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico de Movimientos"

    # Definir Estilos
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Escribir Cabeceras (🚨 Añadimos Empleado de Retorno)
    headers = [
        "Fecha", "Tipo", "Cód. Producto", "Nombre Producto", "Talla", "Unidad", "Categoría", 
        "Cantidad", "Stock Final", "Proveedor (Compras)", "Empleado (Retorno)", "Empleado Solicitante", 
        "Área", "Cargo", "Doc. Referencia", "Fecha Factura", "Guía", "Observación"
    ]
    
    ws.append(headers)
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin

    # 3. LLENAR LOS DATOS
    for m in movimientos:
        es_entrada = m.tipo_movimiento == 'ENTRADA'
        talla_val = m.talla 
        
        if es_entrada:
            fecha = m.entrada_rel.fecha_ingreso.strftime('%d-%m-%Y') if m.entrada_rel and m.entrada_rel.fecha_ingreso else "-"
            prov = m.entrada_rel.proveedor.razon_social if m.entrada_rel and m.entrada_rel.proveedor else "-"
            doc = m.entrada_rel.nro_factura or "-"
            f_fac = m.entrada_rel.fecha_factura.strftime('%d-%m-%Y') if m.entrada_rel and m.entrada_rel.fecha_factura else "-"
            guia = m.entrada_rel.nro_guia or "-"
            obs = m.entrada_rel.obs_entrada or "-"
            
            # Rescatamos el empleado de recupero si lo hay
            emp_retorno = Empleado.query.get(m.id_empleado_recupero).nombres if m.id_empleado_recupero else "-"
            
            emp, area, cargo = "-", "-", "-"
            cantidad_str = f"+{m.cantidad}"
            cant_color = "16A34A" # Verde
        else:
            fecha = m.salida_rel.fecha_salida.strftime('%d-%m-%Y') if m.salida_rel and m.salida_rel.fecha_salida else "-"
            
            # HEREDAMOS TALLA Y DOCS DE LA ÚLTIMA ENTRADA (Igual que en la ruta JSON)
            ultima_entrada = MovimientoDetalle.query.filter_by(id_producto=m.id_producto, tipo_movimiento='ENTRADA').order_by(MovimientoDetalle.id_movimiento.desc()).first()
            if not talla_val and ultima_entrada and ultima_entrada.talla:
                talla_val = ultima_entrada.talla
                
            f_fac = ultima_entrada.entrada_rel.fecha_factura.strftime('%d-%m-%Y') if (ultima_entrada and ultima_entrada.entrada_rel.fecha_factura) else "-"
            doc = ultima_entrada.entrada_rel.nro_factura if ultima_entrada else "-"
            guia = ultima_entrada.entrada_rel.nro_guia if ultima_entrada else "-"
            
            emp = m.salida_rel.empleado.nombres if m.salida_rel and m.salida_rel.empleado else "-"
            area = m.salida_rel.empleado.area if m.salida_rel and m.salida_rel.empleado else "-"
            cargo = m.salida_rel.empleado.cargo if m.salida_rel and m.salida_rel.empleado else "-"
            obs = m.salida_rel.obs_salida or "-"
            
            prov = "-"
            emp_retorno = "-" # Las salidas no tienen empleado de retorno
            
            cantidad_str = f"-{m.cantidad}"
            cant_color = "EF4444" # Rojo
            
        talla_mostrar = talla_val if talla_val else "-"
        
        # 🚨 LA CORRECCIÓN DEL HISTÓRICO: Tomar la foto o usar el actual
        stock_kardex = float(m.stock_historico if m.stock_historico is not None else m.producto_rel.stock)
            
        # INCORPORAMOS 'stock_kardex' EN LA POSICIÓN CORRECTA
        row_data = [
            fecha, m.tipo_movimiento, m.producto_rel.codigo_identificador, 
            m.producto_rel.nombre_prod, talla_mostrar, m.producto_rel.unidad_medida, 
            m.producto_rel.categoria.tipo_categoria, cantidad_str, stock_kardex, # <--- 🚨 CORRECCIÓN APLICADA AQUÍ
            prov, emp_retorno, emp, area, cargo, doc, f_fac, guia, obs
        ]
        
        ws.append(row_data)
        
        # Estilos por fila (bordes y colores dinámicos)
        current_row = ws[ws.max_row]
        for idx, cell in enumerate(current_row):
            cell.border = border_thin
            # Índices actualizados (3:Nombre, 9:Prov, 10:EmpRetorno, 11:Emp, 17:Obs)
            cell.alignment = align_left if idx in [3, 9, 10, 11, 17] else align_center 
            
            # Pintar la celda de Cantidad de Verde o Rojo (Ahora es el índice 7)
            if idx == 7: 
                cell.font = Font(color=cant_color, bold=True)

    # 4. AUTOAJUSTAR ANCHO DE COLUMNAS (Se añade 'R' y se mueven las letras)
    column_widths = {
        'A': 12, 'B': 12, 'C': 15, 'D': 40, 'E': 10, 'F': 10, 'G': 18, 
        'H': 12, 'I': 12, 'J': 35, 'K': 35, 'L': 35, 'M': 20, 
        'N': 20, 'O': 18, 'P': 15, 'Q': 18, 'R': 40
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 5. PREPARAR DESCARGA
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output, 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        as_attachment=True, 
        download_name="Reporte_Kardex.xlsx"
    )


@app.route('/almacen/api/exportar-excel-inventario', methods=['GET'])
def exportar_excel_inventario():
    # 1. Recibimos la búsqueda, limpiamos espacios extra y convertimos a mayúsculas
    search = request.args.get('search', '', type=str).strip().upper()
    
    # 2. Separamos la búsqueda en palabras individuales (ej: ["CINTA", "NEGRA"])
    palabras_busqueda = search.split() if search else []

    # Traemos todos los lotes de entrada
    movimientos = MovimientoDetalle.query.filter_by(tipo_movimiento='ENTRADA').order_by(MovimientoDetalle.id_movimiento.desc()).all()

    # Configurar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Físico"

    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Cabeceras (🚨 12 columnas en total ahora)
    headers = [
        "Fecha Ingreso", "Cód. Identificador", "Nombre Producto", "Talla", "Proveedor", 
        "Categoría", "Unidad", "Stock", "Precio (S/)", "Total (S/)", 
        "Conteo Físico", "Diferencia"
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin

    suma_precio = 0
    suma_total = 0

    # Llenar datos
    for m in movimientos:
        fecha = m.entrada_rel.fecha_ingreso.strftime('%d-%m-%Y') if m.entrada_rel and m.entrada_rel.fecha_ingreso else "-"
        codigo = m.producto_rel.codigo_identificador or "-"
        nombre = m.producto_rel.nombre_prod or "-"
        
        # 🚨 EXTRAEMOS LA TALLA AQUÍ
        talla_val = m.talla if m.talla else "-"
        
        proveedor = m.entrada_rel.proveedor.razon_social if m.entrada_rel and m.entrada_rel.proveedor else "-"
        categoria = m.producto_rel.categoria.tipo_categoria or "-"
        unidad = m.producto_rel.unidad_medida or "-"
        
        stock = float(m.stock_restante or 0)
        precio = float(m.precio_unitario or m.producto_rel.precio_igv or 0)
        total = stock * precio

        # Buscar el último conteo físico
        ultimo_conteo = InventarioAuditoria.query.filter(
            InventarioAuditoria.observaciones.like(f"%Lote ID: {m.id_movimiento}%")
        ).order_by(InventarioAuditoria.id_auditoria.desc()).first()
        
        conteo_val = float(ultimo_conteo.conteo_fisico) if ultimo_conteo else ""
        diferencia = (conteo_val - stock) if conteo_val != "" else ""

        # ==============================================================
        # 🚨 NUEVO FILTRO INTELIGENTE: Búsqueda por múltiples palabras
        # ==============================================================
        if palabras_busqueda:
            fila_texto = f"{fecha} {codigo} {nombre} {talla_val} {proveedor} {categoria}".upper()
            # Validamos que TODAS las palabras buscadas existan en la fila
            # Si alguna falta, saltamos a la siguiente fila sin procesarla
            if not all(palabra in fila_texto for palabra in palabras_busqueda):
                continue 

        # Sumatorias
        suma_precio += precio
        suma_total += total
        
        # Formato de la diferencia visual
        dif_str = ""
        dif_color = "475569" # Gris
        if diferencia != "":
            dif_str = f"+{diferencia}" if diferencia >= 0 else str(diferencia)
            if diferencia > 0: dif_color = "10B981" # Verde
            elif diferencia < 0: dif_color = "EF4444" # Rojo

        # 🚨 AGREGAMOS LA TALLA EN EL ARRAY (Es el índice 3)
        row_data = [fecha, codigo, nombre, talla_val, proveedor, categoria, unidad, stock, precio, total, conteo_val, dif_str]
        ws.append(row_data)

        # Aplicar estilos a la fila recién agregada
        current_row = ws[ws.max_row]
        for idx, cell in enumerate(current_row):
            cell.border = border_thin
            
            # 🚨 Los índices se movieron. 2 es Nombre, 4 es Proveedor
            cell.alignment = align_left if idx in [2, 4] else align_center 
            
            # 🚨 Precio y Total ahora son índices 8 y 9
            if idx in [8, 9]: 
                cell.number_format = '"S/" #,##0.00'
                
            # 🚨 Diferencia ahora es índice 11
            if idx == 11 and dif_str != "":
                cell.font = Font(color=dif_color, bold=True)

    # Agregar fila de Totales Generales al final (🚨 Ahora son 12 casilleros, se movieron los totales a la derecha)
    ws.append(["", "", "", "", "", "", "", "TOTALES:", suma_precio, suma_total, "", ""])
    last_row = ws[ws.max_row]
    
    # Índice 7 es la palabra "TOTALES:"
    last_row[7].font = Font(bold=True) 
    last_row[7].alignment = Alignment(horizontal="right")
    
    # Índices 8 y 9 son los números de los totales
    for idx in [8, 9]:
        last_row[idx].font = Font(color="0369A1", bold=True)
        last_row[idx].number_format = '"S/" #,##0.00'
        last_row[idx].border = border_thin
        last_row[idx].alignment = align_center

    # Autoajuste de columnas (🚨 Agregada la letra L y ajustados todos los anchos)
    column_widths = {
        'A': 15, 'B': 15, 'C': 40, 'D': 10, 'E': 35, 'F': 20, 
        'G': 10, 'H': 10, 'I': 15, 'J': 15, 'K': 15, 'L': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Preparar archivo para descarga
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output, 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        as_attachment=True, 
        download_name="Reporte_Inventario_Fisico.xlsx"
    )

##### MODULO DE REVALIDACION DE LECTURAS #####
@app.route('/subir_matriz_csv', methods=['POST'])
def subir_matriz_csv():
    print("\n[DEBUG] === INICIANDO SUBIDA DE MATRIZ CSV ===")
    
    # 1. Validar que el archivo venga en la petición
    if 'archivo_csv' not in request.files:
        print("[DEBUG] Error: No se encontró 'archivo_csv' en request.files")
        return jsonify({'error': 'No se encontró el archivo en la petición.'}), 400
    
    file = request.files['archivo_csv']
    print(f"[DEBUG] Archivo recibido: {file.filename}")
    
    if file.filename == '':
        print("[DEBUG] Error: El nombre del archivo está vacío.")
        return jsonify({'error': 'No seleccionó ningún archivo.'}), 400
        
    if not file.filename.lower().endswith('.csv'):
        print("[DEBUG] Error: Extensión no válida.")
        return jsonify({'error': 'El formato debe ser .csv estrictamente.'}), 400

    try:
        # 2. Leer y decodificar el archivo en memoria
        print("[DEBUG] Leyendo y decodificando archivo...")
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        
        # 3. Detectar delimitador
        first_line = stream.readline()
        delimiter = ';' if ';' in first_line else ','
        print(f"[DEBUG] Delimitador detectado: '{delimiter}'")
        
        stream.seek(0) # Volver el cursor al inicio del archivo
        
        reader = csv.reader(stream, delimiter=delimiter)
        header = next(reader, None) # Saltar la primera fila (Cabeceras)
        print(f"[DEBUG] Cabeceras extraídas: {header}")
        
        registros_insertados = 0
        filas_ignoradas = 0
        
        def clean_val(val):
            return val.strip() if val and val.strip() else None

        print("[DEBUG] Iniciando procesamiento de filas...")
        
        # 4. Iterar sobre las filas e instanciar el modelo
        for indice, row in enumerate(reader, start=2): # start=2 porque la fila 1 es la cabecera
            
            # ¡CORRECCIÓN CLAVE AQUÍ! Ahora evalúa si tiene menos de 9 columnas
            if not row or len(row) < 9:
                print(f"[DEBUG - IGNORADA Fila {indice}] Longitud: {len(row) if row else 0} | Contenido: {row}")
                filas_ignoradas += 1
                continue
                
            try:
                nueva_matriz = MatrizValidacion(
                    clicodfac=clean_val(row[0]),
                    medcodygo=clean_val(row[1]),
                    lectura=clean_val(row[2]),
                    feclec=clean_val(row[3]),
                    horalec=clean_val(row[4]),
                    obs1=clean_val(row[5]),
                    obs2=clean_val(row[6]),
                    newmed=clean_val(row[7]),
                    operador=clean_val(row[8]),
                    estado='PENDIENTE',
                    fecha_subida=datetime.utcnow() - timedelta(hours=5)
                )
                db.session.add(nueva_matriz)
                registros_insertados += 1
                
            except Exception as row_err:
                print(f"[DEBUG - ERROR Fila {indice}] Falló al preparar registro. Error: {row_err}")
                
        # 5. Confirmar transacción en la base de datos
        print(f"[DEBUG] Resumen iteración -> Insertados: {registros_insertados} | Ignoradas: {filas_ignoradas}")
        db.session.commit()
        print("[DEBUG] Commit ejecutado exitosamente en la base de datos.")
        print("[DEBUG] === FIN DE SUBIDA ===\n")
        
        return jsonify({
            'success': True, 
            'mensaje': f'Se guardaron {registros_insertados} registros correctamente en la base de datos.'
        }), 200

    except Exception as e:
        db.session.rollback() # Revertir cambios si algo falla
        print(f"[ERROR DB] Error general al procesar CSV de Matriz: {e}")
        return jsonify({'error': f'Ocurrió un error interno: {str(e)}'}), 500


@app.route('/obtener_operarios_matriz', methods=['GET'])
def obtener_operarios_matriz():
    try:
        # Consultar todos los nombres de operadores únicos que no sean nulos
        operarios_db = db.session.query(MatrizValidacion.operador).filter(MatrizValidacion.operador != None).distinct().all()
        # Convertir la lista de tuplas en una lista simple de strings limpios
        lista_operarios = sorted([op[0].strip() for op in operarios_db if op[0].strip()])
        return jsonify({'success': True, 'operarios': lista_operarios})
    except Exception as e:
        print(f"[ERROR] No se pudieron obtener los operarios: {e}")
        return jsonify({'success': False, 'operarios': []})


@app.route('/obtener_lecturas', methods=['GET'])
def obtener_lecturas():
    page = request.args.get('page', 1, type=int)
    operador_filtro = request.args.get('operador', '').strip()
    fecha_filtro = request.args.get('fecha', '').strip() # Viene en formato YYYY-MM-DD
    per_page = 10

    query = MatrizValidacion.query

    # 1. Aplicar filtro de operario
    if operador_filtro and operador_filtro.upper() != 'TODOS':
        query = query.filter(MatrizValidacion.operador.ilike(f'%{operador_filtro}%'))

    # 2. Aplicar filtro de fecha
    if fecha_filtro:
        try:
            # El input type="date" envía YYYY-MM-DD, pero en el CSV/BD está como DD/MM/YYYY
            fecha_obj = datetime.strptime(fecha_filtro, '%Y-%m-%d')
            fecha_formateada = fecha_obj.strftime('%d/%m/%Y')
            # Buscamos coincidencias exactas con la fecha del CSV
            query = query.filter(MatrizValidacion.feclec == fecha_formateada)
        except ValueError:
            pass # Si la fecha viene en un formato extraño, la ignoramos

    query = query.order_by(MatrizValidacion.id_matriz.desc())
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)

    lecturas = []
    for item in paginated.items:
        lecturas.append({
            'id': item.id_matriz,
            'clicodfac': item.clicodfac or '-',
            'medcodygo': item.medcodygo or '-',
            'lectura': item.lectura or '-',
            'feclec': item.feclec or '-',
            'estado': item.estado or 'PENDIENTE'
        })

    return jsonify({
        'lecturas': lecturas,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': page,
        'per_page': per_page
    })


@app.route('/api/descargar_matriz/<int:user_id>', methods=['GET'])
def descargar_matriz(user_id):
    try:
        # 1. Buscar al Usuario y Empleado
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return jsonify({'success': False, 'message': 'Usuario no encontrado.'}), 404

        empleado = Empleado.query.get(usuario.id_empleado)
        if not empleado:
            return jsonify({'success': False, 'message': 'Empleado no asignado a este usuario.'}), 404

        # 2. Limpiar el nombre del empleado de la BD
        nombre_db = empleado.nombres.strip().upper()

        # 3. Traer TODAS las órdenes que estén PENDIENTES
        # (En lugar de filtrar en SQL, filtramos en Python con Inteligencia)
        todas_pendientes = MatrizValidacion.query.filter_by(estado='PENDIENTE').all()
        
        asignaciones_usuario = []

        # 4. MOTOR DE BÚSQUEDA DIFUSA (FUZZY MATCHING)
        for registro in todas_pendientes:
            nombre_csv = (registro.operador or "").strip().upper()
            
            # Si la columna en el CSV vino vacía, la ignoramos
            if not nombre_csv:
                continue

            # Calcula el porcentaje de coincidencia (0 a 100)
            # token_set_ratio ignora palabras extra, desordenadas y tolera typos
            similitud = fuzz.token_set_ratio(nombre_db, nombre_csv)

            # Umbral de confianza: 75% suele ser el punto dulce ideal
            if similitud >= 75:
                # Opcional: Imprimir en consola para ver cómo funciona el algoritmo
                print(f"[FUZZY MATCH] BD: '{nombre_db}' | CSV: '{nombre_csv}' | Similitud: {similitud}% -> APROBADO")
                asignaciones_usuario.append(registro)

        # 5. Si no encontró ninguna, avisamos
        if not asignaciones_usuario:
            return jsonify({'success': False, 'message': 'No se encontraron órdenes pendientes para este operador.'}), 404

        # 6. Cambiar estado a las órdenes que hicieron Match
        for registro in asignaciones_usuario:
            registro.estado = 'DESCARGADO'
            
        db.session.commit()
        print(f"[DEBUG] Se han actualizado {len(asignaciones_usuario)} registros a 'DESCARGADO'.")

        # 7. Formatear los datos para el JSON
        data = [{
            'id_matriz': a.id_matriz,
            'clicodfac': a.clicodfac or '-',
            'medcodygo': a.medcodygo or '-',
            'lectura': a.lectura or '-',
            'feclec': a.feclec or '-',
            'estado': a.estado,
            'obs1': a.obs1 or '',
            'newmed': a.newmed or ''
        } for a in asignaciones_usuario]

        return jsonify({
            'success': True,
            'data': data,
            'message': f'Se descargaron {len(data)} registros (Fuzzy Match aplicado).'
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR BACKEND] Error descargando matriz: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/actualizar_revalidacion', methods=['POST'])
def actualizar_revalidacion():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos válidos.'}), 400

        # 1. Extraer los datos enviados desde la app móvil (Retrofit)
        id_matriz = data.get('idMatriz')
        estado = data.get('estado')
        nueva_lect = data.get('nuevaLec')
        nueva_obs = data.get('nuevaObs')
        nuevo_med = data.get('nuevoMed')
        fecha_validacion = data.get('fechaHoraCalculada')

        if not id_matriz or not estado:
            return jsonify({'success': False, 'message': 'Faltan parámetros obligatorios (idMatriz o estado).'}), 400

        # 2. Buscar el registro exacto en la base de datos usando SQLAlchemy
        registro = MatrizValidacion.query.get(id_matriz)
        
        if not registro:
            return jsonify({'success': False, 'message': f'El registro {id_matriz} no existe.'}), 404

        # 3. Lógica limpia para convertir strings vacíos ("") en NULL (None en Python)
        # Esto asegura que si el técnico no modificó nada, se guarde como NULL en MySQL/SQL Server
        registro.estado = estado.upper()
        registro.nueva_lect = nueva_lect.strip() if nueva_lect and nueva_lect.strip() != "" else None
        registro.nueva_obs = nueva_obs.strip() if nueva_obs and nueva_obs.strip() != "" else None
        registro.nuevo_med = nuevo_med.strip() if nuevo_med and nuevo_med.strip() != "" else None
        registro.fecha_validacion = fecha_validacion.strip() if fecha_validacion and fecha_validacion.strip() != "" else None

        # 4. Guardar los cambios definitivamente
        db.session.commit()
        print(f"[DEBUG] Suministro {id_matriz} actualizado correctamente a estado: {estado.upper()}")

        return jsonify({
            'success': True,
            'message': f'Registro actualizado exitosamente a {estado.upper()}.'
        }), 200

    except Exception as e:
        # 5. Si algo falla, revertimos el cambio para proteger la base de datos (igual que en tu código)
        db.session.rollback()
        print(f"[ERROR BACKEND] Error actualizando revalidación: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500


@app.route('/api/operarios_por_fecha', methods=['GET'])
def get_operarios_por_fecha():
    fecha_html = request.args.get('fecha') # Viene como '2026-06-11'
    
    if not fecha_html:
        return jsonify([])

    try:
        # 1. Convertimos la fecha de 'YYYY-MM-DD' a 'DD/MM/YYYY' (el formato de tu CSV)
        fecha_obj = datetime.strptime(fecha_html, '%Y-%m-%d')
        fecha_db = fecha_obj.strftime('%d/%m/%Y') 
        
        # 2. Buscamos usando la fecha ya formateada
        operarios = db.session.query(MatrizValidacion.operador)\
            .filter(MatrizValidacion.feclec == fecha_db)\
            .distinct().all()

        # Convertir a lista de strings
        lista_operarios = [op[0] for op in operarios if op[0]]
        
        return jsonify(lista_operarios), 200

    except Exception as e:
        print(f"Error parseando fecha: {e}")
        return jsonify([])


@app.route('/api/matriz_revision', methods=['GET'])
def get_matriz_revision():
    fecha_html = request.args.get('fecha')
    operario = request.args.get('operario')
    estado_filtro = request.args.get('estado')
    page = request.args.get('page', 1, type=int) # Recibimos la página actual
    per_page = 15 # Cantidad de registros por página

    # Lógica de estado
    estado_busqueda = estado_filtro if estado_filtro else 'POR MODIFICAR'

    query = MatrizValidacion.query.filter(MatrizValidacion.estado == estado_busqueda)

    if fecha_html:
        try:
            fecha_obj = datetime.strptime(fecha_html, '%Y-%m-%d')
            fecha_db = fecha_obj.strftime('%d/%m/%Y')
            query = query.filter(MatrizValidacion.feclec == fecha_db)
        except Exception as e:
            print(f"Error parseando fecha: {e}")
    
    if operario:
        query = query.filter(MatrizValidacion.operador == operario)

    # Ordenar y PAGINAR
    query = query.order_by(MatrizValidacion.id_matriz.desc())
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)

    data = [{
        'id_matriz': r.id_matriz,
        'suministro': r.clicodfac or '-',
        'newmed': r.newmed or '-',
        'lectura_nueva': r.nueva_lect or r.lectura or '-', 
        'observacion_nueva': r.nueva_obs or r.obs1 or 'SIN OBSERVACIÓN',
        'estado': r.estado,
        'feclec': r.feclec or ''
    } for r in paginated.items] # Extraemos solo los items de la página actual

    return jsonify({
        'success': True, 
        'data': data,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': page,
        'per_page': per_page
    }), 200


@app.route('/api/cambiar_estado_revision', methods=['POST'])
def cambiar_estado_revision():
    try:
        data = request.get_json()
        id_matriz = data.get('id_matriz')
        nuevo_estado = data.get('estado') # Recibirá 'MODIFICADO' o 'RECHAZADO'

        if not id_matriz or not nuevo_estado:
            return jsonify({'success': False, 'message': 'Faltan parámetros obligatorios.'}), 400

        # Buscar el registro exacto en la tabla matriz_validacion
        registro = MatrizValidacion.query.get(id_matriz)
        if not registro:
            return jsonify({'success': False, 'message': 'El registro no existe en la base de datos.'}), 404

        # Actualizar el estado
        registro.estado = nuevo_estado.upper()
        
        # Guardar definitivamente en la BD
        db.session.commit()
        print(f"[DEBUG] Registro {id_matriz} actualizado exitosamente a: {nuevo_estado.upper()}")

        return jsonify({
            'success': True, 
            'message': f'Suministro actualizado correctamente a {nuevo_estado.upper()}.'
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Error al cambiar estado en revisión: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500


@app.route('/api/avance_validacion', methods=['GET'])
def get_avance_validacion():
    fecha_html = request.args.get('fecha')
    operario_filtro = request.args.get('operario', '').strip()

    if not fecha_html:
        return jsonify({'success': False, 'error': 'La fecha es obligatoria para calcular el avance.'}), 400

    try:
        # Transformar YYYY-MM-DD a DD/MM/YYYY para tu base de datos
        fecha_obj = datetime.strptime(fecha_html, '%Y-%m-%d')
        fecha_db = fecha_obj.strftime('%d/%m/%Y')
    except ValueError:
        return jsonify({'success': False, 'error': 'Formato de fecha inválido.'}), 400

    # Magia de SQLAlchemy: Contar total y contar solo los validados/modificados
    query = db.session.query(
        MatrizValidacion.operador,
        func.count(MatrizValidacion.id_matriz).label('total'),
        func.sum(
            case(
                (MatrizValidacion.estado.in_(['VALIDADO', 'MODIFICADO']), 1), 
                else_=0
            )
        ).label('procesados')
    ).filter(MatrizValidacion.feclec == fecha_db)

    # Si escribió un operario, filtramos. Si no, trae todos los del día.
    if operario_filtro and operario_filtro.upper() != 'TODOS':
        query = query.filter(MatrizValidacion.operador.ilike(f'%{operario_filtro}%'))

    # Agrupamos por el nombre del operario para sacar el resumen de cada uno
    query = query.group_by(MatrizValidacion.operador).order_by(MatrizValidacion.operador)

    resultados = query.all()

    datos_avance = []
    for row in resultados:
        datos_avance.append({
            'operario': row.operador or 'SIN ASIGNAR',
            'total': row.total or 0,
            'procesados': row.procesados or 0
        })

    return jsonify({'success': True, 'data': datos_avance})



# ==========================================
# 1. RUTA PARA PREVISUALIZAR EN LA WEB
# ==========================================
@app.route('/api/reportes/previsualizar', methods=['POST'])
def previsualizar_reporte():
    data = request.json
    operarios = data.get('operarios', [])
    estado = data.get('estado', 'TODOS')

    query = MatrizValidacion.query

    # Filtros
    if operarios and "TODOS" not in operarios:
        query = query.filter(MatrizValidacion.operador.in_(operarios))
    if estado and estado != "TODOS":
        query = query.filter(MatrizValidacion.estado == estado)

    # Limitamos a 50 para no colgar la web en la previsualización
    resultados = query.limit(50).all() 

    datos = [{
        "clicodfac": r.clicodfac or "-",
        "medcodygo": r.medcodygo or "-",
        "lectura": r.lectura or "-",
        "feclec": r.feclec or "-",
        "operador": r.operador or "-",
        "estado": r.estado or "-",
        "nueva_lect": r.nueva_lect or "-",
        "nuevo_med": r.nuevo_med or "-"
    } for r in resultados]

    # Contamos el total real sin el límite
    total_registros = query.count()

    return jsonify({"success": True, "data": datos, "total": total_registros})

# ==========================================
# 2. RUTA PARA GENERAR EL EXCEL PROFESIONAL
# ==========================================
@app.route('/api/reportes/descargar', methods=['POST'])
def descargar_excel():
    data = request.json
    fecha_html = data.get('fecha') # Viene como 'YYYY-MM-DD'
    operarios = data.get('operarios', [])
    estado = data.get('estado', 'TODOS')

    query = MatrizValidacion.query

    # 1. Filtro de Fecha (Reutilizando tu lógica de conversión)
    if fecha_html:
        fecha_obj = datetime.strptime(fecha_html, '%Y-%m-%d')
        fecha_db = fecha_obj.strftime('%d/%m/%Y')
        query = query.filter(MatrizValidacion.feclec == fecha_db)

    # 2. Filtro de Operarios
    if operarios and "TODOS" not in operarios:
        query = query.filter(MatrizValidacion.operador.in_(operarios))
        
    # 3. Filtro de Estado
    if estado and estado != "TODOS":
        query = query.filter(MatrizValidacion.estado == estado)

    resultados = query.all()

    # Mapeo exacto solicitado
    df_data = []
    for r in resultados:
        df_data.append({
            "SUMINISTRO": r.clicodfac,
            "MEDIDOR": r.medcodygo,
            "FECHA LECTURA": r.feclec,
            "LECTURA DIGITADA": r.lectura,
            "OBS1 DIGITADA": r.obs1,
            "MEDIDOR DIGITADO": r.newmed,
            "LECTURISTA": r.operador,
            "FECHA REVALI": r.fecha_validacion.strftime('%Y-%m-%d %H:%M') if r.fecha_validacion else "",
            "ESTADO": r.estado,
            "NUEVA LECTURA": r.nueva_lect,
            "NUEVA OBS1": r.nueva_obs,
            "NUEVO MEDIDOR": r.nuevo_med
        })

    # Crear DataFrame de Pandas
    df = pd.DataFrame(df_data)

    # Generar Excel en Memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data_Validacion')
        worksheet = writer.sheets['Data_Validacion']

        # Estilos corporativos (Cabecera Azul Oscuro, Letras Blancas)
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Aplicar estilos a la cabecera
        for col_num, cell in enumerate(worksheet[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Aplicar bordes a las filas y autoajustar el tamaño de las columnas mágicamente
        for column in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            
            for cell in column:
                cell.border = thin_border # Borde para toda la tabla
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ajuste de ancho con un poco de padding
            adjusted_width = (max_length + 3)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    output.seek(0)
    
    # Retornar el archivo virtual como descarga
    return send_file(
        output, 
        download_name="Reporte_Operativo_Techdito.xlsx", 
        as_attachment=True, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



# ==============================================================================
# MÓDULO: GESTIÓN DE CARTAS
# ==============================================================================

# Configuración de subidas (Crea la carpeta si no existe)
UPLOAD_FOLDER_CARTAS = os.path.join(app.root_path, 'static', 'uploads', 'cartas')
os.makedirs(UPLOAD_FOLDER_CARTAS, exist_ok=True)
ALLOWED_EXTENSIONS_CARTAS = {'pdf'}

def allowed_file_cartas(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_CARTAS

# ==============================================================================
# GOOGLE CLOUD STORAGE HELPERS
# ==============================================================================
# IVARGAS - 11/07/2026
def get_gcs_client():
    if storage is None:
        raise RuntimeError("google-cloud-storage no está instalado. Instala google-cloud-storage en tu entorno.")

    credentials_file = app.config.get('GCS_CREDENTIALS_FILE')
    if credentials_file:
        if not os.path.exists(credentials_file):
            raise RuntimeError(f"Archivo de credenciales GCS no encontrado: {credentials_file}")
        credentials = Credentials.from_service_account_file(credentials_file)
        return storage.Client(credentials=credentials)

    return storage.Client()


def upload_pdf_to_gcs(local_path, blob_name):
    bucket_name = app.config.get('GCS_BUCKET_NAME')
    if not bucket_name:
        raise RuntimeError('La variable GCS_BUCKET_NAME no está configurada.')

    client = get_gcs_client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    blob.upload_from_filename(local_path, content_type='application/pdf')
    return blob.name


def get_signed_url(blob_name, as_attachment=False, filename=None):
    bucket_name = app.config.get('GCS_BUCKET_NAME')
    if not bucket_name:
        raise RuntimeError('La variable GCS_BUCKET_NAME no está configurada.')

    client = get_gcs_client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    expiration = timedelta(minutes=app.config.get('GCS_SIGNED_URL_EXPIRATION', 15))

    if as_attachment:
        if not filename:
            filename = os.path.basename(blob_name)
        return blob.generate_signed_url(
            expiration=expiration,
            version='v4',
            method='GET',
            response_disposition=f'attachment; filename="{filename}"'
        )

    return blob.generate_signed_url(
        expiration=expiration,
        version='v4',
        method='GET'
    )


def delete_blob_from_gcs(blob_name):
    """Elimina un archivo del bucket GCS."""
    bucket_name = app.config.get('GCS_BUCKET_NAME')
    if not bucket_name:
        raise RuntimeError('La variable GCS_BUCKET_NAME no está configurada.')

    try:
        client = get_gcs_client()
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        blob.delete()
        print(f"[OK] Eliminado de GCS: {blob_name}")
        return True
    except Exception as e:
        print(f"[AVISO] Error al eliminar de GCS: {e}")
        return False


def carta_to_dict(carta):
    datos = carta.to_dict()
    datos['ruta_pdf'] = url_for('obtener_documento_carta', carta_id=carta.id)
    datos['ruta_descarga'] = url_for('descargar_carta', carta_id=carta.id)
    return datos


def reducir_pdf_tamano_raster(input_path, output_path, max_width=1000, image_quality=70):
    """Reduce el tamaño del PDF rasterizando cada página con PyMuPDF (fitz).

    - max_width: ancho máximo en px para las páginas renderizadas (reduce resolución si es mayor).
    - image_quality: calidad JPEG (1-100) para la compresión de las imágenes.

    Nota: Este método rasteriza las páginas (pierde texto seleccionable). Úsalo cuando
    el objetivo sea reducir peso en producción y la pérdida de búsqueda sea aceptable.
    """
    try:
        doc = fitz.open(input_path)
        new_doc = fitz.open()

        for page in doc:
            rect = page.rect
            width = rect.width
            scale = 1.0
            if width > max_width:
                scale = max_width / width

            mat = fitz.Matrix(scale, scale)
            pix = page.get_pixmap(matrix=mat, alpha=False)

            # Obtener JPEG con la calidad deseada
            jpg_bytes = pix.tobytes('jpg', quality=image_quality)

            # Crear página nueva con el tamaño del pixmap
            new_page = new_doc.new_page(width=pix.width, height=pix.height)
            new_page.insert_image(new_page.rect, stream=jpg_bytes)

        # Guardar documento resultado
        new_doc.save(output_path, garbage=4, deflate=True)
        doc.close()
        new_doc.close()
        return True
    except Exception as e:
        print(f"[AVISO] Falló compresión de PDF: {e}")
        try:
            doc.close()
        except:
            pass
        try:
            new_doc.close()
        except:
            pass
        return False
# ========================================

# ==============================================================================
# 1. ANÁLISIS DE PDF LOCAL
# ==============================================================================
@app.route('/api/cartas/analizar-pdf', methods=['POST'])
def analizar_pdf_ocr():
    if 'archivo_pdf' not in request.files:
        return jsonify({"error": "No se adjuntó archivo"}), 400
        
    file = request.files['archivo_pdf']
    if file.filename == '':
        return jsonify({"error": "Archivo inválido"}), 400

    temp_path = os.path.join(tempfile.gettempdir(), secure_filename(file.filename))
    file.save(temp_path)

    try:
        import fitz  # PyMuPDF que ya tienes instalado localmente
        import re
        
        # 1. Leer el archivo localmente
        doc = fitz.open(temp_path)
        texto_extraido = doc[0].get_text()
        doc.close()
        os.remove(temp_path)

        # 2. Si es una foto pura, el texto estará vacío.
        # Fallback instantáneo a modo manual, sin hacer esperar al usuario.
        if len(texto_extraido.strip()) < 15:
            return jsonify({
                "exito": True, 
                "datos": {"numero_carta": "", "asunto": ""}, 
                "alerta": "Documento escaneado detectado. Por favor, digite el Número y Asunto manualmente."
            }), 200

        # 3. Si el PDF es digital y tiene texto, lo extraemos con Regex
        datos_sugeridos = {"numero_carta": "", "asunto": ""}
        
        match_carta = re.search(r'CARTA\s*(?:N[°|º|.]?|NRO[.]?)?\s*([0-9A-Za-z-]+)', texto_extraido, re.IGNORECASE)
        if match_carta:
            datos_sugeridos["numero_carta"] = "Carta N° " + match_carta.group(1).strip()

        match_asunto = re.search(r'Asunto[:\s]+([^\n]+)', texto_extraido, re.IGNORECASE)
        if match_asunto:
            datos_sugeridos["asunto"] = match_asunto.group(1).strip()

        return jsonify({"exito": True, "datos": datos_sugeridos}), 200

    except Exception as e:
        if os.path.exists(temp_path): os.remove(temp_path)
        print(f"⚠️ [AVISO] Fallo en la lectura del PDF local: {e}")
        # En caso de error, abrimos el formulario para llenado manual de inmediato
        return jsonify({"exito": True, "datos": {"numero_carta": "", "asunto": ""}}), 200

# ---------------------------------------------------------
# 2. GUARDAR CARTA Y ARMAR EL HILO
# ---------------------------------------------------------

# --ivargas 
@app.route('/api/cartas/registrar', methods=['POST'])
def registrar_carta():
    try:
        numero_carta = request.form.get('numero_carta')
        asunto = request.form.get('asunto')
        tipo = request.form.get('tipo')
        fecha_str = request.form.get('fecha')
        fecha_limite_str = request.form.get('fecha_limite')
        estado_form = request.form.get('estado')
        referencia_id = request.form.get('carta_referencia_id')
        
        file = request.files.get('archivo_pdf')

        if not file or not allowed_file_cartas(file.filename):
            return jsonify({"error": "Debe adjuntar un documento PDF válido"}), 400

        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else None
        fecha_limite_obj = datetime.strptime(fecha_limite_str, '%Y-%m-%d').date() if fecha_limite_str else None

        filename = secure_filename(file.filename)
        nombre_unico = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"

        # IVARGAS - 11/07/2026
        # ========================================
        temp_filepath = os.path.join(tempfile.gettempdir(), nombre_unico)
        file.save(temp_filepath)

        # Intentar comprimir el PDF antes de enviarlo a GCS
        compressed_path = os.path.join(tempfile.gettempdir(), f"compressed_{nombre_unico}")
        try:
            comprimido = reducir_pdf_tamano_raster(temp_filepath, compressed_path, max_width=1000, image_quality=70)
        except Exception as e:
            print(f"[AVISO] Error al intentar comprimir: {e}")
            comprimido = False

        upload_target = compressed_path if comprimido and os.path.exists(compressed_path) else temp_filepath

        blob_name = f"cartas/{nombre_unico}"
        upload_pdf_to_gcs(upload_target, blob_name)

        # Limpiar temporales
        try:
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
        except Exception:
            pass
        try:
            if upload_target != temp_filepath and os.path.exists(upload_target):
                os.remove(upload_target)
        except Exception:
            pass

        ruta_relativa_bd = blob_name
        # ========================================

        nueva_carta = Carta(
            numero_carta=numero_carta,
            asunto=asunto,
            tipo=tipo,
            fecha_emision=fecha_obj if tipo == 'EMITIDA' else None,
            fecha_recepcion=fecha_obj if tipo == 'RECIBIDA' else None,
            fecha_limite=fecha_limite_obj,
            ruta_pdf=ruta_relativa_bd,
            estado=estado_form
        )
        db.session.add(nueva_carta)
        db.session.flush()

        if referencia_id:
            carta_origen = Carta.query.get(referencia_id)
            if carta_origen:
                nueva_carta.referencias_pasadas.append(carta_origen)
                if tipo == 'EMITIDA' and carta_origen.tipo == 'RECIBIDA' and carta_origen.estado == 'PENDIENTE':
                    carta_origen.estado = 'ATENDIDA'

        db.session.commit()
        return jsonify({"exito": True, "mensaje": "Documento registrado correctamente"}), 201

    except Exception as e:
        db.session.rollback()
        
        # IVARGAS - 11/07/2026
        # ========================================
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
            os.remove(temp_filepath)
        # ========================================

        print(f"Error Guardado: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Fallo al guardar: {str(e)}"}), 500

# ==============================================================================
# LISTAR CARTAS
# ==============================================================================
@app.route('/api/cartas/listar', methods=['GET'])
def listar_cartas():
    try:
        # 1. Atrapamos los parámetros que envía JavaScript
        page = request.args.get('page', 1, type=int)
        search = request.args.get('search', '').strip()
        tipo = request.args.get('tipo', '').strip()
        estado = request.args.get('estado', '').strip()

        # 2. Iniciamos la consulta base
        query = Carta.query

        # 3. Aplicamos los filtros dinámicamente si es que existen
        if search:
            # Filtra si el texto coincide con el Número de carta O con el Asunto (Ignora mayúsculas/minúsculas)
            query = query.filter(db.or_(
                Carta.numero_carta.ilike(f'%{search}%'),
                Carta.asunto.ilike(f'%{search}%')
            ))
        
        if tipo:
            query = query.filter(Carta.tipo == tipo)
            
        if estado:
            query = query.filter(Carta.estado == estado)

        # 4. Ordenamos por las más recientes primero y paginamos
        paginacion = query.order_by(Carta.id.desc()).paginate(page=page, per_page=10, error_out=False)

        # IVARGAS - 11/07/2026
        # ========================================
        datos = [carta_to_dict(carta) for carta in paginacion.items]
        # ========================================

        meta = {
            "total_items": paginacion.total,
            "total_pages": paginacion.pages,
            "current_page": paginacion.page,
            "has_next": paginacion.has_next,
            "has_prev": paginacion.has_prev
        }

        return jsonify({"exito": True, "datos": datos, "meta": meta}), 200

    except Exception as e:
        print(f"Error al listar cartas: {e}")
        return jsonify({"error": str(e)}), 500
    
# ==============================================================================
# OBTENER CARTAS PARA EL BUSCADOR DEL MODAL
# ==============================================================================
@app.route('/api/cartas/todas-basico', methods=['GET'])
def listar_cartas_basico():
    try:
        # Traemos todas las cartas (solo los campos necesarios para el buscador)
        cartas = Carta.query.order_by(Carta.id.desc()).all()
        datos = [{"id": c.id, "numero_carta": c.numero_carta, "asunto": c.asunto} for c in cartas]
        return jsonify({"exito": True, "datos": datos}), 200
    except Exception as e:
        print(f"Error al listar cartas para el buscador: {e}")
        return jsonify({"error": str(e)}), 500


# IVARGAS - 11/07/2026
# ========================================
@app.route('/api/cartas/documento/<int:carta_id>', methods=['GET'])
def obtener_documento_carta(carta_id):
    try:
        carta = Carta.query.get_or_404(carta_id)
        if not carta.ruta_pdf:
            return jsonify({"error": "No se encontró un documento asociado a esta carta."}), 404

        signed_url = get_signed_url(carta.ruta_pdf)
        return jsonify({"exito": True, "url": signed_url}), 200

    except Exception as e:
        print(f"Error generando URL de documento GCS: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/cartas/descargar/<int:carta_id>', methods=['GET'])
def descargar_carta(carta_id):
    try:
        carta = Carta.query.get_or_404(carta_id)
        if not carta.ruta_pdf:
            return jsonify({"error": "No se encontró un documento asociado a esta carta."}), 404

        filename = os.path.basename(carta.ruta_pdf)
        signed_url = get_signed_url(carta.ruta_pdf, as_attachment=True, filename=filename)
        return redirect(signed_url)

    except Exception as e:
        print(f"Error generando descarga GCS: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/cartas/eliminar/<int:carta_id>', methods=['DELETE'])
def eliminar_carta(carta_id):
    try:
        carta = Carta.query.get_or_404(carta_id)
        ruta_pdf = carta.ruta_pdf

        # 1. Eliminar de GCS
        if ruta_pdf:
            delete_blob_from_gcs(ruta_pdf)

        # 2. Eliminar referencias (lazy=True para no cargar todas de una vez)
        carta.referencias_pasadas.clear()
        # Si hay referencias futuras, también limpiarlas
        for hijo in Carta.query.filter(Carta.referencias_pasadas.any(Carta.id == carta.id)).all():
            hijo.referencias_pasadas.remove(carta)

        # 3. Eliminar de BD
        db.session.delete(carta)
        db.session.commit()

        return jsonify({"exito": True, "mensaje": "Carta eliminada correctamente"}), 200

    except Exception as e:
        db.session.rollback()
        print(f"Error eliminando carta: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/cartas/actualizar/<int:carta_id>', methods=['PUT'])
def actualizar_carta(carta_id):
    try:
        carta = Carta.query.get_or_404(carta_id)

        # Capturar campos a actualizar (pueden ser parciales)
        numero_carta = request.form.get('numero_carta')
        asunto = request.form.get('asunto')
        tipo = request.form.get('tipo')
        fecha_str = request.form.get('fecha')
        fecha_limite_str = request.form.get('fecha_limite')
        estado_form = request.form.get('estado')
        file = request.files.get('archivo_pdf')

        # Actualizar campos simples
        if numero_carta:
            carta.numero_carta = numero_carta
        if asunto:
            carta.asunto = asunto.upper()
        if tipo in ['EMITIDA', 'RECIBIDA']:
            carta.tipo = tipo
        if estado_form in ['PENDIENTE', 'ATENDIDA', 'ARCHIVADA']:
            carta.estado = estado_form

        if fecha_str:
            try:
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                if tipo == 'EMITIDA':
                    carta.fecha_emision = fecha_obj
                elif tipo == 'RECIBIDA':
                    carta.fecha_recepcion = fecha_obj
            except ValueError:
                pass

        if fecha_limite_str:
            try:
                fecha_limite_obj = datetime.strptime(fecha_limite_str, '%Y-%m-%d').date()
                carta.fecha_limite = fecha_limite_obj
            except ValueError:
                pass

        # Si se envía un nuevo archivo, reemplazar
        if file and allowed_file_cartas(file.filename):
            # Eliminar antiguo de GCS
            if carta.ruta_pdf:
                delete_blob_from_gcs(carta.ruta_pdf)

            # Subir nuevo
            filename = secure_filename(file.filename)
            nombre_unico = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
            temp_filepath = os.path.join(tempfile.gettempdir(), nombre_unico)
            file.save(temp_filepath)

            # Comprimir
            compressed_path = os.path.join(tempfile.gettempdir(), f"compressed_{nombre_unico}")
            try:
                comprimido = reducir_pdf_tamano_raster(temp_filepath, compressed_path, max_width=1000, image_quality=70)
            except Exception as e:
                print(f"[AVISO] Error al intentar comprimir: {e}")
                comprimido = False

            upload_target = compressed_path if comprimido and os.path.exists(compressed_path) else temp_filepath
            blob_name = f"cartas/{nombre_unico}"
            upload_pdf_to_gcs(upload_target, blob_name)

            # Limpiar temporales
            try:
                if os.path.exists(temp_filepath):
                    os.remove(temp_filepath)
            except Exception:
                pass
            try:
                if upload_target != temp_filepath and os.path.exists(upload_target):
                    os.remove(upload_target)
            except Exception:
                pass

            carta.ruta_pdf = blob_name

        db.session.commit()
        return jsonify({"exito": True, "mensaje": "Carta actualizada correctamente"}), 200

    except Exception as e:
        db.session.rollback()
        print(f"Error actualizando carta: {e}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
# ========================================

# ==============================================================================
# OBTENER EL HILO COMPLETO DE UNA CARTA (EXPEDIENTE)
# ==============================================================================
@app.route('/api/cartas/hilo/<int:carta_id>', methods=['GET'])
def obtener_hilo_carta(carta_id):
    try:
        # Encontramos la carta solicitada
        carta_actual = Carta.query.get_or_404(carta_id)
        
        hilo_completo = []
        
        # 1. Función para subir hacia el origen (padres)
        def buscar_padres(carta):
            # Si tu modelo usa "referencias_pasadas" como colección ManyToMany:
            if carta.referencias_pasadas:
                for padre in carta.referencias_pasadas:
                    if padre not in hilo_completo:
                        hilo_completo.append(padre)
                        buscar_padres(padre)

        # 2. Función para bajar hacia el futuro (hijos)
        # Esto busca qué cartas tienen a ESTA carta como su referencia pasada.
        def buscar_hijos(carta):
            # En SQLAlchemy, si definiste un backref (ej. 'respuestas'), úsalo.
            # Si no, podemos consultar la BD:
            todas_las_cartas = Carta.query.all()
            for posible_hijo in todas_las_cartas:
                if carta in posible_hijo.referencias_pasadas:
                    if posible_hijo not in hilo_completo:
                        hilo_completo.append(posible_hijo)
                        buscar_hijos(posible_hijo)

        # Agregamos la carta que buscó el usuario
        hilo_completo.append(carta_actual)
        
        # Poblamos la lista hacia atrás y hacia adelante
        buscar_padres(carta_actual)
        buscar_hijos(carta_actual)

        # Ordenamos del más reciente al más antiguo basados en la fecha del documento
        hilo_completo = sorted(
            hilo_completo, 
            key=lambda x: x.fecha_emision if x.tipo == 'EMITIDA' else x.fecha_recepcion, 
            reverse=True
        )

        # Retornamos los datos limpios para el Frontend
        # IVARGAS - 11/07/2026
        # =====================================
        datos = [carta_to_dict(c) for c in hilo_completo]
        # =====================================

        return jsonify({"exito": True, "datos": datos}), 200

    except Exception as e:
        print(f"Error armando el hilo: {e}")
        return jsonify({"error": str(e)}), 500


